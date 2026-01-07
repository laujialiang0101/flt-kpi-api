"""
FLT KPI Tracker API
FastAPI backend using pre-aggregated materialized views for fast queries.
"""
from fastapi import FastAPI, Query, HTTPException, UploadFile, File, Depends
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from starlette.middleware.gzip import GZipMiddleware
from contextlib import asynccontextmanager
from datetime import date, datetime, timedelta
from typing import Optional, List
from pydantic import BaseModel
import asyncpg
import asyncio
import os
import io
import secrets
import json
import random

# Optional imports for Excel handling
try:
    import openpyxl
    from openpyxl import Workbook
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

# Optional imports for Push Notifications
try:
    from pywebpush import webpush, WebPushException
    PUSH_AVAILABLE = True
except ImportError:
    PUSH_AVAILABLE = False
    print("Warning: pywebpush not installed. Push notifications disabled.")

# Optional imports for Password Hashing
try:
    import bcrypt
    BCRYPT_AVAILABLE = True
except ImportError:
    BCRYPT_AVAILABLE = False
    print("Warning: bcrypt not installed. Using hashlib fallback.")
    import hashlib

# VAPID Configuration for Web Push
# Generate new keys: python -c "from pywebpush import generate_vapid_keys; keys=generate_vapid_keys(); print(keys)"
VAPID_PRIVATE_KEY = os.getenv('VAPID_PRIVATE_KEY', 'Yt3wLOw0I2VT0pr-7abhp9MqklTv2dUef9bIRiFcGQY')
VAPID_PUBLIC_KEY = os.getenv('VAPID_PUBLIC_KEY', 'BNEaN-5grwfKBkK2JPQCCFnLIgW8CSUs0LU2dI5rGlJGzauTBinEfYnf0wOLKTmIqgBnfN1N9W7F1dq_7K-5hHc')
VAPID_CLAIMS = {"sub": "mailto:admin@farmasilautan.com"}

# KPI Category Mapping - AcStockUDGroup1ID values in database
# Maps display names to actual database values
KPI_CATEGORIES = {
    'HOUSE_BRAND': 'FLTHB',
    'FOCUSED_1': 'FLTF1',
    'FOCUSED_2': 'FLTF2',
    'FOCUSED_3': 'FLTF3',
    'STOCK_CLEARANCE': 'STOCK CLEARANCE',
    'PWP': 'PWP',  # PWP uses promotion system, not UD1
}


# ============================================================================
# Pydantic Models
# ============================================================================

class LoginRequest(BaseModel):
    code: str
    password: str


class SetPasswordRequest(BaseModel):
    code: str
    new_password: str


class LoginResponse(BaseModel):
    success: bool
    user: Optional[dict] = None
    token: Optional[str] = None
    error: Optional[str] = None
    needs_password_setup: bool = False


class TargetUploadRow(BaseModel):
    staff_id: str
    year_month: int  # YYYYMM format
    total_sales: float = 0
    house_brand: float = 0
    focused_1: float = 0
    focused_2: float = 0
    focused_3: float = 0
    pwp: float = 0
    clearance: float = 0
    transactions: int = 0


# ============================================================================
# Role Mapping
# ============================================================================

ROLE_MAPPING = {
    'ADMINISTRATORS': 'admin',
    'COO': 'admin',  # Chief Operating Officer - full admin access
    'CMO': 'admin',  # Chief Marketing Officer - full admin access
    'SUPERVISOR': 'supervisor',
    'PIC OUTLET': 'pic',
    'PIC': 'pic',
    'AREA_MANAGER': 'area_manager',
    'AREA MANAGER': 'area_manager',
    'OPERATIONS_MANAGER': 'operations_manager',
    'OPERATIONS MANAGER': 'operations_manager',
    'OOM': 'operations_manager',  # OOM = Outlet Operations Manager
    'CASHIER': 'staff',
    'PRICE CHECKER': 'staff',
}

ROLE_PERMISSIONS = {
    'admin': {
        'can_view_own_kpi': True,
        'can_view_leaderboard': True,
        'can_submit_audit': True,
        'can_upload_targets': True,
        'can_view_all_staff': True,
        'can_manage_roles': True
    },
    'operations_manager': {
        'can_view_own_kpi': True,
        'can_view_leaderboard': True,
        'can_submit_audit': True,
        'can_upload_targets': True,
        'can_view_all_staff': True,
        'can_manage_roles': False
    },
    'area_manager': {
        'can_view_own_kpi': True,
        'can_view_leaderboard': True,
        'can_submit_audit': True,
        'can_upload_targets': False,
        'can_view_all_staff': True,
        'can_manage_roles': False
    },
    'supervisor': {
        'can_view_own_kpi': True,
        'can_view_leaderboard': True,
        'can_submit_audit': True,
        'can_upload_targets': False,
        'can_view_all_staff': True,
        'can_manage_roles': False
    },
    'pic': {
        'can_view_own_kpi': True,
        'can_view_leaderboard': True,
        'can_submit_audit': False,
        'can_upload_targets': False,
        'can_view_all_staff': True,
        'can_manage_roles': False
    },
    'staff': {
        'can_view_own_kpi': True,
        'can_view_leaderboard': True,
        'can_submit_audit': False,
        'can_upload_targets': False,
        'can_view_all_staff': False,
        'can_manage_roles': False
    }
}

# In-memory session store (for production, use Redis)
sessions = {}

# Database configuration
# Internal hostname = DB ID with -a suffix, NO domain - only works within Render network
# External hostname = full domain with -a suffix - works from anywhere
INTERNAL_HOST = 'dpg-d4pr99je5dus73eb5730-a'  # From Render dashboard internal URL
EXTERNAL_HOST = 'dpg-d4pr99je5dus73eb5730-a.singapore-postgres.render.com'
DB_PORT = int(os.getenv('DB_PORT', 5432))
DB_NAME = os.getenv('DB_NAME', 'flt_sales_commission_db')
DB_USER = os.getenv('DB_USER', 'flt_sales_commission_db_user')
DB_PASSWORD = os.getenv('DB_PASSWORD', 'Wy0ZP1wjLPsIta0YLpYLeRWgdITbya2m')

# Global connection pool
pool: asyncpg.Pool = None
connected_host: str = None  # Track which host we're connected to (description for display)


async def create_pool_with_retry():
    """Create connection pool - try internal first (faster), then external."""
    global connected_host
    import sys

    # Try internal first (faster, private network) - no SSL needed for internal
    print(f"Trying INTERNAL host (no SSL): {INTERNAL_HOST}", flush=True)
    for attempt in range(2):
        try:
            print(f"  Attempt {attempt + 1}/2...", flush=True)
            created_pool = await asyncpg.create_pool(
                host=INTERNAL_HOST,
                port=DB_PORT,
                database=DB_NAME,
                user=DB_USER,
                password=DB_PASSWORD,
                ssl=False,  # Internal network doesn't need SSL
                min_size=1,
                max_size=10,
                command_timeout=60,
            )
            async with created_pool.acquire() as conn:
                await conn.fetchval("SELECT 1")
            print(f"SUCCESS with internal host (no SSL)!", flush=True)
            connected_host = f"{INTERNAL_HOST} (internal, no SSL)"
            return created_pool
        except Exception as e:
            print(f"  Failed: {type(e).__name__}: {e}", flush=True)
            if attempt < 1:
                await asyncio.sleep(1)

    # Try internal with SSL (in case it's required)
    print(f"Trying INTERNAL host (with SSL): {INTERNAL_HOST}", flush=True)
    for attempt in range(2):
        try:
            print(f"  Attempt {attempt + 1}/2...", flush=True)
            created_pool = await asyncpg.create_pool(
                host=INTERNAL_HOST,
                port=DB_PORT,
                database=DB_NAME,
                user=DB_USER,
                password=DB_PASSWORD,
                ssl='require',
                min_size=1,
                max_size=10,
                command_timeout=60,
            )
            async with created_pool.acquire() as conn:
                await conn.fetchval("SELECT 1")
            print(f"SUCCESS with internal host (SSL)!", flush=True)
            connected_host = f"{INTERNAL_HOST} (internal, SSL)"
            return created_pool
        except Exception as e:
            print(f"  Failed: {type(e).__name__}: {e}", flush=True)
            if attempt < 1:
                await asyncio.sleep(1)

    # Fallback to external (always works, but slower)
    print(f"Trying EXTERNAL host (SSL): {EXTERNAL_HOST}", flush=True)
    for attempt in range(3):
        try:
            print(f"  Attempt {attempt + 1}/3...", flush=True)
            created_pool = await asyncpg.create_pool(
                host=EXTERNAL_HOST,
                port=DB_PORT,
                database=DB_NAME,
                user=DB_USER,
                password=DB_PASSWORD,
                ssl='require',
                min_size=1,
                max_size=10,
                command_timeout=60,
            )
            async with created_pool.acquire() as conn:
                await conn.fetchval("SELECT 1")
            print(f"SUCCESS with external host (SSL)!", flush=True)
            connected_host = f"{EXTERNAL_HOST} (external, SSL)"
            return created_pool
        except Exception as e:
            print(f"  Failed: {type(e).__name__}: {e}", flush=True)
            if attempt < 2:
                await asyncio.sleep(2)

    raise Exception("All connection attempts failed")


@asynccontextmanager
async def lifespan(app: FastAPI):
    global pool
    pool = await create_pool_with_retry()

    # Create kpi_user_auth table if not exists
    async with pool.acquire() as conn:
        await conn.execute("""
            CREATE TABLE IF NOT EXISTS kpi_user_auth (
                code VARCHAR(50) PRIMARY KEY,
                password_hash VARCHAR(255) NOT NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)

        # Ensure OutletTargets has gross_profit_target column (migration for existing tables)
        await conn.execute("""
            DO $$
            BEGIN
                IF EXISTS (
                    SELECT 1 FROM information_schema.tables
                    WHERE table_name = 'OutletTargets'
                ) AND NOT EXISTS (
                    SELECT 1 FROM information_schema.columns
                    WHERE table_name = 'OutletTargets' AND column_name = 'gross_profit_target'
                ) THEN
                    ALTER TABLE "OutletTargets" ADD COLUMN gross_profit_target DECIMAL(15,2) DEFAULT 0;
                END IF;
            END $$;
        """)

    yield
    await pool.close()


# ============================================================================
# Password Hashing Utilities
# ============================================================================

def hash_password(password: str) -> str:
    """Hash a password using bcrypt or fallback to SHA256."""
    if BCRYPT_AVAILABLE:
        salt = bcrypt.gensalt()
        return bcrypt.hashpw(password.encode('utf-8'), salt).decode('utf-8')
    else:
        # Fallback to SHA256 with a simple salt
        salt = secrets.token_hex(16)
        hashed = hashlib.sha256((salt + password).encode()).hexdigest()
        return f"sha256${salt}${hashed}"


def verify_password(password: str, password_hash: str) -> bool:
    """Verify a password against its hash."""
    if BCRYPT_AVAILABLE and not password_hash.startswith('sha256$'):
        try:
            return bcrypt.checkpw(password.encode('utf-8'), password_hash.encode('utf-8'))
        except:
            return False
    elif password_hash.startswith('sha256$'):
        # Fallback SHA256 verification
        parts = password_hash.split('$')
        if len(parts) != 3:
            return False
        salt = parts[1]
        stored_hash = parts[2]
        computed_hash = hashlib.sha256((salt + password).encode()).hexdigest()
        return computed_hash == stored_hash
    return False


app = FastAPI(
    title="FLT KPI Tracker API",
    description="API for Farmasi Lautan Staff KPI Tracking",
    version="1.0.0",
    lifespan=lifespan
)

# CORS for frontend
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# GZip compression for responses
app.add_middleware(GZipMiddleware, minimum_size=500)


# ============================================================================
# Helper Functions
# ============================================================================

def get_role_from_group(user_group: str, is_supervisor: bool = False) -> str:
    """Map AcPOSUserGroupID to application role.
    
    Role is determined SOLELY by AcPOSUserGroupID (POS user group).
    The IsSupervisor checkbox is NOT used for role determination.
    """
    group = (user_group or '').upper().strip()
    role = ROLE_MAPPING.get(group, 'staff')
    # Note: is_supervisor parameter kept for backwards compatibility but NOT used
    # Role should be determined by POS user group only
    return role


async def get_current_user(token: str = Query(None, alias="token")):
    """Validate session token and return user info."""
    if not token or token not in sessions:
        raise HTTPException(status_code=401, detail="Invalid or expired session")
    session = sessions[token]
    # Check expiry
    if datetime.now() > session['expires_at']:
        del sessions[token]
        raise HTTPException(status_code=401, detail="Session expired")
    return session['user']


# ============================================================================
# Password Encoding Functions (Dynamod XOR-based)
# ============================================================================

def encode_password_dynamod(plain_password: str, user_code: str) -> str:
    """
    Encode password using Dynamod's XOR-based algorithm.
    The XOR key for each position is derived from the user code characters.
    """
    if not plain_password or not user_code:
        return plain_password

    encoded = []
    user_code_upper = user_code.upper()

    for i, char in enumerate(plain_password):
        code_char = user_code_upper[i % len(user_code_upper)]
        xor_key = ord(code_char) - 70
        encoded_char = chr(ord(char) ^ xor_key)
        encoded.append(encoded_char)

    return ''.join(encoded)


def check_password_dynamod(plain_password: str, stored_password: str, user_code: str) -> bool:
    """
    Check if password matches using multiple encoding methods.
    Handles Dynamod POS XOR-based password encoding.

    Password encoding varies by user:
    - Some passwords are stored in plain text (no encoding)
    - Some are XOR encoded with patterns derived from user code
    - For IC users (12-digit numeric codes), password = last 4 digits
    """
    if not stored_password:
        return False

    # Method 1: Direct comparison (if stored password is plain text)
    if plain_password == stored_password:
        return True

    # Method 2: For IC users (12-digit numeric code), password should be last 4 digits
    # We verify by checking if the entered password matches expected and stored is encoded version
    if len(user_code) == 12 and user_code.isdigit():
        expected_password = user_code[-4:]
        if plain_password == expected_password:
            # Verify the stored password length matches (4 digits)
            if len(stored_password) == 4:
                # For IC users, we trust that if they enter the correct expected password
                # (last 4 digits of IC), they are authorized
                return True
        # For IC users, ONLY accept the expected password (last 4 digits)
        # Don't allow other passwords even if they would pass other methods
        return False

    # Method 4: Try fixed XOR key patterns discovered from known passwords
    fixed_patterns = [
        [2, 8, 5, 9, 2, 8],   # Pattern from LTK (506050 -> 383938)
        [4, 1, 5, 14],        # Pattern from ID 30 (9202 -> 535<)
        [6, 3, 4, 2, 3, 4],   # Common 6-char pattern
        [6, 3, 4],            # Short repeating pattern
        [15, 13, 13, 15],     # Pattern from IC 980101115197
    ]
    for pattern in fixed_patterns:
        try:
            encoded = []
            for i, char in enumerate(plain_password):
                xor_key = pattern[i % len(pattern)]
                encoded.append(chr(ord(char) ^ xor_key))
            if ''.join(encoded) == stored_password:
                return True
        except:
            pass

    # Method 5: Try various XOR key offsets based on user code
    for offset in range(65, 80):
        try:
            encoded = []
            user_code_upper = user_code.upper()
            for i, char in enumerate(plain_password):
                code_char = user_code_upper[i % len(user_code_upper)]
                xor_key = ord(code_char) - offset
                if xor_key < 0:
                    xor_key = abs(xor_key)
                encoded.append(chr(ord(char) ^ xor_key))
            if ''.join(encoded) == stored_password:
                return True
        except:
            pass

    return False


# ============================================================================
# Authentication Endpoints
# ============================================================================

@app.post("/api/v1/auth/login")
async def login(request: LoginRequest):
    """Authenticate user with KPI Tracker credentials.

    Uses kpi.staff_list_master for all user info (role, permissions, outlets).
    Uses kpi_user_auth for password verification.
    First-time users (no KPI password set) will be prompted to set one.

    OPTIMIZED: Single query to staff_list_master instead of 5+ queries to raw tables.
    """
    try:
        async with pool.acquire() as conn:
            # Query unified staff table for all user info
            staff = await conn.fetchrow("""
                SELECT
                    staff_id, staff_name, role, pos_user_group, is_supervisor,
                    can_view_own_kpi, can_view_leaderboard, can_submit_audit,
                    can_upload_targets, can_view_all_staff, can_manage_roles,
                    primary_outlet, primary_outlet_name,
                    allowed_outlets, allowed_outlet_names,
                    is_active, region, area_manager_id, area_manager_name
                FROM kpi.staff_list_master
                WHERE UPPER(staff_id) = UPPER($1)
                  AND is_active = true
            """, request.code)

            if not staff:
                return {"success": False, "error": "Invalid credentials or inactive account"}

            # Check if user has set a KPI Tracker password
            kpi_auth = await conn.fetchrow("""
                SELECT password_hash FROM kpi_user_auth WHERE UPPER(code) = UPPER($1)
            """, request.code)

            if not kpi_auth:
                # First-time login - user needs to set password
                return {
                    "success": False,
                    "needs_password_setup": True,
                    "user": {"code": staff['staff_id'], "name": staff['staff_name']},
                    "error": "First-time login. Please set your KPI Tracker password."
                }

            # Verify password against kpi_user_auth
            if not verify_password(request.password, kpi_auth['password_hash']):
                return {"success": False, "error": "Invalid credentials"}

            # Build allowed_outlets list from pre-computed arrays
            allowed_outlets_ids = staff['allowed_outlets'] or []
            allowed_outlets_names = staff['allowed_outlet_names'] or []
            allowed_outlets = [
                {'id': outlet_id, 'name': allowed_outlets_names[i] if i < len(allowed_outlets_names) else outlet_id}
                for i, outlet_id in enumerate(allowed_outlets_ids)
            ]

            # Determine outlet_id based on role
            role = staff['role']
            if role in ['admin', 'operations_manager', 'area_manager']:
                outlet_id = None  # Multi-outlet access
                group_id = None
            else:
                outlet_id = staff['primary_outlet']
                group_id = staff['primary_outlet']

            # Build permissions from pre-computed columns
            permissions = {
                'can_view_own_kpi': staff['can_view_own_kpi'],
                'can_view_leaderboard': staff['can_view_leaderboard'],
                'can_submit_audit': staff['can_submit_audit'],
                'can_upload_targets': staff['can_upload_targets'],
                'can_view_all_staff': staff['can_view_all_staff'],
                'can_manage_roles': staff['can_manage_roles']
            }

            # Generate session token
            token = secrets.token_urlsafe(32)

            # Store session (expires in 24 hours)
            user_data = {
                'code': staff['staff_id'],
                'name': staff['staff_name'],
                'role': role,
                'outlet_id': outlet_id,
                'group_id': group_id,
                'allowed_outlets': allowed_outlets,
                'is_supervisor': staff['is_supervisor'],
                'user_group': staff['pos_user_group'],
                'permissions': permissions,
                'region': staff['region'],
                'area_manager_id': staff['area_manager_id'],
                'area_manager_name': staff['area_manager_name']
            }

            sessions[token] = {
                'user': user_data,
                'expires_at': datetime.now() + timedelta(hours=24)
            }

            return {
                "success": True,
                "user": user_data,
                "token": token
            }

    except Exception as e:
        return {"success": False, "error": f"Login failed: {str(e)}"}


@app.get("/api/v1/auth/session")
async def get_session(token: str = Query(..., description="Session token")):
    """Validate session and return user info."""
    if token not in sessions:
        return {"success": False, "error": "Invalid session"}

    session = sessions[token]
    if datetime.now() > session['expires_at']:
        del sessions[token]
        return {"success": False, "error": "Session expired"}

    return {"success": True, "user": session['user']}


@app.post("/api/v1/auth/logout")
async def logout(token: str = Query(..., description="Session token")):
    """Logout and invalidate session."""
    if token in sessions:
        del sessions[token]
    return {"success": True, "message": "Logged out successfully"}


@app.post("/api/v1/auth/refresh-session")
async def refresh_session(token: str = Query(..., description="Session token")):
    """Refresh session data from database.

    Use this to update user's role/permissions without requiring logout/login.
    Useful when user's role has been changed in the database.

    OPTIMIZED: Single query to kpi.staff_list_master instead of multiple queries.
    """
    if token not in sessions:
        return {"success": False, "error": "Invalid session"}

    session = sessions[token]
    if datetime.now() > session['expires_at']:
        del sessions[token]
        return {"success": False, "error": "Session expired"}

    old_user = session['user']
    user_code = old_user['code']

    try:
        async with pool.acquire() as conn:
            # Query unified staff table for all user info
            staff = await conn.fetchrow("""
                SELECT
                    staff_id, staff_name, role, pos_user_group, is_supervisor,
                    can_view_own_kpi, can_view_leaderboard, can_submit_audit,
                    can_upload_targets, can_view_all_staff, can_manage_roles,
                    primary_outlet, primary_outlet_name,
                    allowed_outlets, allowed_outlet_names,
                    is_active, region, area_manager_id, area_manager_name
                FROM kpi.staff_list_master
                WHERE UPPER(staff_id) = UPPER($1)
                  AND is_active = true
            """, user_code)

            if not staff:
                del sessions[token]
                return {"success": False, "error": "User no longer active"}

            # Build allowed_outlets list from pre-computed arrays
            allowed_outlets_ids = staff['allowed_outlets'] or []
            allowed_outlets_names = staff['allowed_outlet_names'] or []
            allowed_outlets = [
                {'id': outlet_id, 'name': allowed_outlets_names[i] if i < len(allowed_outlets_names) else outlet_id}
                for i, outlet_id in enumerate(allowed_outlets_ids)
            ]

            # Determine outlet_id based on role
            role = staff['role']
            if role in ['admin', 'operations_manager', 'area_manager']:
                outlet_id = None
                group_id = None
            else:
                outlet_id = staff['primary_outlet']
                group_id = staff['primary_outlet']

            # Build permissions from pre-computed columns
            permissions = {
                'can_view_own_kpi': staff['can_view_own_kpi'],
                'can_view_leaderboard': staff['can_view_leaderboard'],
                'can_submit_audit': staff['can_submit_audit'],
                'can_upload_targets': staff['can_upload_targets'],
                'can_view_all_staff': staff['can_view_all_staff'],
                'can_manage_roles': staff['can_manage_roles']
            }

            # Update session with fresh data
            updated_user = {
                'code': staff['staff_id'],
                'name': staff['staff_name'],
                'role': role,
                'outlet_id': outlet_id,
                'group_id': group_id,
                'allowed_outlets': allowed_outlets,
                'is_supervisor': staff['is_supervisor'],
                'user_group': staff['pos_user_group'],
                'permissions': permissions,
                'region': staff['region'],
                'area_manager_id': staff['area_manager_id'],
                'area_manager_name': staff['area_manager_name']
            }

            sessions[token]['user'] = updated_user

            # Log the change if role changed
            old_role = old_user.get('role', 'unknown')
            if old_role != role:
                print(f"[SESSION REFRESH] {staff['staff_name']} ({user_code}): role changed from '{old_role}' to '{role}'", flush=True)

            return {
                "success": True,
                "message": "Session refreshed successfully",
                "user": updated_user,
                "role_changed": old_role != role,
                "old_role": old_role,
                "new_role": role
            }

    except Exception as e:
        return {"success": False, "error": f"Failed to refresh session: {str(e)}"}


@app.post("/api/v1/auth/set-password")
async def set_password(request: SetPasswordRequest):
    """Set KPI Tracker password for first-time login.

    Only works for users who exist in AcPersonal but haven't set a KPI password yet.
    """
    try:
        # Validate password requirements
        if len(request.new_password) < 4:
            return {"success": False, "error": "Password must be at least 4 characters"}

        async with pool.acquire() as conn:
            # Verify user exists in AcPersonal
            user = await conn.fetchrow("""
                SELECT "Code" as code, "Name" as name
                FROM "AcPersonal"
                WHERE UPPER("Code") = UPPER($1) AND "Active" = 'Y'
            """, request.code)

            if not user:
                return {"success": False, "error": "User not found or inactive"}

            # Check if password already set
            existing = await conn.fetchrow("""
                SELECT code FROM kpi_user_auth WHERE UPPER(code) = UPPER($1)
            """, request.code)

            if existing:
                return {"success": False, "error": "Password already set. Use change-password instead."}

            # Hash and store the password
            password_hash = hash_password(request.new_password)
            await conn.execute("""
                INSERT INTO kpi_user_auth (code, password_hash)
                VALUES ($1, $2)
            """, user['code'], password_hash)

            return {
                "success": True,
                "message": "Password set successfully. You can now login."
            }

    except Exception as e:
        return {"success": False, "error": f"Failed to set password: {str(e)}"}


@app.get("/health")
async def health():
    try:
        async with pool.acquire() as conn:
            await conn.fetchval("SELECT 1")
        return {
            "status": "healthy",
            "database": "connected",
            "host": connected_host,
            "is_internal": "internal" in (connected_host or "")
        }
    except Exception as e:
        return {"status": "unhealthy", "error": str(e), "host": connected_host}


@app.get("/api/v1/kpi/me")
async def get_my_dashboard(
    staff_id: str = Query(..., description="Staff ID"),
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None)
):
    """Get personal KPI dashboard - HYBRID: MV for history + real-time for today.

    Uses materialized views for historical data (fast) and queries base tables
    only for today's data (small dataset, real-time).
    """
    if not start_date:
        start_date = date.today().replace(day=1)
    if not end_date:
        end_date = date.today()

    today = date.today()
    # Use timestamp range for index efficiency (avoid ::date cast which causes seq scan)
    today_start = datetime.combine(today, datetime.min.time())
    today_end = datetime.combine(today + timedelta(days=1), datetime.min.time())

    try:
        async with pool.acquire() as conn:
            # HYBRID APPROACH: MV for historical + real-time for today
            # Step 1: Get historical data from materialized view (fast)
            mv_summary = await conn.fetchrow("""
                SELECT
                    staff_id,
                    MAX(outlet_id) as outlet_id,
                    COALESCE(SUM(transactions), 0) as transactions,
                    COALESCE(SUM(total_sales), 0) as total_sales,
                    COALESCE(SUM(gross_profit), 0) as gross_profit,
                    COALESCE(SUM(house_brand_sales), 0) as house_brand_sales,
                    COALESCE(SUM(focused_1_sales), 0) as focused_1_sales,
                    COALESCE(SUM(focused_2_sales), 0) as focused_2_sales,
                    COALESCE(SUM(focused_3_sales), 0) as focused_3_sales,
                    COALESCE(SUM(pwp_sales), 0) as pwp_sales,
                    COALESCE(SUM(clearance_sales), 0) as clearance_sales
                FROM analytics.mv_staff_daily_kpi
                WHERE staff_id = $1
                  AND sale_date BETWEEN $2 AND $3
                  AND sale_date < $4
                GROUP BY staff_id
            """, staff_id, start_date, end_date, today)

            # Step 2: Get today's data from cache (10-40x faster than raw tables)
            # Cache is updated every 60 seconds by sync service
            today_summary = None
            if end_date >= today:
                today_summary = await conn.fetchrow("""
                    SELECT
                        outlet_id,
                        transactions,
                        total_sales,
                        gross_profit,
                        house_brand_sales,
                        focused_1_sales AS focused_1_sales,
                        focused_2_sales AS focused_2_sales,
                        focused_3_sales AS focused_3_sales,
                        clearance_sales,
                        pwp_sales
                    FROM kpi.today_sales_cache
                    WHERE staff_id = $1 AND sale_date = CURRENT_DATE
                """, staff_id)

            # Combine MV + today's data
            def safe_float(val):
                return float(val) if val else 0.0

            def safe_int(val):
                return int(val) if val else 0

            total_sales = safe_float(mv_summary['total_sales'] if mv_summary else 0) + safe_float(today_summary['total_sales'] if today_summary else 0)
            transactions = safe_int(mv_summary['transactions'] if mv_summary else 0) + safe_int(today_summary['transactions'] if today_summary else 0)
            gross_profit = safe_float(mv_summary['gross_profit'] if mv_summary else 0) + safe_float(today_summary['gross_profit'] if today_summary else 0)
            house_brand = safe_float(mv_summary['house_brand_sales'] if mv_summary else 0) + safe_float(today_summary['house_brand_sales'] if today_summary else 0)
            focused_1 = safe_float(mv_summary['focused_1_sales'] if mv_summary else 0) + safe_float(today_summary['focused_1_sales'] if today_summary else 0)
            focused_2 = safe_float(mv_summary['focused_2_sales'] if mv_summary else 0) + safe_float(today_summary['focused_2_sales'] if today_summary else 0)
            focused_3 = safe_float(mv_summary['focused_3_sales'] if mv_summary else 0) + safe_float(today_summary['focused_3_sales'] if today_summary else 0)
            pwp = safe_float(mv_summary['pwp_sales'] if mv_summary else 0) + safe_float(today_summary['pwp_sales'] if today_summary else 0)
            clearance = safe_float(mv_summary['clearance_sales'] if mv_summary else 0) + safe_float(today_summary['clearance_sales'] if today_summary else 0)

            # Get outlet_id from either source
            outlet_id = (today_summary['outlet_id'] if today_summary and today_summary['outlet_id']
                        else mv_summary['outlet_id'] if mv_summary else None)

            if total_sales == 0 and transactions == 0:
                # Check if staff exists
                staff_check = await conn.fetchrow("""
                    SELECT "AcSalesmanID", "AcSalesmanName", s."AcSalesmanGroupID" as outlet_id
                    FROM "AcSalesman" s WHERE "AcSalesmanID" = $1
                """, staff_id)
                if not staff_check:
                    raise HTTPException(status_code=404, detail="Staff not found")
                outlet_id = staff_check['outlet_id']

            # Get staff name
            staff_info = await conn.fetchrow("""
                SELECT "AcSalesmanName" FROM "AcSalesman" WHERE "AcSalesmanID" = $1
            """, staff_id)

            # Get outlet name
            outlet_info = await conn.fetchrow("""
                SELECT "AcLocationDesc" as outlet_name FROM "AcLocation" WHERE "AcLocationID" = $1
            """, outlet_id)

            # Get rankings from MV
            rankings = await conn.fetchrow("""
                SELECT outlet_rank_sales, company_rank_sales, sales_percentile
                FROM analytics.mv_staff_rankings
                WHERE staff_id = $1
                  AND month = DATE_TRUNC('month', $2::date)
            """, staff_id, start_date)

            # Get daily breakdown from MV (fast)
            daily = await conn.fetch("""
                SELECT
                    sale_date,
                    transactions,
                    total_sales,
                    house_brand_sales
                FROM analytics.mv_staff_daily_kpi
                WHERE staff_id = $1
                  AND sale_date BETWEEN $2 AND $3
                ORDER BY sale_date
            """, staff_id, start_date, end_date)

            return {
                "success": True,
                "data": {
                    "staff_id": staff_id,
                    "staff_name": staff_info['AcSalesmanName'] if staff_info else "Unknown",
                    "outlet_id": outlet_id,
                    "outlet_name": outlet_info['outlet_name'] if outlet_info else "Unknown",
                    "period": {
                        "start": start_date.isoformat(),
                        "end": end_date.isoformat()
                    },
                    "kpis": {
                        "total_sales": round(total_sales, 2),
                        "house_brand": round(house_brand, 2),
                        "focused_1": round(focused_1, 2),
                        "focused_2": round(focused_2, 2),
                        "focused_3": round(focused_3, 2),
                        "pwp": round(pwp, 2),
                        "clearance": round(clearance, 2),
                        "transactions": transactions,
                        "gross_profit": round(gross_profit, 2)
                    },
                    "rankings": {
                        "outlet_rank": rankings['outlet_rank_sales'] if rankings else None,
                        "company_rank": rankings['company_rank_sales'] if rankings else None,
                        "percentile": float(rankings['sales_percentile']) if rankings and rankings['sales_percentile'] else None
                    },
                    "daily": [
                        {
                            "date": row['sale_date'].isoformat(),
                            "sales": float(row['total_sales'] or 0),
                            "house_brand": float(row['house_brand_sales'] or 0)
                        }
                        for row in daily
                    ]
                }
            }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Database error: {str(e)}")


@app.get("/api/v1/kpi/leaderboard")
async def get_leaderboard(
    month: Optional[str] = Query(None, description="Month in YYYY-MM format"),
    staff_id: Optional[str] = Query(None, description="Current user's staff ID to include their position"),
    limit: int = Query(20, ge=1, le=100)
):
    """Get company-wide staff rankings leaderboard with HYBRID approach.

    Returns top N + logged-in user's position if not in top N.
    Uses MV for historical data + real-time query for today's data.
    """
    if month:
        try:
            period = datetime.strptime(month, "%Y-%m").date()
        except:
            raise HTTPException(status_code=400, detail="Invalid month format")
    else:
        period = date.today()

    today = date.today()
    month_start = period.replace(day=1)
    # For current month, include up to today; for past months, include whole month
    if period.year == today.year and period.month == today.month:
        month_end = today
    else:
        # Last day of the month
        next_month = month_start.replace(day=28) + timedelta(days=4)
        month_end = next_month - timedelta(days=next_month.day)

    today_start = datetime.combine(today, datetime.min.time())
    today_end = datetime.combine(today + timedelta(days=1), datetime.min.time())

    async with pool.acquire() as conn:
        # HYBRID APPROACH: MV for historical + real-time for today
        # Step 1: Get all staff data from MV (historical, excludes today)
        mv_data = await conn.fetch("""
            SELECT
                k.staff_id,
                s."AcSalesmanName" as staff_name,
                s."AcSalesmanGroupID" as outlet_id,
                COALESCE(SUM(k.total_sales), 0) as total_sales,
                COALESCE(SUM(k.house_brand_sales), 0) as house_brand_sales,
                COALESCE(SUM(k.focused_1_sales), 0) as focused_1_sales,
                COALESCE(SUM(k.focused_2_sales), 0) as focused_2_sales,
                COALESCE(SUM(k.focused_3_sales), 0) as focused_3_sales,
                COALESCE(SUM(k.pwp_sales), 0) as pwp_sales,
                COALESCE(SUM(k.clearance_sales), 0) as clearance_sales,
                COALESCE(SUM(k.transactions), 0) as transactions
            FROM analytics.mv_staff_daily_kpi k
            LEFT JOIN "AcSalesman" s ON k.staff_id = s."AcSalesmanID"
            WHERE k.sale_date >= $1 AND k.sale_date <= $2
              AND k.sale_date < $3
            GROUP BY k.staff_id, s."AcSalesmanName", s."AcSalesmanGroupID"
        """, month_start, month_end, today)

        # Step 2: Get today's data from cache (10-40x faster than raw tables)
        today_data = {}
        if month_end >= today:
            today_rows = await conn.fetch("""
                SELECT
                    staff_id,
                    transactions,
                    total_sales,
                    house_brand_sales,
                    focused_1_sales,
                    focused_2_sales,
                    focused_3_sales,
                    clearance_sales,
                    pwp_sales
                FROM kpi.today_sales_cache
                WHERE sale_date = CURRENT_DATE
            """)

            for row in today_rows:
                today_data[row['staff_id']] = dict(row)

        # Step 3: Combine MV + today's data
        combined = {}
        for row in mv_data:
            sid = row['staff_id']
            combined[sid] = {
                'staff_id': sid,
                'staff_name': row['staff_name'] or 'Unknown',
                'outlet_id': row['outlet_id'],
                'total_sales': float(row['total_sales'] or 0),
                'house_brand': float(row['house_brand_sales'] or 0),
                'focused_1': float(row['focused_1_sales'] or 0),
                'focused_2': float(row['focused_2_sales'] or 0),
                'focused_3': float(row['focused_3_sales'] or 0),
                'pwp': float(row['pwp_sales'] or 0),
                'clearance': float(row['clearance_sales'] or 0),
                'transactions': int(row['transactions'] or 0),
            }

        # Add today's data
        for sid, tdata in today_data.items():
            if sid in combined:
                combined[sid]['total_sales'] += float(tdata['total_sales'] or 0)
                combined[sid]['house_brand'] += float(tdata['house_brand_sales'] or 0)
                combined[sid]['focused_1'] += float(tdata['focused_1_sales'] or 0)
                combined[sid]['focused_2'] += float(tdata['focused_2_sales'] or 0)
                combined[sid]['focused_3'] += float(tdata['focused_3_sales'] or 0)
                combined[sid]['pwp'] += float(tdata['pwp_sales'] or 0)
                combined[sid]['clearance'] += float(tdata['clearance_sales'] or 0)
                combined[sid]['transactions'] += int(tdata['transactions'] or 0)
            else:
                # New staff who only has today's data
                staff_info = await conn.fetchrow("""
                    SELECT "AcSalesmanName", "AcSalesmanGroupID"
                    FROM "AcSalesman" WHERE "AcSalesmanID" = $1
                """, sid)
                combined[sid] = {
                    'staff_id': sid,
                    'staff_name': staff_info['AcSalesmanName'] if staff_info else 'Unknown',
                    'outlet_id': staff_info['AcSalesmanGroupID'] if staff_info else None,
                    'total_sales': float(tdata['total_sales'] or 0),
                    'house_brand': float(tdata['house_brand_sales'] or 0),
                    'focused_1': float(tdata['focused_1_sales'] or 0),
                    'focused_2': float(tdata['focused_2_sales'] or 0),
                    'focused_3': float(tdata['focused_3_sales'] or 0),
                    'pwp': float(tdata['pwp_sales'] or 0),
                    'clearance': float(tdata['clearance_sales'] or 0),
                    'transactions': int(tdata['transactions'] or 0),
                }

        # Step 4: Sort by total_sales and assign ranks
        sorted_staff = sorted(combined.values(), key=lambda x: x['total_sales'], reverse=True)
        total_count = len(sorted_staff)

        # Assign ranks and percentiles
        for i, staff in enumerate(sorted_staff):
            staff['rank'] = i + 1
            staff['percentile'] = round((1 - (i / total_count)) * 100, 1) if total_count > 0 else 0

        # Get top N
        top_n = sorted_staff[:limit]

        # Find user position if not in top N
        user_position = None
        if staff_id:
            top_ids = [s['staff_id'] for s in top_n]
            if staff_id not in top_ids:
                for staff in sorted_staff:
                    if staff['staff_id'] == staff_id:
                        user_position = staff
                        break

        response_data = {
            "period": period.strftime("%Y-%m"),
            "scope": "company",
            "rankings": top_n
        }

        if user_position:
            response_data["user_position"] = user_position

        return {
            "success": True,
            "data": response_data
        }


@app.get("/api/v1/kpi/team")
async def get_team_overview(
    outlet_id: Optional[str] = Query(None, description="Physical outlet ID for summary (empty/ALL = all outlets)"),
    group_id: Optional[str] = Query(None, description="Staff group ID for filtering (defaults to outlet_id)"),
    outlet_ids: Optional[str] = Query(None, description="Comma-separated outlet IDs for filtering (for Admin/OOM)"),
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None)
):
    """Get team overview - HYBRID: MV for history + real-time for today.

    - outlet_id: Physical outlet where sales are made (for summary totals). Empty/ALL = aggregate all.
    - outlet_ids: Comma-separated list of outlets to filter (for Admin/OOM viewing multiple outlets)
    - group_id: Staff team/group for filtering assigned staff (defaults to outlet_id)
    """
    if not start_date:
        start_date = date.today().replace(day=1)
    if not end_date:
        end_date = date.today()

    today = date.today()
    today_start = datetime.combine(today, datetime.min.time())
    today_end = datetime.combine(today + timedelta(days=1), datetime.min.time())

    # Determine if viewing ALL outlets or specific outlet
    view_all = not outlet_id or outlet_id.upper() == 'ALL'
    outlet_list = []
    if outlet_ids:
        outlet_list = [o.strip() for o in outlet_ids.split(',') if o.strip()]

    staff_group = group_id or outlet_id

    async with pool.acquire() as conn:
        # If no group_id provided, try to find corresponding group for the outlet
        # This handles cases like HQ (location) → BG (group)
        if not group_id and outlet_id and not view_all:
            # Check if outlet_id is also a valid AcSalesmanGroupID
            group_check = await conn.fetchrow("""
                SELECT COUNT(*) as cnt FROM "AcSalesman"
                WHERE "AcSalesmanGroupID" = $1 AND "Active" = 'Y'
            """, outlet_id)

            if group_check and group_check['cnt'] == 0:
                # No staff in this group - find the actual group from sales data
                actual_group = await conn.fetchrow("""
                    SELECT DISTINCT s."AcSalesmanGroupID" as group_id
                    FROM analytics.mv_staff_daily_kpi k
                    INNER JOIN "AcSalesman" s ON k.staff_id = s."AcSalesmanID"
                    WHERE k.outlet_id = $1
                    LIMIT 1
                """, outlet_id)
                if actual_group and actual_group['group_id']:
                    staff_group = actual_group['group_id']

        # HYBRID: Get outlet summary from MV (excluding today) + real-time for today
        if view_all and outlet_list:
            # Aggregate across specified outlets
            mv_summary = await conn.fetchrow("""
                SELECT
                    COALESCE(SUM(transactions), 0) as transactions,
                    COALESCE(SUM(total_sales), 0) as total_sales,
                    COALESCE(SUM(gross_profit), 0) as gross_profit,
                    COALESCE(SUM(house_brand_sales), 0) as house_brand_sales,
                    COALESCE(SUM(focused_1_sales), 0) as focused_1_sales,
                    COALESCE(SUM(focused_2_sales), 0) as focused_2_sales,
                    COALESCE(SUM(focused_3_sales), 0) as focused_3_sales,
                    COALESCE(SUM(pwp_sales), 0) as pwp_sales,
                    COALESCE(SUM(clearance_sales), 0) as clearance_sales
                FROM analytics.mv_outlet_daily_kpi
                WHERE outlet_id = ANY($1)
                  AND sale_date BETWEEN $2 AND $3
                  AND sale_date < $4
            """, outlet_list, start_date, end_date, today)
        elif view_all:
            # Aggregate across ALL outlets
            mv_summary = await conn.fetchrow("""
                SELECT
                    COALESCE(SUM(transactions), 0) as transactions,
                    COALESCE(SUM(total_sales), 0) as total_sales,
                    COALESCE(SUM(gross_profit), 0) as gross_profit,
                    COALESCE(SUM(house_brand_sales), 0) as house_brand_sales,
                    COALESCE(SUM(focused_1_sales), 0) as focused_1_sales,
                    COALESCE(SUM(focused_2_sales), 0) as focused_2_sales,
                    COALESCE(SUM(focused_3_sales), 0) as focused_3_sales,
                    COALESCE(SUM(pwp_sales), 0) as pwp_sales,
                    COALESCE(SUM(clearance_sales), 0) as clearance_sales
                FROM analytics.mv_outlet_daily_kpi
                WHERE sale_date BETWEEN $1 AND $2
                  AND sale_date < $3
            """, start_date, end_date, today)
        else:
            # Single outlet
            mv_summary = await conn.fetchrow("""
                SELECT
                    COALESCE(SUM(transactions), 0) as transactions,
                    COALESCE(SUM(total_sales), 0) as total_sales,
                    COALESCE(SUM(gross_profit), 0) as gross_profit,
                    COALESCE(SUM(house_brand_sales), 0) as house_brand_sales,
                    COALESCE(SUM(focused_1_sales), 0) as focused_1_sales,
                    COALESCE(SUM(focused_2_sales), 0) as focused_2_sales,
                    COALESCE(SUM(focused_3_sales), 0) as focused_3_sales,
                    COALESCE(SUM(pwp_sales), 0) as pwp_sales,
                    COALESCE(SUM(clearance_sales), 0) as clearance_sales
                FROM analytics.mv_outlet_daily_kpi
                WHERE outlet_id = $1
                  AND sale_date BETWEEN $2 AND $3
                  AND sale_date < $4
            """, outlet_id, start_date, end_date, today)

        # Step 2: Today's data from base tables (real-time)
        # Note: asyncpg doesn't support parallel queries on single connection, run sequentially
        # Requires indexes: idx_accsm_location_date, idx_accusinvoicem_location_date
        today_summary = None
        if end_date >= today:
            if view_all and outlet_list:
                # Filter by specific allowed outlets
                cash_result = await conn.fetchrow("""
                    SELECT
                        COUNT(DISTINCT m."DocumentNo") as transactions,
                        COALESCE(SUM(d."ItemTotal"), 0) as total_sales,
                        COALESCE(SUM(d."ItemTotal" - COALESCE(d."ItemCost" * d."ItemQuantity", 0)), 0) as gross_profit,
                        COALESCE(SUM(CASE WHEN s."AcStockUDGroup1ID" = 'FLTHB' THEN d."ItemTotal" ELSE 0 END), 0) as house_brand_sales,
                        COALESCE(SUM(CASE WHEN s."AcStockUDGroup1ID" = 'FLTF1' THEN d."ItemTotal" ELSE 0 END), 0) as focused_1_sales,
                        COALESCE(SUM(CASE WHEN s."AcStockUDGroup1ID" = 'FLTF2' THEN d."ItemTotal" ELSE 0 END), 0) as focused_2_sales,
                        COALESCE(SUM(CASE WHEN s."AcStockUDGroup1ID" = 'FLTF3' THEN d."ItemTotal" ELSE 0 END), 0) as focused_3_sales,
                        COALESCE(SUM(CASE WHEN s."AcStockUDGroup1ID" = 'STOCK CLEARANCE' THEN d."ItemTotal" ELSE 0 END), 0) as clearance_sales
                    FROM "AcCSM" m
                    INNER JOIN "AcCSD" d ON m."DocumentNo" = d."DocumentNo"
                    LEFT JOIN "AcStockCompany" s ON d."AcStockID" = s."AcStockID" AND d."AcStockUOMID" = s."AcStockUOMID"
                    WHERE m."DocumentDate" >= $1
                      AND m."DocumentDate" < $2
                      AND m."AcLocationID" = ANY($3)
                """, today_start, today_end, outlet_list)

                invoice_result = await conn.fetchrow("""
                    SELECT
                        COUNT(DISTINCT m."AcCusInvoiceMID") as transactions,
                        COALESCE(SUM(d."ItemTotalPrice"), 0) as total_sales,
                        COALESCE(SUM(CASE WHEN s."AcStockUDGroup1ID" = 'FLTHB' THEN d."ItemTotalPrice" ELSE 0 END), 0) as house_brand_sales,
                        COALESCE(SUM(CASE WHEN s."AcStockUDGroup1ID" = 'FLTF1' THEN d."ItemTotalPrice" ELSE 0 END), 0) as focused_1_sales,
                        COALESCE(SUM(CASE WHEN s."AcStockUDGroup1ID" = 'FLTF2' THEN d."ItemTotalPrice" ELSE 0 END), 0) as focused_2_sales,
                        COALESCE(SUM(CASE WHEN s."AcStockUDGroup1ID" = 'FLTF3' THEN d."ItemTotalPrice" ELSE 0 END), 0) as focused_3_sales,
                        COALESCE(SUM(CASE WHEN s."AcStockUDGroup1ID" = 'STOCK CLEARANCE' THEN d."ItemTotalPrice" ELSE 0 END), 0) as clearance_sales
                    FROM "AcCusInvoiceM" m
                    INNER JOIN "AcCusInvoiceD" d ON m."AcCusInvoiceMID" = d."AcCusInvoiceMID"
                    LEFT JOIN "AcStockCompany" s ON d."AcStockID" = s."AcStockID" AND d."AcStockUOMID" = s."AcStockUOMID"
                    WHERE m."DocumentDate" >= $1
                      AND m."DocumentDate" < $2
                      AND m."AcLocationID" = ANY($3)
                """, today_start, today_end, outlet_list)

                pwp_result = await conn.fetchrow("""
                    SELECT COALESCE(SUM(d."ItemTotal"), 0) as pwp_sales
                    FROM "AcCSD" d
                    INNER JOIN "AcCSM" m ON d."DocumentNo" = m."DocumentNo"
                    INNER JOIN "AcCSDPromotionType" pt ON d."DocumentNo" = pt."DocumentNo" AND d."ItemNo" = pt."ItemNo"
                    WHERE pt."AcPromotionSettingID" = 'PURCHASE WITH PURCHASE'
                      AND m."DocumentDate" >= $1
                      AND m."DocumentDate" < $2
                      AND m."AcLocationID" = ANY($3)
                """, today_start, today_end, outlet_list)
            elif view_all:
                # All outlets - no location filter
                cash_result = await conn.fetchrow("""
                    SELECT
                        COUNT(DISTINCT m."DocumentNo") as transactions,
                        COALESCE(SUM(d."ItemTotal"), 0) as total_sales,
                        COALESCE(SUM(d."ItemTotal" - COALESCE(d."ItemCost" * d."ItemQuantity", 0)), 0) as gross_profit,
                        COALESCE(SUM(CASE WHEN s."AcStockUDGroup1ID" = 'FLTHB' THEN d."ItemTotal" ELSE 0 END), 0) as house_brand_sales,
                        COALESCE(SUM(CASE WHEN s."AcStockUDGroup1ID" = 'FLTF1' THEN d."ItemTotal" ELSE 0 END), 0) as focused_1_sales,
                        COALESCE(SUM(CASE WHEN s."AcStockUDGroup1ID" = 'FLTF2' THEN d."ItemTotal" ELSE 0 END), 0) as focused_2_sales,
                        COALESCE(SUM(CASE WHEN s."AcStockUDGroup1ID" = 'FLTF3' THEN d."ItemTotal" ELSE 0 END), 0) as focused_3_sales,
                        COALESCE(SUM(CASE WHEN s."AcStockUDGroup1ID" = 'STOCK CLEARANCE' THEN d."ItemTotal" ELSE 0 END), 0) as clearance_sales
                    FROM "AcCSM" m
                    INNER JOIN "AcCSD" d ON m."DocumentNo" = d."DocumentNo"
                    LEFT JOIN "AcStockCompany" s ON d."AcStockID" = s."AcStockID" AND d."AcStockUOMID" = s."AcStockUOMID"
                    WHERE m."DocumentDate" >= $1
                      AND m."DocumentDate" < $2
                """, today_start, today_end)

                invoice_result = await conn.fetchrow("""
                    SELECT
                        COUNT(DISTINCT m."AcCusInvoiceMID") as transactions,
                        COALESCE(SUM(d."ItemTotalPrice"), 0) as total_sales,
                        COALESCE(SUM(CASE WHEN s."AcStockUDGroup1ID" = 'FLTHB' THEN d."ItemTotalPrice" ELSE 0 END), 0) as house_brand_sales,
                        COALESCE(SUM(CASE WHEN s."AcStockUDGroup1ID" = 'FLTF1' THEN d."ItemTotalPrice" ELSE 0 END), 0) as focused_1_sales,
                        COALESCE(SUM(CASE WHEN s."AcStockUDGroup1ID" = 'FLTF2' THEN d."ItemTotalPrice" ELSE 0 END), 0) as focused_2_sales,
                        COALESCE(SUM(CASE WHEN s."AcStockUDGroup1ID" = 'FLTF3' THEN d."ItemTotalPrice" ELSE 0 END), 0) as focused_3_sales,
                        COALESCE(SUM(CASE WHEN s."AcStockUDGroup1ID" = 'STOCK CLEARANCE' THEN d."ItemTotalPrice" ELSE 0 END), 0) as clearance_sales
                    FROM "AcCusInvoiceM" m
                    INNER JOIN "AcCusInvoiceD" d ON m."AcCusInvoiceMID" = d."AcCusInvoiceMID"
                    LEFT JOIN "AcStockCompany" s ON d."AcStockID" = s."AcStockID" AND d."AcStockUOMID" = s."AcStockUOMID"
                    WHERE m."DocumentDate" >= $1
                      AND m."DocumentDate" < $2
                """, today_start, today_end)

                pwp_result = await conn.fetchrow("""
                    SELECT COALESCE(SUM(d."ItemTotal"), 0) as pwp_sales
                    FROM "AcCSD" d
                    INNER JOIN "AcCSM" m ON d."DocumentNo" = m."DocumentNo"
                    INNER JOIN "AcCSDPromotionType" pt ON d."DocumentNo" = pt."DocumentNo" AND d."ItemNo" = pt."ItemNo"
                    WHERE pt."AcPromotionSettingID" = 'PURCHASE WITH PURCHASE'
                      AND m."DocumentDate" >= $1
                      AND m."DocumentDate" < $2
                """, today_start, today_end)
            else:
                # Single outlet - filter by location
                cash_result = await conn.fetchrow("""
                    SELECT
                        COUNT(DISTINCT m."DocumentNo") as transactions,
                        COALESCE(SUM(d."ItemTotal"), 0) as total_sales,
                        COALESCE(SUM(d."ItemTotal" - COALESCE(d."ItemCost" * d."ItemQuantity", 0)), 0) as gross_profit,
                        COALESCE(SUM(CASE WHEN s."AcStockUDGroup1ID" = 'FLTHB' THEN d."ItemTotal" ELSE 0 END), 0) as house_brand_sales,
                        COALESCE(SUM(CASE WHEN s."AcStockUDGroup1ID" = 'FLTF1' THEN d."ItemTotal" ELSE 0 END), 0) as focused_1_sales,
                        COALESCE(SUM(CASE WHEN s."AcStockUDGroup1ID" = 'FLTF2' THEN d."ItemTotal" ELSE 0 END), 0) as focused_2_sales,
                        COALESCE(SUM(CASE WHEN s."AcStockUDGroup1ID" = 'FLTF3' THEN d."ItemTotal" ELSE 0 END), 0) as focused_3_sales,
                        COALESCE(SUM(CASE WHEN s."AcStockUDGroup1ID" = 'STOCK CLEARANCE' THEN d."ItemTotal" ELSE 0 END), 0) as clearance_sales
                    FROM "AcCSM" m
                    INNER JOIN "AcCSD" d ON m."DocumentNo" = d."DocumentNo"
                    LEFT JOIN "AcStockCompany" s ON d."AcStockID" = s."AcStockID" AND d."AcStockUOMID" = s."AcStockUOMID"
                    WHERE m."AcLocationID" = $1
                      AND m."DocumentDate" >= $2
                      AND m."DocumentDate" < $3
                """, outlet_id, today_start, today_end)

                invoice_result = await conn.fetchrow("""
                    SELECT
                        COUNT(DISTINCT m."AcCusInvoiceMID") as transactions,
                        COALESCE(SUM(d."ItemTotalPrice"), 0) as total_sales,
                        COALESCE(SUM(CASE WHEN s."AcStockUDGroup1ID" = 'FLTHB' THEN d."ItemTotalPrice" ELSE 0 END), 0) as house_brand_sales,
                        COALESCE(SUM(CASE WHEN s."AcStockUDGroup1ID" = 'FLTF1' THEN d."ItemTotalPrice" ELSE 0 END), 0) as focused_1_sales,
                        COALESCE(SUM(CASE WHEN s."AcStockUDGroup1ID" = 'FLTF2' THEN d."ItemTotalPrice" ELSE 0 END), 0) as focused_2_sales,
                        COALESCE(SUM(CASE WHEN s."AcStockUDGroup1ID" = 'FLTF3' THEN d."ItemTotalPrice" ELSE 0 END), 0) as focused_3_sales,
                        COALESCE(SUM(CASE WHEN s."AcStockUDGroup1ID" = 'STOCK CLEARANCE' THEN d."ItemTotalPrice" ELSE 0 END), 0) as clearance_sales
                    FROM "AcCusInvoiceM" m
                    INNER JOIN "AcCusInvoiceD" d ON m."AcCusInvoiceMID" = d."AcCusInvoiceMID"
                    LEFT JOIN "AcStockCompany" s ON d."AcStockID" = s."AcStockID" AND d."AcStockUOMID" = s."AcStockUOMID"
                    WHERE m."AcLocationID" = $1
                      AND m."DocumentDate" >= $2
                      AND m."DocumentDate" < $3
                """, outlet_id, today_start, today_end)

                pwp_result = await conn.fetchrow("""
                    SELECT COALESCE(SUM(d."ItemTotal"), 0) as pwp_sales
                    FROM "AcCSD" d
                    INNER JOIN "AcCSM" m ON d."DocumentNo" = m."DocumentNo"
                    INNER JOIN "AcCSDPromotionType" pt ON d."DocumentNo" = pt."DocumentNo" AND d."ItemNo" = pt."ItemNo"
                    WHERE pt."AcPromotionSettingID" = 'PURCHASE WITH PURCHASE'
                      AND m."AcLocationID" = $1
                      AND m."DocumentDate" >= $2
                      AND m."DocumentDate" < $3
                """, outlet_id, today_start, today_end)

            # Combine results
            today_summary = {
                'transactions': (cash_result['transactions'] or 0) + (invoice_result['transactions'] or 0),
                'total_sales': float(cash_result['total_sales'] or 0) + float(invoice_result['total_sales'] or 0),
                'gross_profit': float(cash_result['gross_profit'] or 0),  # Invoice GP not tracked in real-time
                'house_brand_sales': float(cash_result['house_brand_sales'] or 0) + float(invoice_result['house_brand_sales'] or 0),
                'focused_1_sales': float(cash_result['focused_1_sales'] or 0) + float(invoice_result['focused_1_sales'] or 0),
                'focused_2_sales': float(cash_result['focused_2_sales'] or 0) + float(invoice_result['focused_2_sales'] or 0),
                'focused_3_sales': float(cash_result['focused_3_sales'] or 0) + float(invoice_result['focused_3_sales'] or 0),
                'clearance_sales': float(cash_result['clearance_sales'] or 0) + float(invoice_result['clearance_sales'] or 0),
                'pwp_sales': float(pwp_result['pwp_sales'] or 0)
            }

        # Combine MV + today
        def sf(val): return float(val) if val else 0.0
        def si(val): return int(val) if val else 0

        summary = {
            "transactions": si(mv_summary['transactions']) + si(today_summary['transactions'] if today_summary else 0),
            "total_sales": sf(mv_summary['total_sales']) + sf(today_summary['total_sales'] if today_summary else 0),
            "gross_profit": sf(mv_summary['gross_profit']) + sf(today_summary['gross_profit'] if today_summary else 0),
            "house_brand": sf(mv_summary['house_brand_sales']) + sf(today_summary['house_brand_sales'] if today_summary else 0),
            "focused_1": sf(mv_summary['focused_1_sales']) + sf(today_summary['focused_1_sales'] if today_summary else 0),
            "focused_2": sf(mv_summary['focused_2_sales']) + sf(today_summary['focused_2_sales'] if today_summary else 0),
            "focused_3": sf(mv_summary['focused_3_sales']) + sf(today_summary['focused_3_sales'] if today_summary else 0),
            "pwp": sf(mv_summary['pwp_sales']) + sf(today_summary['pwp_sales'] if today_summary else 0),
            "clearance": sf(mv_summary['clearance_sales']) + sf(today_summary['clearance_sales'] if today_summary else 0),
        }

        # Get outlet name
        outlet_name = "All Outlets" if view_all else None
        if not view_all:
            outlet_info = await conn.fetchrow("""
                SELECT "AcLocationDesc" as outlet_name FROM "AcLocation" WHERE "AcLocationID" = $1
            """, outlet_id)
            outlet_name = outlet_info['outlet_name'] if outlet_info else outlet_id

        # Staff performance - HYBRID: MV for history + real-time for today
        staff = []
        staff_data = {}  # staff_id -> aggregated data

        # Helper to safely get historical date range (excluding today)
        hist_end = min(end_date, today - timedelta(days=1)) if end_date >= today else end_date
        has_historical = start_date <= hist_end
        has_today = end_date >= today

        if view_all:
            # All outlets - show top performers across all allowed outlets
            # Step 1: Get historical data from MV (excluding today)
            if has_historical:
                if outlet_list:
                    mv_staff = await conn.fetch("""
                        SELECT
                            k.staff_id,
                            s."AcSalesmanName" as staff_name,
                            s."AcSalesmanGroupID" as outlet_id,
                            COALESCE(SUM(k.transactions), 0) as transactions,
                            COALESCE(SUM(k.total_sales), 0) as total_sales,
                            COALESCE(SUM(k.house_brand_sales), 0) as house_brand_sales,
                            COALESCE(SUM(k.focused_1_sales), 0) as focused_1_sales,
                            COALESCE(SUM(k.focused_2_sales), 0) as focused_2_sales,
                            COALESCE(SUM(k.focused_3_sales), 0) as focused_3_sales,
                            COALESCE(SUM(k.pwp_sales), 0) as pwp_sales,
                            COALESCE(SUM(k.clearance_sales), 0) as clearance_sales,
                            r.company_rank_sales as rank
                        FROM analytics.mv_staff_daily_kpi k
                        LEFT JOIN "AcSalesman" s ON k.staff_id = s."AcSalesmanID"
                        LEFT JOIN analytics.mv_staff_rankings r
                            ON k.staff_id = r.staff_id
                            AND r.month = DATE_TRUNC('month', $1::date)
                            AND r.outlet_id = k.outlet_id
                        WHERE k.sale_date BETWEEN $1 AND $2
                          AND k.outlet_id = ANY($3)
                        GROUP BY k.staff_id, s."AcSalesmanName", s."AcSalesmanGroupID", r.company_rank_sales
                    """, start_date, hist_end, outlet_list)
                else:
                    mv_staff = await conn.fetch("""
                        SELECT
                            k.staff_id,
                            s."AcSalesmanName" as staff_name,
                            s."AcSalesmanGroupID" as outlet_id,
                            COALESCE(SUM(k.transactions), 0) as transactions,
                            COALESCE(SUM(k.total_sales), 0) as total_sales,
                            COALESCE(SUM(k.house_brand_sales), 0) as house_brand_sales,
                            COALESCE(SUM(k.focused_1_sales), 0) as focused_1_sales,
                            COALESCE(SUM(k.focused_2_sales), 0) as focused_2_sales,
                            COALESCE(SUM(k.focused_3_sales), 0) as focused_3_sales,
                            COALESCE(SUM(k.pwp_sales), 0) as pwp_sales,
                            COALESCE(SUM(k.clearance_sales), 0) as clearance_sales,
                            r.company_rank_sales as rank
                        FROM analytics.mv_staff_daily_kpi k
                        LEFT JOIN "AcSalesman" s ON k.staff_id = s."AcSalesmanID"
                        LEFT JOIN analytics.mv_staff_rankings r
                            ON k.staff_id = r.staff_id
                            AND r.month = DATE_TRUNC('month', $1::date)
                            AND r.outlet_id = k.outlet_id
                        WHERE k.sale_date BETWEEN $1 AND $2
                        GROUP BY k.staff_id, s."AcSalesmanName", s."AcSalesmanGroupID", r.company_rank_sales
                    """, start_date, hist_end)

                for row in mv_staff:
                    sid = row['staff_id']
                    staff_data[sid] = {
                        'staff_id': sid,
                        'staff_name': row['staff_name'],
                        'outlet_id': row['outlet_id'],
                        'transactions': int(row['transactions'] or 0),
                        'total_sales': float(row['total_sales'] or 0),
                        'house_brand_sales': float(row['house_brand_sales'] or 0),
                        'focused_1_sales': float(row['focused_1_sales'] or 0),
                        'focused_2_sales': float(row['focused_2_sales'] or 0),
                        'focused_3_sales': float(row['focused_3_sales'] or 0),
                        'pwp_sales': float(row['pwp_sales'] or 0),
                        'clearance_sales': float(row['clearance_sales'] or 0),
                        'rank': row['rank']
                    }

            # Step 2: Get today's data from base tables (real-time)
            if has_today:
                if outlet_list:
                    today_staff = await conn.fetch("""
                        SELECT
                            d."AcSalesmanID" as staff_id,
                            s."AcSalesmanName" as staff_name,
                            s."AcSalesmanGroupID" as outlet_id,
                            COUNT(DISTINCT d."DocumentNo") as transactions,
                            COALESCE(SUM(d."ItemTotal"), 0) as total_sales,
                            COALESCE(SUM(CASE WHEN sc."AcStockUDGroup1ID" = 'FLTHB' THEN d."ItemTotal" ELSE 0 END), 0) as house_brand_sales,
                            COALESCE(SUM(CASE WHEN sc."AcStockUDGroup1ID" = 'FLTF1' THEN d."ItemTotal" ELSE 0 END), 0) as focused_1_sales,
                            COALESCE(SUM(CASE WHEN sc."AcStockUDGroup1ID" = 'FLTF2' THEN d."ItemTotal" ELSE 0 END), 0) as focused_2_sales,
                            COALESCE(SUM(CASE WHEN sc."AcStockUDGroup1ID" = 'FLTF3' THEN d."ItemTotal" ELSE 0 END), 0) as focused_3_sales,
                            COALESCE(SUM(CASE WHEN sc."AcStockUDGroup1ID" = 'FLTSC' THEN d."ItemTotal" ELSE 0 END), 0) as clearance_sales
                        FROM "AcCSD" d
                        INNER JOIN "AcCSM" m ON d."DocumentNo" = m."DocumentNo"
                        LEFT JOIN "AcSalesman" s ON d."AcSalesmanID" = s."AcSalesmanID"
                        LEFT JOIN "AcStockCompany" sc ON d."AcStockID" = sc."AcStockID" AND d."AcStockUOMID" = sc."AcStockUOMID"
                        WHERE m."DocumentDate" >= $1 AND m."DocumentDate" < $2
                          AND m."AcLocationID" = ANY($3)
                        GROUP BY d."AcSalesmanID", s."AcSalesmanName", s."AcSalesmanGroupID"
                    """, today_start, today_end, outlet_list)
                else:
                    today_staff = await conn.fetch("""
                        SELECT
                            d."AcSalesmanID" as staff_id,
                            s."AcSalesmanName" as staff_name,
                            s."AcSalesmanGroupID" as outlet_id,
                            COUNT(DISTINCT d."DocumentNo") as transactions,
                            COALESCE(SUM(d."ItemTotal"), 0) as total_sales,
                            COALESCE(SUM(CASE WHEN sc."AcStockUDGroup1ID" = 'FLTHB' THEN d."ItemTotal" ELSE 0 END), 0) as house_brand_sales,
                            COALESCE(SUM(CASE WHEN sc."AcStockUDGroup1ID" = 'FLTF1' THEN d."ItemTotal" ELSE 0 END), 0) as focused_1_sales,
                            COALESCE(SUM(CASE WHEN sc."AcStockUDGroup1ID" = 'FLTF2' THEN d."ItemTotal" ELSE 0 END), 0) as focused_2_sales,
                            COALESCE(SUM(CASE WHEN sc."AcStockUDGroup1ID" = 'FLTF3' THEN d."ItemTotal" ELSE 0 END), 0) as focused_3_sales,
                            COALESCE(SUM(CASE WHEN sc."AcStockUDGroup1ID" = 'FLTSC' THEN d."ItemTotal" ELSE 0 END), 0) as clearance_sales
                        FROM "AcCSD" d
                        INNER JOIN "AcCSM" m ON d."DocumentNo" = m."DocumentNo"
                        LEFT JOIN "AcSalesman" s ON d."AcSalesmanID" = s."AcSalesmanID"
                        LEFT JOIN "AcStockCompany" sc ON d."AcStockID" = sc."AcStockID" AND d."AcStockUOMID" = sc."AcStockUOMID"
                        WHERE m."DocumentDate" >= $1 AND m."DocumentDate" < $2
                        GROUP BY d."AcSalesmanID", s."AcSalesmanName", s."AcSalesmanGroupID"
                    """, today_start, today_end)

                # Get today's PWP data separately
                if outlet_list:
                    today_pwp = await conn.fetch("""
                        SELECT d."AcSalesmanID" as staff_id, COALESCE(SUM(d."ItemTotal"), 0) as pwp_sales
                        FROM "AcCSD" d
                        INNER JOIN "AcCSM" m ON d."DocumentNo" = m."DocumentNo"
                        INNER JOIN "AcCSDPromotionType" pt ON d."DocumentNo" = pt."DocumentNo" AND d."ItemNo" = pt."ItemNo"
                        WHERE pt."AcPromotionSettingID" = 'PURCHASE WITH PURCHASE'
                          AND m."DocumentDate" >= $1 AND m."DocumentDate" < $2
                          AND m."AcLocationID" = ANY($3)
                        GROUP BY d."AcSalesmanID"
                    """, today_start, today_end, outlet_list)
                else:
                    today_pwp = await conn.fetch("""
                        SELECT d."AcSalesmanID" as staff_id, COALESCE(SUM(d."ItemTotal"), 0) as pwp_sales
                        FROM "AcCSD" d
                        INNER JOIN "AcCSM" m ON d."DocumentNo" = m."DocumentNo"
                        INNER JOIN "AcCSDPromotionType" pt ON d."DocumentNo" = pt."DocumentNo" AND d."ItemNo" = pt."ItemNo"
                        WHERE pt."AcPromotionSettingID" = 'PURCHASE WITH PURCHASE'
                          AND m."DocumentDate" >= $1 AND m."DocumentDate" < $2
                        GROUP BY d."AcSalesmanID"
                    """, today_start, today_end)

                pwp_map = {row['staff_id']: float(row['pwp_sales'] or 0) for row in today_pwp}

                # Merge today's data
                for row in today_staff:
                    sid = row['staff_id']
                    if sid in staff_data:
                        staff_data[sid]['transactions'] += int(row['transactions'] or 0)
                        staff_data[sid]['total_sales'] += float(row['total_sales'] or 0)
                        staff_data[sid]['house_brand_sales'] += float(row['house_brand_sales'] or 0)
                        staff_data[sid]['focused_1_sales'] += float(row['focused_1_sales'] or 0)
                        staff_data[sid]['focused_2_sales'] += float(row['focused_2_sales'] or 0)
                        staff_data[sid]['focused_3_sales'] += float(row['focused_3_sales'] or 0)
                        staff_data[sid]['clearance_sales'] += float(row['clearance_sales'] or 0)
                        staff_data[sid]['pwp_sales'] += pwp_map.get(sid, 0)
                    else:
                        staff_data[sid] = {
                            'staff_id': sid,
                            'staff_name': row['staff_name'],
                            'outlet_id': row['outlet_id'],
                            'transactions': int(row['transactions'] or 0),
                            'total_sales': float(row['total_sales'] or 0),
                            'house_brand_sales': float(row['house_brand_sales'] or 0),
                            'focused_1_sales': float(row['focused_1_sales'] or 0),
                            'focused_2_sales': float(row['focused_2_sales'] or 0),
                            'focused_3_sales': float(row['focused_3_sales'] or 0),
                            'pwp_sales': pwp_map.get(sid, 0),
                            'clearance_sales': float(row['clearance_sales'] or 0),
                            'rank': None
                        }

            # Convert to list and sort
            staff = sorted(staff_data.values(), key=lambda x: x['total_sales'], reverse=True)

        elif staff_group:
            # Single outlet with staff group - HYBRID approach
            # Step 1: Get all active staff in this group
            all_staff = await conn.fetch("""
                SELECT s."AcSalesmanID" as staff_id, s."AcSalesmanName" as staff_name,
                       r.outlet_rank_sales as rank
                FROM "AcSalesman" s
                LEFT JOIN analytics.mv_staff_rankings r
                    ON s."AcSalesmanID" = r.staff_id
                    AND r.outlet_id = $1
                    AND r.month = DATE_TRUNC('month', $2::date)
                WHERE s."AcSalesmanGroupID" = $3 AND s."Active" = 'Y'
            """, outlet_id, start_date, staff_group)

            for row in all_staff:
                staff_data[row['staff_id']] = {
                    'staff_id': row['staff_id'],
                    'staff_name': row['staff_name'],
                    'transactions': 0,
                    'total_sales': 0,
                    'house_brand_sales': 0,
                    'focused_1_sales': 0,
                    'focused_2_sales': 0,
                    'focused_3_sales': 0,
                    'pwp_sales': 0,
                    'clearance_sales': 0,
                    'rank': row['rank']
                }

            # Step 2: Get historical data from MV (excluding today)
            if has_historical:
                mv_staff = await conn.fetch("""
                    SELECT
                        k.staff_id,
                        COALESCE(SUM(k.transactions), 0) as transactions,
                        COALESCE(SUM(k.total_sales), 0) as total_sales,
                        COALESCE(SUM(k.house_brand_sales), 0) as house_brand_sales,
                        COALESCE(SUM(k.focused_1_sales), 0) as focused_1_sales,
                        COALESCE(SUM(k.focused_2_sales), 0) as focused_2_sales,
                        COALESCE(SUM(k.focused_3_sales), 0) as focused_3_sales,
                        COALESCE(SUM(k.pwp_sales), 0) as pwp_sales,
                        COALESCE(SUM(k.clearance_sales), 0) as clearance_sales
                    FROM analytics.mv_staff_daily_kpi k
                    WHERE k.outlet_id = $1 AND k.sale_date BETWEEN $2 AND $3
                    GROUP BY k.staff_id
                """, outlet_id, start_date, hist_end)

                for row in mv_staff:
                    sid = row['staff_id']
                    if sid in staff_data:
                        staff_data[sid]['transactions'] += int(row['transactions'] or 0)
                        staff_data[sid]['total_sales'] += float(row['total_sales'] or 0)
                        staff_data[sid]['house_brand_sales'] += float(row['house_brand_sales'] or 0)
                        staff_data[sid]['focused_1_sales'] += float(row['focused_1_sales'] or 0)
                        staff_data[sid]['focused_2_sales'] += float(row['focused_2_sales'] or 0)
                        staff_data[sid]['focused_3_sales'] += float(row['focused_3_sales'] or 0)
                        staff_data[sid]['pwp_sales'] += float(row['pwp_sales'] or 0)
                        staff_data[sid]['clearance_sales'] += float(row['clearance_sales'] or 0)

            # Step 3: Get today's data from base tables (real-time)
            if has_today:
                today_staff = await conn.fetch("""
                    SELECT
                        d."AcSalesmanID" as staff_id,
                        COUNT(DISTINCT d."DocumentNo") as transactions,
                        COALESCE(SUM(d."ItemTotal"), 0) as total_sales,
                        COALESCE(SUM(CASE WHEN sc."AcStockUDGroup1ID" = 'FLTHB' THEN d."ItemTotal" ELSE 0 END), 0) as house_brand_sales,
                        COALESCE(SUM(CASE WHEN sc."AcStockUDGroup1ID" = 'FLTF1' THEN d."ItemTotal" ELSE 0 END), 0) as focused_1_sales,
                        COALESCE(SUM(CASE WHEN sc."AcStockUDGroup1ID" = 'FLTF2' THEN d."ItemTotal" ELSE 0 END), 0) as focused_2_sales,
                        COALESCE(SUM(CASE WHEN sc."AcStockUDGroup1ID" = 'FLTF3' THEN d."ItemTotal" ELSE 0 END), 0) as focused_3_sales,
                        COALESCE(SUM(CASE WHEN sc."AcStockUDGroup1ID" = 'FLTSC' THEN d."ItemTotal" ELSE 0 END), 0) as clearance_sales
                    FROM "AcCSD" d
                    INNER JOIN "AcCSM" m ON d."DocumentNo" = m."DocumentNo"
                    LEFT JOIN "AcStockCompany" sc ON d."AcStockID" = sc."AcStockID" AND d."AcStockUOMID" = sc."AcStockUOMID"
                    WHERE m."AcLocationID" = $1
                      AND m."DocumentDate" >= $2 AND m."DocumentDate" < $3
                    GROUP BY d."AcSalesmanID"
                """, outlet_id, today_start, today_end)

                # Get today's PWP
                today_pwp = await conn.fetch("""
                    SELECT d."AcSalesmanID" as staff_id, COALESCE(SUM(d."ItemTotal"), 0) as pwp_sales
                    FROM "AcCSD" d
                    INNER JOIN "AcCSM" m ON d."DocumentNo" = m."DocumentNo"
                    INNER JOIN "AcCSDPromotionType" pt ON d."DocumentNo" = pt."DocumentNo" AND d."ItemNo" = pt."ItemNo"
                    WHERE pt."AcPromotionSettingID" = 'PURCHASE WITH PURCHASE'
                      AND m."AcLocationID" = $1
                      AND m."DocumentDate" >= $2 AND m."DocumentDate" < $3
                    GROUP BY d."AcSalesmanID"
                """, outlet_id, today_start, today_end)

                pwp_map = {row['staff_id']: float(row['pwp_sales'] or 0) for row in today_pwp}

                for row in today_staff:
                    sid = row['staff_id']
                    if sid in staff_data:
                        staff_data[sid]['transactions'] += int(row['transactions'] or 0)
                        staff_data[sid]['total_sales'] += float(row['total_sales'] or 0)
                        staff_data[sid]['house_brand_sales'] += float(row['house_brand_sales'] or 0)
                        staff_data[sid]['focused_1_sales'] += float(row['focused_1_sales'] or 0)
                        staff_data[sid]['focused_2_sales'] += float(row['focused_2_sales'] or 0)
                        staff_data[sid]['focused_3_sales'] += float(row['focused_3_sales'] or 0)
                        staff_data[sid]['clearance_sales'] += float(row['clearance_sales'] or 0)
                        staff_data[sid]['pwp_sales'] += pwp_map.get(sid, 0)

            # Convert to list and sort
            staff = sorted(staff_data.values(), key=lambda x: x['total_sales'], reverse=True)

        return {
            "success": True,
            "data": {
                "outlet_id": outlet_id or "ALL",
                "group_id": staff_group,
                "outlet_name": outlet_name,
                "view_all": view_all,
                "period": {
                    "start": start_date.isoformat(),
                    "end": end_date.isoformat()
                },
                "summary": {
                    "total_sales": round(summary['total_sales'], 2),
                    "gross_profit": round(summary['gross_profit'], 2),
                    "house_brand": round(summary['house_brand'], 2),
                    "focused_1": round(summary['focused_1'], 2),
                    "focused_2": round(summary['focused_2'], 2),
                    "focused_3": round(summary['focused_3'], 2),
                    "pwp": round(summary['pwp'], 2),
                    "clearance": round(summary['clearance'], 2),
                    "transactions": summary['transactions'],
                    "staff_count": len(staff)
                },
                "staff": [
                    {
                        "staff_id": row['staff_id'],
                        "staff_name": row['staff_name'] or "Unknown",
                        "outlet_id": row.get('outlet_id') if view_all else None,  # Show outlet for all-outlets view
                        "total_sales": float(row['total_sales'] or 0),
                        "house_brand": float(row['house_brand_sales'] or 0),
                        "focused_1": float(row['focused_1_sales'] or 0),
                        "focused_2": float(row['focused_2_sales'] or 0),
                        "focused_3": float(row['focused_3_sales'] or 0),
                        "pwp": float(row['pwp_sales'] or 0),
                        "clearance": float(row['clearance_sales'] or 0),
                        "transactions": int(row['transactions'] or 0),
                        "rank": row['rank']
                    }
                    for row in staff
                ]
            }
        }


@app.get("/api/v1/kpi/team/export")
async def export_team_performance(
    outlet_id: Optional[str] = Query(None, description="Outlet ID or 'ALL' for all outlets"),
    outlet_ids: Optional[str] = Query(None, description="Comma-separated list of allowed outlet IDs"),
    start_date: Optional[str] = Query(None, description="Start date YYYY-MM-DD"),
    end_date: Optional[str] = Query(None, description="End date YYYY-MM-DD")
):
    """Export staff performance to Excel file.

    For Admin/OOM: Can export all outlets or specific outlet
    Returns Excel file with all staff KPI data
    """
    if not EXCEL_AVAILABLE:
        raise HTTPException(status_code=500, detail="Excel export not available - openpyxl not installed")

    # Parse dates
    today = date.today()
    if start_date:
        try:
            start = datetime.strptime(start_date, "%Y-%m-%d").date()
        except:
            start = date(today.year, today.month, 1)
    else:
        start = date(today.year, today.month, 1)

    if end_date:
        try:
            end = datetime.strptime(end_date, "%Y-%m-%d").date()
        except:
            end = today
    else:
        end = today

    view_all = not outlet_id or outlet_id.upper() == 'ALL'
    outlet_list = [o.strip() for o in outlet_ids.split(',')] if outlet_ids else None

    try:
        async with pool.acquire() as conn:
            if view_all:
                if outlet_list:
                    staff = await conn.fetch("""
                        SELECT
                            k.staff_id,
                            s."AcSalesmanName" as staff_name,
                            s."AcSalesmanGroupID" as outlet_id,
                            l."AcLocationDesc" as outlet_name,
                            COALESCE(SUM(k.transactions), 0) as transactions,
                            COALESCE(SUM(k.total_sales), 0) as total_sales,
                            COALESCE(SUM(k.gross_profit), 0) as gross_profit,
                            COALESCE(SUM(k.house_brand_sales), 0) as house_brand_sales,
                            COALESCE(SUM(k.focused_1_sales), 0) as focused_1_sales,
                            COALESCE(SUM(k.focused_2_sales), 0) as focused_2_sales,
                            COALESCE(SUM(k.focused_3_sales), 0) as focused_3_sales,
                            COALESCE(SUM(k.pwp_sales), 0) as pwp_sales,
                            COALESCE(SUM(k.clearance_sales), 0) as clearance_sales,
                            r.company_rank_sales as rank
                        FROM analytics.mv_staff_daily_kpi k
                        LEFT JOIN "AcSalesman" s ON k.staff_id = s."AcSalesmanID"
                        LEFT JOIN "AcLocation" l ON s."AcSalesmanGroupID" = l."AcLocationID"
                        LEFT JOIN analytics.mv_staff_rankings r
                            ON k.staff_id = r.staff_id
                            AND r.month = DATE_TRUNC('month', $1::date)
                            AND r.outlet_id = k.outlet_id
                        WHERE k.sale_date BETWEEN $1 AND $2
                          AND k.outlet_id = ANY($3)
                        GROUP BY k.staff_id, s."AcSalesmanName", s."AcSalesmanGroupID", l."AcLocationDesc", r.company_rank_sales
                        ORDER BY COALESCE(SUM(k.total_sales), 0) DESC
                    """, start, end, outlet_list)
                else:
                    staff = await conn.fetch("""
                        SELECT
                            k.staff_id,
                            s."AcSalesmanName" as staff_name,
                            s."AcSalesmanGroupID" as outlet_id,
                            l."AcLocationDesc" as outlet_name,
                            COALESCE(SUM(k.transactions), 0) as transactions,
                            COALESCE(SUM(k.total_sales), 0) as total_sales,
                            COALESCE(SUM(k.gross_profit), 0) as gross_profit,
                            COALESCE(SUM(k.house_brand_sales), 0) as house_brand_sales,
                            COALESCE(SUM(k.focused_1_sales), 0) as focused_1_sales,
                            COALESCE(SUM(k.focused_2_sales), 0) as focused_2_sales,
                            COALESCE(SUM(k.focused_3_sales), 0) as focused_3_sales,
                            COALESCE(SUM(k.pwp_sales), 0) as pwp_sales,
                            COALESCE(SUM(k.clearance_sales), 0) as clearance_sales,
                            r.company_rank_sales as rank
                        FROM analytics.mv_staff_daily_kpi k
                        LEFT JOIN "AcSalesman" s ON k.staff_id = s."AcSalesmanID"
                        LEFT JOIN "AcLocation" l ON s."AcSalesmanGroupID" = l."AcLocationID"
                        LEFT JOIN analytics.mv_staff_rankings r
                            ON k.staff_id = r.staff_id
                            AND r.month = DATE_TRUNC('month', $1::date)
                            AND r.outlet_id = k.outlet_id
                        WHERE k.sale_date BETWEEN $1 AND $2
                        GROUP BY k.staff_id, s."AcSalesmanName", s."AcSalesmanGroupID", l."AcLocationDesc", r.company_rank_sales
                        ORDER BY COALESCE(SUM(k.total_sales), 0) DESC
                    """, start, end)
            else:
                # Single outlet
                staff = await conn.fetch("""
                    SELECT
                        k.staff_id,
                        s."AcSalesmanName" as staff_name,
                        $1 as outlet_id,
                        l."AcLocationDesc" as outlet_name,
                        COALESCE(SUM(k.transactions), 0) as transactions,
                        COALESCE(SUM(k.total_sales), 0) as total_sales,
                        COALESCE(SUM(k.gross_profit), 0) as gross_profit,
                        COALESCE(SUM(k.house_brand_sales), 0) as house_brand_sales,
                        COALESCE(SUM(k.focused_1_sales), 0) as focused_1_sales,
                        COALESCE(SUM(k.focused_2_sales), 0) as focused_2_sales,
                        COALESCE(SUM(k.focused_3_sales), 0) as focused_3_sales,
                        COALESCE(SUM(k.pwp_sales), 0) as pwp_sales,
                        COALESCE(SUM(k.clearance_sales), 0) as clearance_sales,
                        r.outlet_rank_sales as rank
                    FROM analytics.mv_staff_daily_kpi k
                    LEFT JOIN "AcSalesman" s ON k.staff_id = s."AcSalesmanID"
                    LEFT JOIN "AcLocation" l ON $1 = l."AcLocationID"
                    LEFT JOIN analytics.mv_staff_rankings r
                        ON k.staff_id = r.staff_id
                        AND r.outlet_id = $1
                        AND r.month = DATE_TRUNC('month', $2::date)
                    WHERE k.sale_date BETWEEN $2 AND $3
                      AND k.outlet_id = $1
                    GROUP BY k.staff_id, s."AcSalesmanName", l."AcLocationDesc", r.outlet_rank_sales
                    ORDER BY COALESCE(SUM(k.total_sales), 0) DESC
                """, outlet_id, start, end)

        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Staff Performance"

        # Headers
        headers = [
            "Rank", "Staff ID", "Staff Name", "Outlet ID", "Outlet Name",
            "Total Sales (RM)", "Gross Profit (RM)", "House Brand (RM)",
            "Focused Item 1 (RM)", "Focused Item 2 (RM)", "Focused Item 3 (RM)",
            "PWP (RM)", "Stock Clearance (RM)", "Transactions"
        ]
        ws.append(headers)

        # Style headers
        from openpyxl.styles import Font, PatternFill, Alignment
        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for col, cell in enumerate(ws[1], 1):
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        for row in staff:
            ws.append([
                row['rank'] or "-",
                row['staff_id'],
                row['staff_name'] or "Unknown",
                row['outlet_id'] or "-",
                row['outlet_name'] or "-",
                round(float(row['total_sales'] or 0), 2),
                round(float(row['gross_profit'] or 0), 2),
                round(float(row['house_brand_sales'] or 0), 2),
                round(float(row['focused_1_sales'] or 0), 2),
                round(float(row['focused_2_sales'] or 0), 2),
                round(float(row['focused_3_sales'] or 0), 2),
                round(float(row['pwp_sales'] or 0), 2),
                round(float(row['clearance_sales'] or 0), 2),
                int(row['transactions'] or 0)
            ])

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Format number columns
        from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1
        for row in ws.iter_rows(min_row=2, min_col=6, max_col=13):
            for cell in row:
                cell.number_format = '#,##0.00'

        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        # Generate filename
        outlet_label = "all_outlets" if view_all else outlet_id
        filename = f"staff_performance_{outlet_label}_{start.strftime('%Y%m%d')}_{end.strftime('%Y%m%d')}.xlsx"

        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Export failed: {str(e)}")


@app.get("/api/v1/kpi/outlets")
async def get_outlet_performance(
    outlet_ids: Optional[str] = Query(None, description="Comma-separated list of outlet IDs (empty for all)"),
    start_date: Optional[str] = Query(None, description="Start date YYYY-MM-DD"),
    end_date: Optional[str] = Query(None, description="End date YYYY-MM-DD")
):
    """Get performance summary for each outlet - HYBRID: MV for history + real-time for today.

    Returns KPI metrics aggregated per outlet for comparison.
    """
    # Parse dates
    today = date.today()
    today_start = datetime.combine(today, datetime.min.time())
    today_end = datetime.combine(today + timedelta(days=1), datetime.min.time())

    if start_date:
        try:
            start = datetime.strptime(start_date, "%Y-%m-%d").date()
        except:
            start = date(today.year, today.month, 1)
    else:
        start = date(today.year, today.month, 1)

    if end_date:
        try:
            end = datetime.strptime(end_date, "%Y-%m-%d").date()
        except:
            end = today
    else:
        end = today

    outlet_list = [o.strip() for o in outlet_ids.split(',')] if outlet_ids else None

    try:
        async with pool.acquire() as conn:
            # HYBRID APPROACH: MV for historical data (excluding today) + real-time for today
            # Step 1: Get historical data from MV (excluding today)
            if outlet_list:
                mv_outlets = await conn.fetch("""
                    SELECT
                        k.outlet_id,
                        l."AcLocationDesc" as outlet_name,
                        COALESCE((
                            SELECT COUNT(DISTINCT staff_id)
                            FROM analytics.mv_staff_daily_kpi
                            WHERE outlet_id = k.outlet_id AND sale_date BETWEEN $1 AND $2
                        ), 0) as staff_count,
                        COALESCE(SUM(k.transactions), 0) as transactions,
                        COALESCE(SUM(k.total_sales), 0) as total_sales,
                        COALESCE(SUM(k.gross_profit), 0) as gross_profit,
                        COALESCE(SUM(k.house_brand_sales), 0) as house_brand,
                        COALESCE(SUM(k.focused_1_sales), 0) as focused_1,
                        COALESCE(SUM(k.focused_2_sales), 0) as focused_2,
                        COALESCE(SUM(k.focused_3_sales), 0) as focused_3,
                        COALESCE(SUM(k.pwp_sales), 0) as pwp,
                        COALESCE(SUM(k.clearance_sales), 0) as clearance
                    FROM analytics.mv_outlet_daily_kpi k
                    LEFT JOIN "AcLocation" l ON k.outlet_id = l."AcLocationID"
                    WHERE k.sale_date BETWEEN $1 AND $2
                      AND k.sale_date < $4
                      AND k.outlet_id = ANY($3)
                    GROUP BY k.outlet_id, l."AcLocationDesc"
                """, start, end, outlet_list, today)
            else:
                mv_outlets = await conn.fetch("""
                    SELECT
                        k.outlet_id,
                        l."AcLocationDesc" as outlet_name,
                        COALESCE((
                            SELECT COUNT(DISTINCT staff_id)
                            FROM analytics.mv_staff_daily_kpi
                            WHERE outlet_id = k.outlet_id AND sale_date BETWEEN $1 AND $2
                        ), 0) as staff_count,
                        COALESCE(SUM(k.transactions), 0) as transactions,
                        COALESCE(SUM(k.total_sales), 0) as total_sales,
                        COALESCE(SUM(k.gross_profit), 0) as gross_profit,
                        COALESCE(SUM(k.house_brand_sales), 0) as house_brand,
                        COALESCE(SUM(k.focused_1_sales), 0) as focused_1,
                        COALESCE(SUM(k.focused_2_sales), 0) as focused_2,
                        COALESCE(SUM(k.focused_3_sales), 0) as focused_3,
                        COALESCE(SUM(k.pwp_sales), 0) as pwp,
                        COALESCE(SUM(k.clearance_sales), 0) as clearance
                    FROM analytics.mv_outlet_daily_kpi k
                    LEFT JOIN "AcLocation" l ON k.outlet_id = l."AcLocationID"
                    WHERE k.sale_date BETWEEN $1 AND $2
                      AND k.sale_date < $3
                    GROUP BY k.outlet_id, l."AcLocationDesc"
                """, start, end, today)

            # Convert to dict for easy merging
            outlet_data = {}
            for o in mv_outlets:
                outlet_data[o['outlet_id']] = {
                    'outlet_id': o['outlet_id'],
                    'outlet_name': o['outlet_name'] or o['outlet_id'],
                    'staff_count': int(o['staff_count'] or 0),
                    'transactions': int(o['transactions'] or 0),
                    'total_sales': float(o['total_sales'] or 0),
                    'gross_profit': float(o['gross_profit'] or 0),
                    'house_brand': float(o['house_brand'] or 0),
                    'focused_1': float(o['focused_1'] or 0),
                    'focused_2': float(o['focused_2'] or 0),
                    'focused_3': float(o['focused_3'] or 0),
                    'pwp': float(o['pwp'] or 0),
                    'clearance': float(o['clearance'] or 0),
                }

            # Step 2: Get today's real-time data per outlet (if date range includes today)
            if end >= today:
                # Cash sales per outlet
                if outlet_list:
                    today_cash = await conn.fetch("""
                        SELECT
                            m."AcLocationID" as outlet_id,
                            COUNT(DISTINCT m."DocumentNo") as transactions,
                            COALESCE(SUM(d."ItemTotal"), 0) as total_sales,
                            COALESCE(SUM(d."ItemTotal" - COALESCE(d."ItemCost" * d."ItemQuantity", 0)), 0) as gross_profit,
                            COALESCE(SUM(CASE WHEN s."AcStockUDGroup1ID" = 'FLTHB' THEN d."ItemTotal" ELSE 0 END), 0) as house_brand,
                            COALESCE(SUM(CASE WHEN s."AcStockUDGroup1ID" = 'FLTF1' THEN d."ItemTotal" ELSE 0 END), 0) as focused_1,
                            COALESCE(SUM(CASE WHEN s."AcStockUDGroup1ID" = 'FLTF2' THEN d."ItemTotal" ELSE 0 END), 0) as focused_2,
                            COALESCE(SUM(CASE WHEN s."AcStockUDGroup1ID" = 'FLTF3' THEN d."ItemTotal" ELSE 0 END), 0) as focused_3,
                            COALESCE(SUM(CASE WHEN s."AcStockUDGroup1ID" = 'STOCK CLEARANCE' THEN d."ItemTotal" ELSE 0 END), 0) as clearance
                        FROM "AcCSM" m
                        INNER JOIN "AcCSD" d ON m."DocumentNo" = d."DocumentNo"
                        LEFT JOIN "AcStockCompany" s ON d."AcStockID" = s."AcStockID" AND d."AcStockUOMID" = s."AcStockUOMID"
                        WHERE m."DocumentDate" >= $1 AND m."DocumentDate" < $2
                          AND m."AcLocationID" = ANY($3)
                        GROUP BY m."AcLocationID"
                    """, today_start, today_end, outlet_list)
                else:
                    today_cash = await conn.fetch("""
                        SELECT
                            m."AcLocationID" as outlet_id,
                            COUNT(DISTINCT m."DocumentNo") as transactions,
                            COALESCE(SUM(d."ItemTotal"), 0) as total_sales,
                            COALESCE(SUM(d."ItemTotal" - COALESCE(d."ItemCost" * d."ItemQuantity", 0)), 0) as gross_profit,
                            COALESCE(SUM(CASE WHEN s."AcStockUDGroup1ID" = 'FLTHB' THEN d."ItemTotal" ELSE 0 END), 0) as house_brand,
                            COALESCE(SUM(CASE WHEN s."AcStockUDGroup1ID" = 'FLTF1' THEN d."ItemTotal" ELSE 0 END), 0) as focused_1,
                            COALESCE(SUM(CASE WHEN s."AcStockUDGroup1ID" = 'FLTF2' THEN d."ItemTotal" ELSE 0 END), 0) as focused_2,
                            COALESCE(SUM(CASE WHEN s."AcStockUDGroup1ID" = 'FLTF3' THEN d."ItemTotal" ELSE 0 END), 0) as focused_3,
                            COALESCE(SUM(CASE WHEN s."AcStockUDGroup1ID" = 'STOCK CLEARANCE' THEN d."ItemTotal" ELSE 0 END), 0) as clearance
                        FROM "AcCSM" m
                        INNER JOIN "AcCSD" d ON m."DocumentNo" = d."DocumentNo"
                        LEFT JOIN "AcStockCompany" s ON d."AcStockID" = s."AcStockID" AND d."AcStockUOMID" = s."AcStockUOMID"
                        WHERE m."DocumentDate" >= $1 AND m."DocumentDate" < $2
                        GROUP BY m."AcLocationID"
                    """, today_start, today_end)

                # Merge today's cash sales
                for row in today_cash:
                    oid = row['outlet_id']
                    if oid not in outlet_data:
                        # Get outlet name
                        loc = await conn.fetchrow('SELECT "AcLocationDesc" FROM "AcLocation" WHERE "AcLocationID" = $1', oid)
                        outlet_data[oid] = {
                            'outlet_id': oid,
                            'outlet_name': loc['AcLocationDesc'] if loc else oid,
                            'staff_count': 0,
                            'transactions': 0, 'total_sales': 0, 'gross_profit': 0,
                            'house_brand': 0, 'focused_1': 0, 'focused_2': 0, 'focused_3': 0, 'pwp': 0, 'clearance': 0
                        }
                    outlet_data[oid]['transactions'] += int(row['transactions'] or 0)
                    outlet_data[oid]['total_sales'] += float(row['total_sales'] or 0)
                    outlet_data[oid]['gross_profit'] += float(row['gross_profit'] or 0)
                    outlet_data[oid]['house_brand'] += float(row['house_brand'] or 0)
                    outlet_data[oid]['focused_1'] += float(row['focused_1'] or 0)
                    outlet_data[oid]['focused_2'] += float(row['focused_2'] or 0)
                    outlet_data[oid]['focused_3'] += float(row['focused_3'] or 0)
                    outlet_data[oid]['clearance'] += float(row['clearance'] or 0)

                # Invoice sales per outlet
                if outlet_list:
                    today_inv = await conn.fetch("""
                        SELECT
                            m."AcLocationID" as outlet_id,
                            COUNT(DISTINCT m."AcCusInvoiceMID") as transactions,
                            COALESCE(SUM(d."ItemTotalPrice"), 0) as total_sales,
                            COALESCE(SUM(CASE WHEN s."AcStockUDGroup1ID" = 'FLTHB' THEN d."ItemTotalPrice" ELSE 0 END), 0) as house_brand,
                            COALESCE(SUM(CASE WHEN s."AcStockUDGroup1ID" = 'FLTF1' THEN d."ItemTotalPrice" ELSE 0 END), 0) as focused_1,
                            COALESCE(SUM(CASE WHEN s."AcStockUDGroup1ID" = 'FLTF2' THEN d."ItemTotalPrice" ELSE 0 END), 0) as focused_2,
                            COALESCE(SUM(CASE WHEN s."AcStockUDGroup1ID" = 'FLTF3' THEN d."ItemTotalPrice" ELSE 0 END), 0) as focused_3,
                            COALESCE(SUM(CASE WHEN s."AcStockUDGroup1ID" = 'STOCK CLEARANCE' THEN d."ItemTotalPrice" ELSE 0 END), 0) as clearance
                        FROM "AcCusInvoiceM" m
                        INNER JOIN "AcCusInvoiceD" d ON m."AcCusInvoiceMID" = d."AcCusInvoiceMID"
                        LEFT JOIN "AcStockCompany" s ON d."AcStockID" = s."AcStockID" AND d."AcStockUOMID" = s."AcStockUOMID"
                        WHERE m."DocumentDate" >= $1 AND m."DocumentDate" < $2
                          AND m."AcLocationID" = ANY($3)
                        GROUP BY m."AcLocationID"
                    """, today_start, today_end, outlet_list)
                else:
                    today_inv = await conn.fetch("""
                        SELECT
                            m."AcLocationID" as outlet_id,
                            COUNT(DISTINCT m."AcCusInvoiceMID") as transactions,
                            COALESCE(SUM(d."ItemTotalPrice"), 0) as total_sales,
                            COALESCE(SUM(CASE WHEN s."AcStockUDGroup1ID" = 'FLTHB' THEN d."ItemTotalPrice" ELSE 0 END), 0) as house_brand,
                            COALESCE(SUM(CASE WHEN s."AcStockUDGroup1ID" = 'FLTF1' THEN d."ItemTotalPrice" ELSE 0 END), 0) as focused_1,
                            COALESCE(SUM(CASE WHEN s."AcStockUDGroup1ID" = 'FLTF2' THEN d."ItemTotalPrice" ELSE 0 END), 0) as focused_2,
                            COALESCE(SUM(CASE WHEN s."AcStockUDGroup1ID" = 'FLTF3' THEN d."ItemTotalPrice" ELSE 0 END), 0) as focused_3,
                            COALESCE(SUM(CASE WHEN s."AcStockUDGroup1ID" = 'STOCK CLEARANCE' THEN d."ItemTotalPrice" ELSE 0 END), 0) as clearance
                        FROM "AcCusInvoiceM" m
                        INNER JOIN "AcCusInvoiceD" d ON m."AcCusInvoiceMID" = d."AcCusInvoiceMID"
                        LEFT JOIN "AcStockCompany" s ON d."AcStockID" = s."AcStockID" AND d."AcStockUOMID" = s."AcStockUOMID"
                        WHERE m."DocumentDate" >= $1 AND m."DocumentDate" < $2
                        GROUP BY m."AcLocationID"
                    """, today_start, today_end)

                # Merge today's invoice sales
                for row in today_inv:
                    oid = row['outlet_id']
                    if oid not in outlet_data:
                        loc = await conn.fetchrow('SELECT "AcLocationDesc" FROM "AcLocation" WHERE "AcLocationID" = $1', oid)
                        outlet_data[oid] = {
                            'outlet_id': oid,
                            'outlet_name': loc['AcLocationDesc'] if loc else oid,
                            'staff_count': 0,
                            'transactions': 0, 'total_sales': 0, 'gross_profit': 0,
                            'house_brand': 0, 'focused_1': 0, 'focused_2': 0, 'focused_3': 0, 'pwp': 0, 'clearance': 0
                        }
                    outlet_data[oid]['transactions'] += int(row['transactions'] or 0)
                    outlet_data[oid]['total_sales'] += float(row['total_sales'] or 0)
                    outlet_data[oid]['house_brand'] += float(row['house_brand'] or 0)
                    outlet_data[oid]['focused_1'] += float(row['focused_1'] or 0)
                    outlet_data[oid]['focused_2'] += float(row['focused_2'] or 0)
                    outlet_data[oid]['focused_3'] += float(row['focused_3'] or 0)
                    outlet_data[oid]['clearance'] += float(row['clearance'] or 0)

                # PWP per outlet
                if outlet_list:
                    today_pwp = await conn.fetch("""
                        SELECT m."AcLocationID" as outlet_id, COALESCE(SUM(d."ItemTotal"), 0) as pwp
                        FROM "AcCSD" d
                        INNER JOIN "AcCSM" m ON d."DocumentNo" = m."DocumentNo"
                        INNER JOIN "AcCSDPromotionType" pt ON d."DocumentNo" = pt."DocumentNo" AND d."ItemNo" = pt."ItemNo"
                        WHERE pt."AcPromotionSettingID" = 'PURCHASE WITH PURCHASE'
                          AND m."DocumentDate" >= $1 AND m."DocumentDate" < $2
                          AND m."AcLocationID" = ANY($3)
                        GROUP BY m."AcLocationID"
                    """, today_start, today_end, outlet_list)
                else:
                    today_pwp = await conn.fetch("""
                        SELECT m."AcLocationID" as outlet_id, COALESCE(SUM(d."ItemTotal"), 0) as pwp
                        FROM "AcCSD" d
                        INNER JOIN "AcCSM" m ON d."DocumentNo" = m."DocumentNo"
                        INNER JOIN "AcCSDPromotionType" pt ON d."DocumentNo" = pt."DocumentNo" AND d."ItemNo" = pt."ItemNo"
                        WHERE pt."AcPromotionSettingID" = 'PURCHASE WITH PURCHASE'
                          AND m."DocumentDate" >= $1 AND m."DocumentDate" < $2
                        GROUP BY m."AcLocationID"
                    """, today_start, today_end)

                # Merge PWP
                for row in today_pwp:
                    oid = row['outlet_id']
                    if oid in outlet_data:
                        outlet_data[oid]['pwp'] += float(row['pwp'] or 0)

            # Sort by total_sales descending
            outlets_list = sorted(outlet_data.values(), key=lambda x: x['total_sales'], reverse=True)

            # Calculate totals
            total_sales = sum(o['total_sales'] for o in outlets_list)
            total_gp = sum(o['gross_profit'] for o in outlets_list)
            total_hb = sum(o['house_brand'] for o in outlets_list)
            total_f1 = sum(o['focused_1'] for o in outlets_list)
            total_f2 = sum(o['focused_2'] for o in outlets_list)
            total_f3 = sum(o['focused_3'] for o in outlets_list)
            total_pwp = sum(o['pwp'] for o in outlets_list)
            total_clearance = sum(o['clearance'] for o in outlets_list)
            total_txn = sum(o['transactions'] for o in outlets_list)
            total_staff = sum(o['staff_count'] for o in outlets_list)

            return {
                "success": True,
                "data": {
                    "period": {"start": start.isoformat(), "end": end.isoformat()},
                    "summary": {
                        "outlet_count": len(outlets_list),
                        "staff_count": total_staff,
                        "total_sales": round(total_sales, 2),
                        "gross_profit": round(total_gp, 2),
                        "house_brand": round(total_hb, 2),
                        "focused_1": round(total_f1, 2),
                        "focused_2": round(total_f2, 2),
                        "focused_3": round(total_f3, 2),
                        "pwp": round(total_pwp, 2),
                        "clearance": round(total_clearance, 2),
                        "transactions": total_txn
                    },
                    "outlets": [
                        {
                            "outlet_id": o['outlet_id'],
                            "outlet_name": o['outlet_name'],
                            "staff_count": o['staff_count'],
                            "total_sales": round(o['total_sales'], 2),
                            "gross_profit": round(o['gross_profit'], 2),
                            "house_brand": round(o['house_brand'], 2),
                            "focused_1": round(o['focused_1'], 2),
                            "focused_2": round(o['focused_2'], 2),
                            "focused_3": round(o['focused_3'], 2),
                            "pwp": round(o['pwp'], 2),
                            "clearance": round(o['clearance'], 2),
                            "transactions": o['transactions'],
                            "rank": idx + 1
                        }
                        for idx, o in enumerate(outlets_list)
                    ]
                }
            }

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error fetching outlet performance: {str(e)}")


@app.get("/api/v1/kpi/outlets/export")
async def export_outlet_performance(
    outlet_ids: Optional[str] = Query(None, description="Comma-separated list of outlet IDs (empty for all)"),
    start_date: Optional[str] = Query(None, description="Start date YYYY-MM-DD"),
    end_date: Optional[str] = Query(None, description="End date YYYY-MM-DD")
):
    """Export outlet performance to Excel file."""
    if not EXCEL_AVAILABLE:
        raise HTTPException(status_code=500, detail="Excel export not available - openpyxl not installed")

    # Parse dates
    today = date.today()
    if start_date:
        try:
            start = datetime.strptime(start_date, "%Y-%m-%d").date()
        except:
            start = date(today.year, today.month, 1)
    else:
        start = date(today.year, today.month, 1)

    if end_date:
        try:
            end = datetime.strptime(end_date, "%Y-%m-%d").date()
        except:
            end = today
    else:
        end = today

    outlet_list = [o.strip() for o in outlet_ids.split(',')] if outlet_ids else None

    try:
        async with pool.acquire() as conn:
            # Use mv_outlet_daily_kpi (includes ALL sales) for consistent figures with Dynamod
            if outlet_list:
                outlets = await conn.fetch("""
                    SELECT
                        k.outlet_id,
                        l."AcLocationDesc" as outlet_name,
                        COALESCE((
                            SELECT COUNT(DISTINCT staff_id)
                            FROM analytics.mv_staff_daily_kpi
                            WHERE outlet_id = k.outlet_id AND sale_date BETWEEN $1 AND $2
                        ), 0) as staff_count,
                        COALESCE(SUM(k.transactions), 0) as transactions,
                        COALESCE(SUM(k.total_sales), 0) as total_sales,
                        COALESCE(SUM(k.gross_profit), 0) as gross_profit,
                        COALESCE(SUM(k.house_brand_sales), 0) as house_brand,
                        COALESCE(SUM(k.focused_1_sales), 0) as focused_1,
                        COALESCE(SUM(k.focused_2_sales), 0) as focused_2,
                        COALESCE(SUM(k.focused_3_sales), 0) as focused_3,
                        COALESCE(SUM(k.pwp_sales), 0) as pwp,
                        COALESCE(SUM(k.clearance_sales), 0) as clearance
                    FROM analytics.mv_outlet_daily_kpi k
                    LEFT JOIN "AcLocation" l ON k.outlet_id = l."AcLocationID"
                    WHERE k.sale_date BETWEEN $1 AND $2
                      AND k.outlet_id = ANY($3)
                    GROUP BY k.outlet_id, l."AcLocationDesc"
                    ORDER BY COALESCE(SUM(k.total_sales), 0) DESC
                """, start, end, outlet_list)
            else:
                outlets = await conn.fetch("""
                    SELECT
                        k.outlet_id,
                        l."AcLocationDesc" as outlet_name,
                        COALESCE((
                            SELECT COUNT(DISTINCT staff_id)
                            FROM analytics.mv_staff_daily_kpi
                            WHERE outlet_id = k.outlet_id AND sale_date BETWEEN $1 AND $2
                        ), 0) as staff_count,
                        COALESCE(SUM(k.transactions), 0) as transactions,
                        COALESCE(SUM(k.total_sales), 0) as total_sales,
                        COALESCE(SUM(k.gross_profit), 0) as gross_profit,
                        COALESCE(SUM(k.house_brand_sales), 0) as house_brand,
                        COALESCE(SUM(k.focused_1_sales), 0) as focused_1,
                        COALESCE(SUM(k.focused_2_sales), 0) as focused_2,
                        COALESCE(SUM(k.focused_3_sales), 0) as focused_3,
                        COALESCE(SUM(k.pwp_sales), 0) as pwp,
                        COALESCE(SUM(k.clearance_sales), 0) as clearance
                    FROM analytics.mv_outlet_daily_kpi k
                    LEFT JOIN "AcLocation" l ON k.outlet_id = l."AcLocationID"
                    WHERE k.sale_date BETWEEN $1 AND $2
                    GROUP BY k.outlet_id, l."AcLocationDesc"
                    ORDER BY COALESCE(SUM(k.total_sales), 0) DESC
                """, start, end)

        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Outlet Performance"

        # Header row
        headers = [
            "Rank", "Outlet ID", "Outlet Name", "Staff Count",
            "Total Sales", "Gross Profit", "House Brand",
            "Focused 1", "Focused 2", "Focused 3",
            "PWP", "Clearance", "Transactions"
        ]
        ws.append(headers)

        # Style headers
        from openpyxl.styles import Font, PatternFill
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill

        # Data rows
        for idx, row in enumerate(outlets):
            ws.append([
                idx + 1,
                row['outlet_id'],
                row['outlet_name'] or row['outlet_id'],
                int(row['staff_count'] or 0),
                round(float(row['total_sales'] or 0), 2),
                round(float(row['gross_profit'] or 0), 2),
                round(float(row['house_brand'] or 0), 2),
                round(float(row['focused_1'] or 0), 2),
                round(float(row['focused_2'] or 0), 2),
                round(float(row['focused_3'] or 0), 2),
                round(float(row['pwp'] or 0), 2),
                round(float(row['clearance'] or 0), 2),
                int(row['transactions'] or 0)
            ])

        # Add totals row
        total_row = [
            "",
            "TOTAL",
            f"{len(outlets)} outlets",
            sum(int(o['staff_count'] or 0) for o in outlets),
            round(sum(float(o['total_sales'] or 0) for o in outlets), 2),
            round(sum(float(o['gross_profit'] or 0) for o in outlets), 2),
            round(sum(float(o['house_brand'] or 0) for o in outlets), 2),
            round(sum(float(o['focused_1'] or 0) for o in outlets), 2),
            round(sum(float(o['focused_2'] or 0) for o in outlets), 2),
            round(sum(float(o['focused_3'] or 0) for o in outlets), 2),
            round(sum(float(o['pwp'] or 0) for o in outlets), 2),
            round(sum(float(o['clearance'] or 0) for o in outlets), 2),
            sum(int(o['transactions'] or 0) for o in outlets)
        ]
        ws.append(total_row)

        # Style totals row
        total_font = Font(bold=True)
        for cell in ws[ws.max_row]:
            cell.font = total_font

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Format number columns
        from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1
        for row in ws.iter_rows(min_row=2, min_col=5, max_col=12):
            for cell in row:
                cell.number_format = '#,##0.00'

        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        # Generate filename
        filename = f"outlet_performance_{start.strftime('%Y%m%d')}_{end.strftime('%Y%m%d')}.xlsx"

        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Export failed: {str(e)}")


# ============================================================================
# Target Management Endpoints
# ============================================================================

@app.get("/api/v1/targets/me")
async def get_my_targets(
    staff_id: str = Query(..., description="Staff ID"),
    month: Optional[str] = Query(None, description="Month in YYYY-MM format")
):
    """Get staff's targets with current progress."""
    if month:
        try:
            period = datetime.strptime(month, "%Y-%m")
            year_month = int(period.strftime("%Y%m"))
        except:
            raise HTTPException(status_code=400, detail="Invalid month format")
    else:
        period = datetime.now()
        year_month = int(period.strftime("%Y%m"))

    start_date = period.replace(day=1).date()
    # Get last day of month
    if period.month == 12:
        end_date = period.replace(year=period.year+1, month=1, day=1).date() - timedelta(days=1)
    else:
        end_date = period.replace(month=period.month+1, day=1).date() - timedelta(days=1)

    try:
        async with pool.acquire() as conn:
            # Get targets from KPITargets table
            targets = await conn.fetchrow("""
                SELECT
                    total_sales_target,
                    house_brand_target,
                    focused_item_1_target,
                    focused_item_2_target,
                    focused_item_3_target,
                    stock_clearance_target,
                    pwp_target,
                    transaction_count_target
                FROM "KPITargets"
                WHERE salesman_id = $1 AND year_month = $2
            """, staff_id, year_month)

            # Get current KPI values
            current = await conn.fetchrow("""
                SELECT
                    SUM(total_sales) as total_sales,
                    SUM(house_brand_sales) as house_brand,
                    SUM(focused_1_sales) as focused_1,
                    SUM(COALESCE(focused_2_sales, 0)) as focused_2,
                    SUM(COALESCE(focused_3_sales, 0)) as focused_3,
                    SUM(COALESCE(clearance_sales, 0)) as clearance,
                    SUM(COALESCE(pwp_sales, 0)) as pwp,
                    SUM(transactions) as transactions
                FROM analytics.mv_staff_daily_kpi
                WHERE staff_id = $1
                  AND sale_date BETWEEN $2 AND $3
            """, staff_id, start_date, end_date)

            def calc_progress(current_val, target_val):
                if not target_val or target_val == 0:
                    return None
                return round((float(current_val or 0) / float(target_val)) * 100, 1)

            result = {
                "total_sales": {
                    "target": float(targets['total_sales_target'] or 0) if targets else 0,
                    "current": float(current['total_sales'] or 0) if current else 0,
                    "progress": calc_progress(current['total_sales'] if current else 0,
                                            targets['total_sales_target'] if targets else 0)
                },
                "house_brand": {
                    "target": float(targets['house_brand_target'] or 0) if targets else 0,
                    "current": float(current['house_brand'] or 0) if current else 0,
                    "progress": calc_progress(current['house_brand'] if current else 0,
                                            targets['house_brand_target'] if targets else 0)
                },
                "focused_1": {
                    "target": float(targets['focused_item_1_target'] or 0) if targets else 0,
                    "current": float(current['focused_1'] or 0) if current else 0,
                    "progress": calc_progress(current['focused_1'] if current else 0,
                                            targets['focused_item_1_target'] if targets else 0)
                },
                "focused_2": {
                    "target": float(targets['focused_item_2_target'] or 0) if targets else 0,
                    "current": float(current['focused_2'] or 0) if current else 0,
                    "progress": calc_progress(current['focused_2'] if current else 0,
                                            targets['focused_item_2_target'] if targets else 0)
                },
                "focused_3": {
                    "target": float(targets['focused_item_3_target'] or 0) if targets else 0,
                    "current": float(current['focused_3'] or 0) if current else 0,
                    "progress": calc_progress(current['focused_3'] if current else 0,
                                            targets['focused_item_3_target'] if targets else 0)
                },
                "clearance": {
                    "target": float(targets['stock_clearance_target'] or 0) if targets else 0,
                    "current": float(current['clearance'] or 0) if current else 0,
                    "progress": calc_progress(current['clearance'] if current else 0,
                                            targets['stock_clearance_target'] if targets else 0)
                },
                "pwp": {
                    "target": float(targets['pwp_target'] or 0) if targets else 0,
                    "current": float(current['pwp'] or 0) if current else 0,
                    "progress": calc_progress(current['pwp'] if current else 0,
                                            targets['pwp_target'] if targets else 0)
                },
                "transactions": {
                    "target": int(targets['transaction_count_target'] or 0) if targets else 0,
                    "current": int(current['transactions'] or 0) if current else 0,
                    "progress": calc_progress(current['transactions'] if current else 0,
                                            targets['transaction_count_target'] if targets else 0)
                }
            }

            return {"success": True, "data": result, "period": month or period.strftime("%Y-%m")}

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error fetching targets: {str(e)}")


@app.get("/api/v1/targets/template")
async def download_target_template():
    """Download Excel template for target upload."""
    if not EXCEL_AVAILABLE:
        raise HTTPException(status_code=500, detail="Excel support not available. Install openpyxl.")

    wb = Workbook()
    ws = wb.active
    ws.title = "Targets"

    # Headers
    headers = [
        "staff_id", "year_month", "total_sales", "house_brand",
        "focused_1", "focused_2", "focused_3", "pwp", "clearance", "transactions"
    ]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    # Sample row
    sample = ["SM001", 202501, 50000, 5000, 3000, 2000, 1000, 500, 500, 500]
    for col, value in enumerate(sample, 1):
        ws.cell(row=2, column=col, value=value)

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=target_template.xlsx"}
    )


@app.post("/api/v1/targets/upload")
async def upload_targets(
    file: UploadFile = File(...),
    token: str = Query(..., description="Session token")
):
    """Upload targets from Excel file (Operations Manager only)."""
    # Verify session and permissions
    if token not in sessions:
        raise HTTPException(status_code=401, detail="Invalid session")

    user = sessions[token]['user']
    if not user['permissions'].get('can_upload_targets'):
        raise HTTPException(status_code=403, detail="Permission denied. Only Operations Manager can upload targets.")

    if not EXCEL_AVAILABLE:
        raise HTTPException(status_code=500, detail="Excel support not available. Install openpyxl.")

    try:
        # Read Excel file
        contents = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(contents))
        ws = wb.active

        # Parse rows (skip header)
        rows_processed = 0
        errors = []

        async with pool.acquire() as conn:
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row[0]:  # Skip empty rows
                    continue

                try:
                    staff_id = str(row[0]).strip()
                    year_month = int(row[1])
                    total_sales = float(row[2] or 0)
                    house_brand = float(row[3] or 0)
                    focused_1 = float(row[4] or 0)
                    focused_2 = float(row[5] or 0)
                    focused_3 = float(row[6] or 0)
                    pwp = float(row[7] or 0)
                    clearance = float(row[8] or 0)
                    transactions = int(row[9] or 0)

                    # Upsert target
                    await conn.execute("""
                        INSERT INTO "KPITargets" (
                            salesman_id, year_month, total_sales_target, house_brand_target,
                            focused_item_1_target, focused_item_2_target, focused_item_3_target,
                            pwp_target, stock_clearance_target, transaction_count_target,
                            updated_at
                        ) VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9, $10, NOW())
                        ON CONFLICT (salesman_id, year_month)
                        DO UPDATE SET
                            total_sales_target = EXCLUDED.total_sales_target,
                            house_brand_target = EXCLUDED.house_brand_target,
                            focused_item_1_target = EXCLUDED.focused_item_1_target,
                            focused_item_2_target = EXCLUDED.focused_item_2_target,
                            focused_item_3_target = EXCLUDED.focused_item_3_target,
                            pwp_target = EXCLUDED.pwp_target,
                            stock_clearance_target = EXCLUDED.stock_clearance_target,
                            transaction_count_target = EXCLUDED.transaction_count_target,
                            updated_at = NOW()
                    """, staff_id, year_month, total_sales, house_brand, focused_1,
                        focused_2, focused_3, pwp, clearance, transactions)

                    rows_processed += 1

                except Exception as e:
                    errors.append(f"Row {row_idx}: {str(e)}")

        return {
            "success": True,
            "rows_processed": rows_processed,
            "errors": errors if errors else None
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Upload failed: {str(e)}")


# ============================================================================
# Commission Calculation Endpoints
# ============================================================================

@app.get("/api/v1/commission/me")
async def get_my_commission(
    staff_id: str = Query(..., description="Staff ID"),
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None)
):
    """Calculate commission - HYBRID: MV for history + real-time for today."""
    if not start_date:
        start_date = date.today().replace(day=1)
    if not end_date:
        end_date = date.today()

    today = date.today()
    today_start = datetime.combine(today, datetime.min.time())
    today_end = datetime.combine(today + timedelta(days=1), datetime.min.time())

    try:
        async with pool.acquire() as conn:
            # HYBRID: MV for history + real-time for today
            # Step 1: Get historical commission from MV (fast)
            mv_result = await conn.fetchrow("""
                SELECT
                    COALESCE(SUM(transactions), 0) as transaction_count,
                    COALESCE(SUM(total_sales), 0) as total_sales,
                    COALESCE(SUM(commission), 0) as commission
                FROM analytics.mv_staff_daily_commission
                WHERE staff_id = $1
                  AND sale_date BETWEEN $2 AND $3
                  AND sale_date < $4
            """, staff_id, start_date, end_date, today)

            # Step 2: Get today's commission from base tables (real-time)
            today_result = await conn.fetchrow("""
                SELECT
                    COUNT(DISTINCT c."DocumentNo") as transaction_count,
                    COALESCE(SUM(d."ItemAmount"), 0) as total_sales,
                    COALESCE(SUM(d."ItemAmount" * COALESCE(s."CommissionByPercentStockPrice1", 0) / 100), 0) as commission
                FROM "AcCSM" c
                INNER JOIN "AcCSD" d ON c."DocumentNo" = d."DocumentNo"
                LEFT JOIN "AcStockCompany" s
                    ON d."AcStockID" = s."AcStockID"
                    AND d."AcStockUOMID" = s."AcStockUOMID"
                WHERE c."DocumentDate" >= $2 AND c."DocumentDate" < $3
                  AND d."AcSalesmanID" = $1
                  AND d."ItemAmount" > 0
            """, staff_id, today_start, today_end)

            # Combine MV + today
            total_transactions = int(mv_result['transaction_count'] or 0) + int(today_result['transaction_count'] or 0)
            total_sales = float(mv_result['total_sales'] or 0) + float(today_result['total_sales'] or 0)
            total_commission = float(mv_result['commission'] or 0) + float(today_result['commission'] or 0)
            today_commission = float(today_result['commission'] or 0)

            # Get commission breakdown from MV (fast, doesn't need real-time)
            breakdown = await conn.fetch("""
                SELECT
                    'COMBINED' as category,
                    SUM(total_sales) as sales,
                    SUM(commission) as commission
                FROM analytics.mv_staff_daily_commission
                WHERE staff_id = $1
                  AND sale_date BETWEEN $2 AND $3
            """, staff_id, start_date, end_date)

            return {
                "success": True,
                "data": {
                    "period": {
                        "start": start_date.isoformat(),
                        "end": end_date.isoformat()
                    },
                    "summary": {
                        "total_sales": round(total_sales, 2),
                        "commission_earned": round(total_commission, 2),
                        "transaction_count": total_transactions
                    },
                    "today": {
                        "commission_earned": round(today_commission, 2)
                    },
                    "breakdown": [
                        {
                            "category": row['category'],
                            "sales": float(row['sales'] or 0),
                            "commission": float(row['commission'] or 0)
                        }
                        for row in breakdown
                    ]
                }
            }

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error calculating commission: {str(e)}")


# ============================================================================
# Notification Endpoints
# ============================================================================

@app.get("/api/v1/notifications")
async def get_notifications(
    staff_id: str = Query(..., description="Staff ID"),
    limit: int = Query(20, ge=1, le=100),
    unread_only: bool = Query(False)
):
    """Get notifications for a staff member."""
    try:
        async with pool.acquire() as conn:
            # Check if notifications table exists
            table_exists = await conn.fetchval("""
                SELECT EXISTS (
                    SELECT FROM information_schema.tables
                    WHERE table_schema = 'kpi' AND table_name = 'notifications'
                )
            """)

            if not table_exists:
                return {"success": True, "data": [], "message": "Notifications not yet configured"}

            query = """
                SELECT id, type, title, message, data, is_read, created_at
                FROM kpi.notifications
                WHERE staff_id = $1
            """
            if unread_only:
                query += " AND is_read = FALSE"
            query += " ORDER BY created_at DESC LIMIT $2"

            rows = await conn.fetch(query, staff_id, limit)

            return {
                "success": True,
                "data": [
                    {
                        "id": row['id'],
                        "type": row['type'],
                        "title": row['title'],
                        "message": row['message'],
                        "data": json.loads(row['data']) if row['data'] else None,
                        "is_read": row['is_read'],
                        "created_at": row['created_at'].isoformat()
                    }
                    for row in rows
                ]
            }

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error fetching notifications: {str(e)}")


@app.post("/api/v1/notifications/{notification_id}/read")
async def mark_notification_read(notification_id: int):
    """Mark a notification as read."""
    try:
        async with pool.acquire() as conn:
            await conn.execute("""
                UPDATE kpi.notifications SET is_read = TRUE WHERE id = $1
            """, notification_id)
            return {"success": True}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")


@app.post("/api/v1/notifications/read-all")
async def mark_all_notifications_read(staff_id: str = Query(...)):
    """Mark all notifications as read for a staff member."""
    try:
        async with pool.acquire() as conn:
            await conn.execute("""
                UPDATE kpi.notifications SET is_read = TRUE WHERE staff_id = $1
            """, staff_id)
            return {"success": True}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")


# ============================================================================
# Staff List Endpoint (for dropdowns)
# ============================================================================

@app.get("/api/v1/staff/list")
async def get_staff_list(
    outlet_id: Optional[str] = Query(None, description="Filter by outlet"),
    limit: int = Query(100, ge=1, le=500)
):
    """Get list of active staff members."""
    try:
        async with pool.acquire() as conn:
            if outlet_id:
                # Get staff who have sales at this outlet (from materialized view)
                rows = await conn.fetch("""
                    SELECT DISTINCT
                        s."AcSalesmanID" as staff_id,
                        s."AcSalesmanName" as name,
                        $1 as outlet_id
                    FROM "AcSalesman" s
                    INNER JOIN analytics.mv_staff_daily_kpi k ON s."AcSalesmanID" = k.staff_id
                    WHERE k.outlet_id = $1
                    ORDER BY s."AcSalesmanName"
                    LIMIT $2
                """, outlet_id, limit)
            else:
                # Get all staff with their most recent outlet
                rows = await conn.fetch("""
                    SELECT DISTINCT ON (s."AcSalesmanID")
                        s."AcSalesmanID" as staff_id,
                        s."AcSalesmanName" as name,
                        k.outlet_id
                    FROM "AcSalesman" s
                    LEFT JOIN analytics.mv_staff_daily_kpi k ON s."AcSalesmanID" = k.staff_id
                    ORDER BY s."AcSalesmanID", k.sale_date DESC NULLS LAST
                    LIMIT $1
                """, limit)

            return {
                "success": True,
                "data": [
                    {
                        "staff_id": row['staff_id'],
                        "name": row['name'],
                        "outlet_id": row['outlet_id']
                    }
                    for row in rows
                ]
            }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")


# ============================================================================
# Push Notification Endpoints
# ============================================================================

class PushSubscription(BaseModel):
    staff_id: str
    subscription: dict  # Contains endpoint, keys.p256dh, keys.auth


@app.post("/api/v1/push/subscribe")
async def subscribe_to_push(data: PushSubscription):
    """Subscribe to push notifications."""
    try:
        sub = data.subscription
        async with pool.acquire() as conn:
            await conn.execute("""
                INSERT INTO kpi.push_subscriptions (staff_id, endpoint, p256dh, auth, user_agent)
                VALUES ($1, $2, $3, $4, $5)
                ON CONFLICT (staff_id, endpoint)
                DO UPDATE SET p256dh = $3, auth = $4, last_used_at = NOW()
            """,
                data.staff_id,
                sub.get('endpoint'),
                sub.get('keys', {}).get('p256dh'),
                sub.get('keys', {}).get('auth'),
                None  # user_agent can be added later
            )
            return {"success": True, "message": "Subscribed to push notifications"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Subscription failed: {str(e)}")


class UnsubscribeRequest(BaseModel):
    staff_id: str
    endpoint: str


@app.post("/api/v1/push/unsubscribe")
async def unsubscribe_from_push(data: UnsubscribeRequest):
    """Unsubscribe from push notifications."""
    try:
        async with pool.acquire() as conn:
            await conn.execute("""
                DELETE FROM kpi.push_subscriptions
                WHERE staff_id = $1 AND endpoint = $2
            """, data.staff_id, data.endpoint)
            return {"success": True, "message": "Unsubscribed from push notifications"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Unsubscribe failed: {str(e)}")


async def send_push_to_staff(staff_id: str, title: str, message: str, data: dict = None):
    """Send push notification to all subscribed devices of a staff member."""
    if not PUSH_AVAILABLE:
        return {"success": False, "error": "Push notifications not available"}

    if VAPID_PRIVATE_KEY == 'your-private-key-here':
        return {"success": False, "error": "VAPID keys not configured"}

    sent = 0
    failed = 0
    errors = []

    try:
        async with pool.acquire() as conn:
            subscriptions = await conn.fetch("""
                SELECT endpoint, p256dh, auth
                FROM kpi.push_subscriptions
                WHERE staff_id = $1
            """, staff_id)

            payload = json.dumps({
                "title": title,
                "message": message,
                "data": data or {},
                "timestamp": datetime.now().isoformat()
            })

            for sub in subscriptions:
                try:
                    webpush(
                        subscription_info={
                            "endpoint": sub['endpoint'],
                            "keys": {
                                "p256dh": sub['p256dh'],
                                "auth": sub['auth']
                            }
                        },
                        data=payload,
                        vapid_private_key=VAPID_PRIVATE_KEY,
                        vapid_claims=VAPID_CLAIMS
                    )
                    sent += 1
                except WebPushException as e:
                    failed += 1
                    errors.append(str(e))
                    # Remove invalid subscriptions (410 Gone)
                    if e.response and e.response.status_code == 410:
                        await conn.execute("""
                            DELETE FROM kpi.push_subscriptions WHERE endpoint = $1
                        """, sub['endpoint'])

        return {"success": True, "sent": sent, "failed": failed, "errors": errors[:5]}
    except Exception as e:
        return {"success": False, "error": str(e)}


class SendNotificationRequest(BaseModel):
    staff_id: str
    title: str
    message: str
    notification_type: str = "info"  # info, achievement, commission, alert
    data: Optional[dict] = None
    save_to_db: bool = True


@app.post("/api/v1/push/send")
async def send_push_notification(
    request: SendNotificationRequest,
    api_key: str = Query(..., description="API key for authentication")
):
    """Send push notification to a specific staff member.

    Use this endpoint from the sync service to notify on sales events.
    """
    expected_key = os.getenv('PUSH_API_KEY', 'flt-push-2024')
    if api_key != expected_key:
        raise HTTPException(status_code=401, detail="Invalid API key")

    # Optionally save to notifications table
    if request.save_to_db:
        try:
            async with pool.acquire() as conn:
                await conn.execute("""
                    INSERT INTO kpi.notifications (staff_id, title, message, type, data)
                    VALUES ($1, $2, $3, $4, $5)
                """, request.staff_id, request.title, request.message,
                    request.notification_type, json.dumps(request.data or {}))
        except Exception as e:
            print(f"Failed to save notification to DB: {e}")

    # Send push notification
    result = await send_push_to_staff(
        staff_id=request.staff_id,
        title=request.title,
        message=request.message,
        data={"type": request.notification_type, **(request.data or {})}
    )

    return result


class SalesEventNotification(BaseModel):
    staff_id: str
    staff_name: str
    sale_type: str  # house_brand, focused_1, focused_2, focused_3, pwp, clearance
    amount: float
    product_name: Optional[str] = None
    outlet_id: Optional[str] = None


@app.post("/api/v1/push/sales-event")
async def notify_sales_event(
    event: SalesEventNotification,
    api_key: str = Query(..., description="API key for authentication")
):
    """Notify staff of a sales achievement. Called by sync service on new sales.

    Generates motivational messages for different sale types.
    """
    expected_key = os.getenv('PUSH_API_KEY', 'flt-push-2024')
    if api_key != expected_key:
        raise HTTPException(status_code=401, detail="Invalid API key")

    # Generate motivational message based on sale type
    messages = {
        "house_brand": [
            f"House Brand sale RM{event.amount:.2f}! Keep pushing! 🏠",
            f"Nice! RM{event.amount:.2f} House Brand sold! You're on fire! 🔥",
            f"RM{event.amount:.2f} House Brand! Great job promoting our brands! 💪"
        ],
        "focused_1": [
            f"Focused Item 1 sale! RM{event.amount:.2f} - Excellent focus! 🎯",
            f"RM{event.amount:.2f} Focused Item 1! Target locked! 🎯"
        ],
        "focused_2": [
            f"Focused Item 2 sale! RM{event.amount:.2f} - Keep it up! ⭐",
            f"RM{event.amount:.2f} Focused Item 2! You're crushing it! 💥"
        ],
        "focused_3": [
            f"Focused Item 3 sale! RM{event.amount:.2f} - Amazing! 🌟",
        ],
        "pwp": [
            f"PWP sale RM{event.amount:.2f}! Great upselling! 🛒",
            f"RM{event.amount:.2f} PWP! Awesome add-on sale! 🎁"
        ],
        "clearance": [
            f"Stock Clearance RM{event.amount:.2f}! Helping clear inventory! 📦",
        ]
    }

    sale_messages = messages.get(event.sale_type, [f"New sale RM{event.amount:.2f}!"])
    message = random.choice(sale_messages)

    title_map = {
        "house_brand": "🏠 House Brand Sale!",
        "focused_1": "🎯 Focused Item 1!",
        "focused_2": "⭐ Focused Item 2!",
        "focused_3": "🌟 Focused Item 3!",
        "pwp": "🛒 PWP Sale!",
        "clearance": "📦 Clearance Sale!"
    }
    title = title_map.get(event.sale_type, "💰 New Sale!")

    # Send push notification
    result = await send_push_to_staff(
        staff_id=event.staff_id,
        title=title,
        message=message,
        data={
            "type": "sale",
            "sale_type": event.sale_type,
            "amount": event.amount,
            "product": event.product_name,
            "outlet": event.outlet_id
        }
    )

    return result


@app.post("/api/v1/push/test")
async def send_test_notification(
    staff_id: str = Query(..., description="Staff ID to send test notification to")
):
    """Send a test push notification to verify setup."""
    result = await send_push_to_staff(
        staff_id=staff_id,
        title="🔔 Test Notification",
        message="Push notifications are working! You'll receive sales alerts here.",
        data={"type": "test"}
    )
    return result


# ============================================================================
# Scheduled Daily Notifications (for cron jobs)
# ============================================================================

# Gen Z friendly motivational messages - short, affirming, with emojis
MORNING_MOTIVATIONS = [
    "Let's crush it today! 💪",
    "You've got this bestie! ✨",
    "Time to slay! 🔥",
    "Main character energy today! 🌟",
    "Your goals are loading... 📈",
    "New day, new wins! 🎯",
    "Serving excellence today! 💅",
    "Era of success starts now! 🚀",
]

EVENING_MOTIVATIONS_GOOD = [
    "You ate today! 🔥",
    "Slayed! Tomorrow we go again! 💅",
    "W performance! Keep vibing! ✨",
    "That's the energy we need! 🌟",
    "Legend status achieved! 🏆",
    "No cap, you killed it! 💪",
]

EVENING_MOTIVATIONS_NEEDS_PUSH = [
    "Tomorrow is your redemption arc! 💪",
    "Rest up, comeback loading... 🔄",
    "We move! Try again tomorrow! ✨",
    "It's giving 'work in progress' 📈",
    "Not the end, just a plot twist! 🎬",
    "Glow up starts tomorrow! 🌅",
]

@app.post("/api/v1/push/morning-briefing")
async def send_morning_briefing(
    api_key: str = Query(..., description="API key for authentication")
):
    """Send 8am morning briefing to all subscribed staff.

    Schedule with external cron: 0 8 * * * (daily at 8am)

    Message includes:
    - Gap against monthly target
    - Daily target to stay on track
    - Gen Z motivational message
    """
    expected_key = os.getenv('PUSH_API_KEY', 'flt-push-2024')
    if api_key != expected_key:
        raise HTTPException(status_code=401, detail="Invalid API key")

    if not PUSH_AVAILABLE:
        return {"success": False, "error": "Push notifications not available"}

    today = date.today()
    year_month = int(today.strftime('%Y%m'))
    days_in_month = (date(today.year, today.month + 1 if today.month < 12 else 1, 1) - timedelta(days=1)).day if today.month < 12 else 31
    days_remaining = days_in_month - today.day + 1

    sent_count = 0
    errors = []

    try:
        async with pool.acquire() as conn:
            # Get all subscribed staff with their targets and MTD performance
            staff_data = await conn.fetch("""
                WITH subscribed_staff AS (
                    SELECT DISTINCT staff_id
                    FROM kpi.push_subscriptions
                ),
                mtd_sales AS (
                    SELECT
                        staff_id,
                        SUM(total_sales) as mtd_total,
                        SUM(house_brand_sales) as mtd_house_brand
                    FROM analytics.mv_staff_daily_kpi
                    WHERE EXTRACT(YEAR FROM sale_date) = $1
                      AND EXTRACT(MONTH FROM sale_date) = $2
                    GROUP BY staff_id
                )
                SELECT
                    s.staff_id,
                    COALESCE(m.mtd_total, 0) as mtd_total,
                    COALESCE(m.mtd_house_brand, 0) as mtd_house_brand,
                    COALESCE(t.total_sales_target, 0) as target_total,
                    COALESCE(t.house_brand_target, 0) as target_house_brand
                FROM subscribed_staff s
                LEFT JOIN mtd_sales m ON s.staff_id = m.staff_id
                LEFT JOIN "KPITargets" t ON s.staff_id = t.salesman_id AND t.year_month = $3
            """, today.year, today.month, year_month)

            motivation = random.choice(MORNING_MOTIVATIONS)

            for staff in staff_data:
                mtd = float(staff['mtd_total'] or 0)
                target = float(staff['target_total'] or 0)

                if target > 0:
                    gap = target - mtd
                    daily_needed = gap / days_remaining if days_remaining > 0 else gap
                    progress_pct = (mtd / target) * 100

                    if gap <= 0:
                        title = "🎯 You're ahead of target!"
                        msg = f"MTD: RM{mtd:,.0f} | Already hit RM{target:,.0f}! {motivation}"
                    else:
                        title = "☀️ Good morning!"
                        msg = f"Gap: RM{gap:,.0f} | Need RM{daily_needed:,.0f}/day | {days_remaining} days left. {motivation}"
                else:
                    title = "☀️ Rise and grind!"
                    msg = f"MTD: RM{mtd:,.0f} | No target set yet. {motivation}"

                try:
                    await send_push_to_staff(
                        staff_id=staff['staff_id'],
                        title=title,
                        message=msg,
                        data={"type": "morning_briefing", "mtd": mtd, "target": target}
                    )
                    sent_count += 1
                except Exception as e:
                    errors.append(f"{staff['staff_id']}: {str(e)}")

        return {
            "success": True,
            "sent_count": sent_count,
            "errors": errors[:10] if errors else [],
            "timestamp": datetime.now().isoformat()
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Morning briefing failed: {str(e)}")


@app.post("/api/v1/push/evening-recap")
async def send_evening_recap(
    api_key: str = Query(..., description="API key for authentication")
):
    """Send 10:30pm evening recap to all subscribed staff.

    Schedule with external cron: 30 22 * * * (daily at 10:30pm)

    Message includes:
    - Today's performance
    - Gap against monthly target
    - Motivational message based on performance
    """
    expected_key = os.getenv('PUSH_API_KEY', 'flt-push-2024')
    if api_key != expected_key:
        raise HTTPException(status_code=401, detail="Invalid API key")

    if not PUSH_AVAILABLE:
        return {"success": False, "error": "Push notifications not available"}

    today = date.today()
    year_month = int(today.strftime('%Y%m'))
    days_in_month = (date(today.year, today.month + 1 if today.month < 12 else 1, 1) - timedelta(days=1)).day if today.month < 12 else 31
    expected_progress = (today.day / days_in_month) * 100

    sent_count = 0
    errors = []

    try:
        async with pool.acquire() as conn:
            # Get all subscribed staff with today's and MTD performance
            staff_data = await conn.fetch("""
                WITH subscribed_staff AS (
                    SELECT DISTINCT staff_id
                    FROM kpi.push_subscriptions
                ),
                today_sales AS (
                    SELECT
                        staff_id,
                        SUM(total_sales) as today_total,
                        SUM(transactions) as today_trans
                    FROM analytics.mv_staff_daily_kpi
                    WHERE sale_date = $1
                    GROUP BY staff_id
                ),
                mtd_sales AS (
                    SELECT
                        staff_id,
                        SUM(total_sales) as mtd_total
                    FROM analytics.mv_staff_daily_kpi
                    WHERE EXTRACT(YEAR FROM sale_date) = $2
                      AND EXTRACT(MONTH FROM sale_date) = $3
                    GROUP BY staff_id
                )
                SELECT
                    s.staff_id,
                    COALESCE(d.today_total, 0) as today_total,
                    COALESCE(d.today_trans, 0) as today_trans,
                    COALESCE(m.mtd_total, 0) as mtd_total,
                    COALESCE(t.total_sales_target, 0) as target_total
                FROM subscribed_staff s
                LEFT JOIN today_sales d ON s.staff_id = d.staff_id
                LEFT JOIN mtd_sales m ON s.staff_id = m.staff_id
                LEFT JOIN "KPITargets" t ON s.staff_id = t.salesman_id AND t.year_month = $4
            """, today, today.year, today.month, year_month)

            for staff in staff_data:
                today_sales = float(staff['today_total'] or 0)
                today_trans = int(staff['today_trans'] or 0)
                mtd = float(staff['mtd_total'] or 0)
                target = float(staff['target_total'] or 0)

                if target > 0:
                    progress_pct = (mtd / target) * 100
                    gap = target - mtd

                    # Good performance = ahead or on track
                    is_good = progress_pct >= expected_progress * 0.9  # Within 10% of expected
                    motivation = random.choice(EVENING_MOTIVATIONS_GOOD if is_good else EVENING_MOTIVATIONS_NEEDS_PUSH)

                    if progress_pct >= 100:
                        title = "🏆 Target achieved!"
                        msg = f"Today: RM{today_sales:,.0f} | MTD: RM{mtd:,.0f}/{target:,.0f} ({progress_pct:.0f}%). {motivation}"
                    elif is_good:
                        title = "🌙 Great day!"
                        msg = f"Today: RM{today_sales:,.0f} | Gap: RM{gap:,.0f} | {progress_pct:.0f}% done. {motivation}"
                    else:
                        title = "🌙 Day's done!"
                        msg = f"Today: RM{today_sales:,.0f} | Gap: RM{gap:,.0f} | {progress_pct:.0f}% done. {motivation}"
                else:
                    motivation = random.choice(EVENING_MOTIVATIONS_GOOD if today_sales > 500 else EVENING_MOTIVATIONS_NEEDS_PUSH)
                    title = "🌙 That's a wrap!"
                    msg = f"Today: RM{today_sales:,.0f} | {today_trans} transactions. {motivation}"

                try:
                    await send_push_to_staff(
                        staff_id=staff['staff_id'],
                        title=title,
                        message=msg,
                        data={"type": "evening_recap", "today": today_sales, "mtd": mtd, "target": target}
                    )
                    sent_count += 1
                except Exception as e:
                    errors.append(f"{staff['staff_id']}: {str(e)}")

        return {
            "success": True,
            "sent_count": sent_count,
            "errors": errors[:10] if errors else [],
            "timestamp": datetime.now().isoformat()
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Evening recap failed: {str(e)}")


# ============================================================================
# Debug Endpoints
# ============================================================================

# ============================================================================
# Scheduled Refresh Endpoint (for cron jobs)
# ============================================================================

# Background refresh state
_refresh_in_progress = False
_last_refresh_result = None


async def _do_background_refresh():
    """Background task to refresh all materialized views.

    IMPORTANT: Uses dedicated connection with 30-minute timeout instead of pool.
    Pool connections have 60-second command_timeout which is too short for MV refresh.
    """
    global _refresh_in_progress, _last_refresh_result
    import time
    start_time = time.time()
    results = {}

    try:
        # Create dedicated connection with NO command timeout for MV refresh
        # Pool connections have 60s command_timeout which causes TimeoutError
        # MV refresh time scales with data size - should never timeout at Python level
        # PostgreSQL statement_timeout provides the safety net (3 hours)
        # Derive actual hostname from connected_host (which contains description like "host (internal, no SSL)")
        use_internal = connected_host and 'internal' in connected_host
        mv_host = INTERNAL_HOST if use_internal else EXTERNAL_HOST
        mv_conn = await asyncpg.connect(
            host=mv_host,
            port=DB_PORT,
            database=DB_NAME,
            user=DB_USER,
            password=DB_PASSWORD,
            ssl=None if use_internal else 'require',
            command_timeout=None,  # No Python timeout - let PostgreSQL handle it
        )

        try:
            # Set PostgreSQL statement timeout as safety net (3 hours)
            # This ensures runaway queries eventually terminate
            await mv_conn.execute("SET statement_timeout = '10800000'")  # 3 hours

            views = [
                'analytics.mv_staff_daily_kpi',
                'analytics.mv_outlet_daily_kpi',
                'analytics.mv_staff_rankings',
                'analytics.mv_staff_daily_commission'  # Commission MV added
            ]

            for view in views:
                view_start = time.time()
                try:
                    await mv_conn.execute(f'REFRESH MATERIALIZED VIEW CONCURRENTLY {view}')
                except Exception as e:
                    error_type = type(e).__name__
                    error_msg = str(e) or repr(e)
                    if 'unique index' in error_msg.lower() or 'does not exist' in error_msg.lower():
                        try:
                            await mv_conn.execute(f'REFRESH MATERIALIZED VIEW {view}')
                        except Exception as e2:
                            results[view] = f"skipped: {type(e2).__name__}: {str(e2)[:40]}"
                            continue
                    else:
                        results[view] = f"error: {error_type}: {error_msg[:40]}"
                        continue
                results[view] = round(time.time() - view_start, 1)

            _last_refresh_result = {
                "success": True,
                "timing": results,
                "total_seconds": round(time.time() - start_time, 1),
                "completed_at": datetime.now().isoformat()
            }
        finally:
            await mv_conn.close()

    except Exception as e:
        _last_refresh_result = {
            "success": False,
            "error": str(e),
            "completed_at": datetime.now().isoformat()
        }
    finally:
        _refresh_in_progress = False


@app.post("/api/v1/admin/refresh-views")
async def refresh_materialized_views(
    api_key: str = Query(..., description="API key for authentication"),
    wait: bool = Query(False, description="Wait for completion (default: fire-and-forget)")
):
    """Refresh all materialized views. Returns immediately by default (fire-and-forget).

    Use wait=true to wait for completion (may timeout for external calls).
    Check /api/v1/admin/refresh-status for background refresh status.
    """
    global _refresh_in_progress

    expected_key = os.getenv('REFRESH_API_KEY', 'flt-refresh-2024')
    if api_key != expected_key:
        raise HTTPException(status_code=401, detail="Invalid API key")

    if _refresh_in_progress:
        return {
            "success": True,
            "message": "Refresh already in progress",
            "status": "running"
        }

    _refresh_in_progress = True

    if wait:
        # Wait for completion (original behavior)
        await _do_background_refresh()
        return _last_refresh_result
    else:
        # Fire-and-forget: start background task and return immediately
        asyncio.create_task(_do_background_refresh())
        return {
            "success": True,
            "message": "Refresh started in background",
            "status": "started",
            "check_status_at": "/api/v1/admin/refresh-status"
        }


@app.get("/api/v1/admin/refresh-status")
async def get_refresh_status():
    """Check the status of background MV refresh."""
    return {
        "in_progress": _refresh_in_progress,
        "last_result": _last_refresh_result
    }


@app.get("/api/v1/admin/view-status")
async def get_view_status():
    """Check the status of materialized views (latest data date)."""
    try:
        async with pool.acquire() as conn:
            # Get latest dates from each view
            staff_daily = await conn.fetchrow("""
                SELECT MAX(sale_date) as latest_date, COUNT(*) as row_count
                FROM analytics.mv_staff_daily_kpi
            """)

            outlet_daily = await conn.fetchrow("""
                SELECT MAX(sale_date) as latest_date, COUNT(*) as row_count
                FROM analytics.mv_outlet_daily_kpi
            """)

            rankings = await conn.fetchrow("""
                SELECT MAX(month) as latest_month, COUNT(*) as row_count
                FROM analytics.mv_staff_rankings
            """)

            return {
                "success": True,
                "views": {
                    "mv_staff_daily_kpi": {
                        "latest_date": staff_daily['latest_date'].isoformat() if staff_daily['latest_date'] else None,
                        "row_count": staff_daily['row_count']
                    },
                    "mv_outlet_daily_kpi": {
                        "latest_date": outlet_daily['latest_date'].isoformat() if outlet_daily['latest_date'] else None,
                        "row_count": outlet_daily['row_count']
                    },
                    "mv_staff_rankings": {
                        "latest_month": rankings['latest_month'].isoformat() if rankings['latest_month'] else None,
                        "row_count": rankings['row_count']
                    }
                },
                "checked_at": datetime.now().isoformat()
            }

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")


@app.get("/api/v1/regions")
async def get_regions():
    """Get list of regions with their area managers.

    Returns region data dynamically from staff_list_master.
    Frontend can use this instead of hardcoding REGIONS constant.
    """
    try:
        async with pool.acquire() as conn:
            # Get unique regions with their area managers
            rows = await conn.fetch("""
                SELECT DISTINCT
                    region,
                    area_manager_name,
                    area_manager_id
                FROM kpi.staff_list_master
                WHERE region IS NOT NULL AND is_active = true
                ORDER BY region
            """)

            regions = {}
            for row in rows:
                region = row['region']
                if region and region not in regions:
                    regions[region] = {
                        'code': region,
                        'label': f"{region} - {row['area_manager_name']}" if row['area_manager_name'] else region,
                        'area_manager_id': row['area_manager_id'],
                        'area_manager_name': row['area_manager_name']
                    }

            return {
                "success": True,
                "regions": list(regions.values())
            }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")


# ============================================================================
# Debug Endpoints
# ============================================================================

@app.get("/api/v1/debug/sales-breakdown")
async def debug_sales_breakdown(
    outlet_id: str = Query(..., description="Outlet ID"),
    start_date: date = Query(..., description="Start date"),
    end_date: date = Query(..., description="End date")
):
    """Debug endpoint to compare cash vs invoice sales from raw tables."""
    try:
        async with pool.acquire() as conn:
            # Use materialized view for speed - it already has the totals
            mv_total = await conn.fetchrow("""
                SELECT
                    SUM(transactions) as transactions,
                    SUM(total_sales) as total
                FROM analytics.mv_outlet_daily_kpi
                WHERE outlet_id = $1
                  AND sale_date BETWEEN $2 AND $3
            """, outlet_id, start_date, end_date)

            # Quick check for negative ItemTotal (returns/refunds) - simplified
            negative_check = await conn.fetchrow("""
                SELECT
                    COUNT(*) as cnt,
                    COALESCE(SUM(d."ItemTotal"), 0) as total
                FROM "AcCSD" d
                INNER JOIN "AcCSM" m ON d."DocumentNo" = m."DocumentNo"
                WHERE m."AcLocationID" = $1
                  AND m."DocumentDate"::date BETWEEN $2 AND $3
                  AND d."ItemTotal" < 0
            """, outlet_id, start_date, end_date)

            return {
                "success": True,
                "outlet_id": outlet_id,
                "period": {"start": start_date.isoformat(), "end": end_date.isoformat()},
                "materialized_view_total": round(float(mv_total['total'] or 0), 2),
                "negative_items": {
                    "count": int(negative_check['cnt'] or 0),
                    "total": round(float(negative_check['total'] or 0), 2)
                },
                "note": "If negative_items total is significant, Dynamod may be excluding returns"
            }

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")


def decode_dynamod_password(encoded: str) -> str:
    """
    Attempt to decode Dynamod password encoding.

    Dynamod uses a simple XOR-based obfuscation where each character is XORed
    with a position-dependent key. Based on analysis:
    - The encoding XORs each character with a value that varies by position
    - This function attempts to reverse the encoding
    """
    if not encoded:
        return ""

    # Try multiple decoding approaches
    decoded_attempts = []

    # Approach 1: Try XOR with common patterns
    # Based on observed patterns, try different XOR key sequences
    xor_patterns = [
        [6, 3, 4, 2, 3, 4],  # Pattern observed for some passwords
        [6, 8, 5, 9, 6, 8],  # Another observed pattern
        [2, 3, 4, 2, 3, 4],  # Simple repeating
    ]

    for pattern in xor_patterns:
        try:
            decoded = ""
            for i, char in enumerate(encoded):
                key = pattern[i % len(pattern)]
                decoded_char = chr(ord(char) ^ key)
                decoded += decoded_char
            if decoded.isdigit() or decoded.isalnum():
                decoded_attempts.append({"method": f"XOR pattern {pattern}", "result": decoded})
        except:
            pass

    # Approach 2: Simple character shift (Caesar-like)
    for shift in range(-10, 11):
        try:
            decoded = "".join(chr(ord(c) + shift) for c in encoded)
            if decoded.isdigit():
                decoded_attempts.append({"method": f"Shift {shift}", "result": decoded})
        except:
            pass

    return decoded_attempts


@app.get("/api/v1/debug/user-credentials")
async def debug_user_credentials(
    code: str = Query(..., description="User code to check (e.g., 'LJL', '30', 'LTK')")
):
    """Debug endpoint to check what credentials are stored in AcPersonal table.

    IMPORTANT: This endpoint is for debugging only. It shows the stored password
    from the AcPersonal table which is synced from Dynamod SQL Server.
    """
    try:
        async with pool.acquire() as conn:
            # Get user from AcPersonal
            user = await conn.fetchrow("""
                SELECT
                    "Code" as code,
                    "Name" as name,
                    "Password" as password,
                    "Active" as active,
                    "IsSupervisor" as is_supervisor,
                    "AcPOSUserGroupID" as user_group
                FROM "AcPersonal"
                WHERE UPPER("Code") = UPPER($1)
            """, code)

            if not user:
                return {
                    "success": False,
                    "error": f"User '{code}' not found in AcPersonal table"
                }

            stored_password = user['password'] or ""

            # Try to decode the password
            decode_attempts = decode_dynamod_password(stored_password)

            return {
                "success": True,
                "message": "Password retrieved from AcPersonal table (synced from Dynamod)",
                "user": {
                    "code": user['code'],
                    "name": user['name'],
                    "stored_password": stored_password,
                    "stored_password_bytes": [ord(c) for c in stored_password],  # Show ASCII values
                    "active": user['active'],
                    "is_supervisor": user['is_supervisor'],
                    "user_group": user['user_group']
                },
                "decode_attempts": decode_attempts,
                "encoding_analysis": {
                    "length": len(stored_password),
                    "char_codes": [f"{c} (ASCII {ord(c)})" for c in stored_password]
                },
                "note": "The password is ENCODED. Use /api/v1/debug/password-encoder to test encoding."
            }

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")


@app.get("/api/v1/debug/password-encoder")
async def debug_password_encoder(
    plain_password: str = Query(..., description="Plain text password to encode"),
    stored_password: str = Query(..., description="Known stored/encoded password"),
    user_code: str = Query("", description="User code (optional, to test encoding)")
):
    """
    Debug endpoint to figure out the password encoding by comparing plain and stored passwords.

    Provide a known plain password and its stored/encoded version to reverse-engineer the encoding.
    """
    try:
        if len(plain_password) != len(stored_password):
            return {
                "success": False,
                "error": f"Length mismatch: plain={len(plain_password)}, stored={len(stored_password)}",
                "hint": "Passwords should be same length if directly encoded"
            }

        # Calculate XOR keys for each position
        xor_keys = []
        for i, (p, s) in enumerate(zip(plain_password, stored_password)):
            xor_key = ord(p) ^ ord(s)
            xor_keys.append({
                "position": i,
                "plain_char": p,
                "plain_ascii": ord(p),
                "stored_char": s,
                "stored_ascii": ord(s),
                "xor_key": xor_key
            })

        # Check if there's a repeating pattern in XOR keys
        key_values = [k["xor_key"] for k in xor_keys]

        # Try to find repeating pattern
        patterns_found = []
        for pattern_len in range(1, len(key_values) // 2 + 1):
            pattern = key_values[:pattern_len]
            is_repeating = True
            for i in range(len(key_values)):
                if key_values[i] != pattern[i % pattern_len]:
                    is_repeating = False
                    break
            if is_repeating:
                patterns_found.append(pattern)

        # Test if our encoding function works
        encoding_test = None
        if user_code:
            will_match = check_password_dynamod(plain_password, stored_password, user_code)
            encoding_test = {
                "user_code": user_code,
                "login_will_work": will_match,
                "encoded_with_offset_70": encode_password_dynamod(plain_password, user_code)
            }

        return {
            "success": True,
            "analysis": {
                "plain_password": plain_password,
                "stored_password": stored_password,
                "xor_keys": xor_keys,
                "key_pattern": key_values,
                "repeating_patterns": patterns_found if patterns_found else "No repeating pattern found"
            },
            "encoding_test": encoding_test,
            "recommendation": "If login_will_work is True, the password will work with login"
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")


@app.get("/api/v1/debug/ic-password-analysis")
async def debug_ic_password_analysis():
    """
    Analyze password encoding for all users with 12-digit Malaysia IC codes.
    For these users, the password should be the last 4 digits of the IC.
    This helps verify and discover the XOR encoding pattern.
    """
    try:
        async with pool.acquire() as conn:
            # Query all users with 12-digit codes (Malaysia IC format)
            users = await conn.fetch("""
                SELECT
                    "Code" as code,
                    "Name" as name,
                    "Password" as stored_password,
                    "Active" as active
                FROM "AcPersonal"
                WHERE LENGTH("Code") = 12
                  AND "Code" ~ '^[0-9]+$'
                  AND "Password" IS NOT NULL
                  AND "Password" != ''
                ORDER BY "Active" DESC, "Code"
            """)

            results = []
            xor_patterns_discovered = {}
            successful_patterns = []

            for user in users:
                code = user['code']
                stored_password = user['stored_password']
                expected_password = code[-4:]  # Last 4 digits of IC

                # Calculate XOR keys
                xor_keys = []
                if len(stored_password) == 4:  # Password length matches expected
                    for i in range(4):
                        plain_char = expected_password[i]
                        stored_char = stored_password[i]
                        xor_key = ord(plain_char) ^ ord(stored_char)
                        xor_keys.append(xor_key)

                    # Check if login would work with current implementation
                    login_works = check_password_dynamod(expected_password, stored_password, code)

                    result = {
                        "code": code,
                        "name": user['name'],
                        "active": user['active'],
                        "expected_password": expected_password,
                        "stored_password": stored_password,
                        "stored_ascii": [ord(c) for c in stored_password],
                        "xor_pattern": xor_keys,
                        "login_works": login_works
                    }
                    results.append(result)

                    # Track discovered patterns
                    pattern_key = tuple(xor_keys)
                    if pattern_key not in xor_patterns_discovered:
                        xor_patterns_discovered[pattern_key] = []
                    xor_patterns_discovered[pattern_key].append(code)

                    if login_works:
                        successful_patterns.append(pattern_key)
                else:
                    results.append({
                        "code": code,
                        "name": user['name'],
                        "active": user['active'],
                        "expected_password": expected_password,
                        "stored_password": stored_password,
                        "error": f"Password length mismatch: expected 4, got {len(stored_password)}"
                    })

            # Summarize patterns
            pattern_summary = []
            for pattern, codes in xor_patterns_discovered.items():
                pattern_summary.append({
                    "pattern": list(pattern),
                    "count": len(codes),
                    "works_with_current_impl": pattern in successful_patterns,
                    "sample_codes": codes[:5]  # Show first 5
                })

            # Sort by count descending
            pattern_summary.sort(key=lambda x: x['count'], reverse=True)

            return {
                "success": True,
                "total_ic_users": len(users),
                "users_analyzed": len(results),
                "users_with_working_login": len([r for r in results if r.get('login_works')]),
                "pattern_summary": pattern_summary,
                "details": results,
                "recommendation": "Add missing patterns to fixed_patterns in check_password_dynamod()"
            }

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")


# ============================================================================
# Outlet Targets Endpoints (Team-level targets)
# ============================================================================

@app.get("/api/v1/outlet-targets/template")
async def download_outlet_target_template():
    """Download Excel template for outlet target upload."""
    if not EXCEL_AVAILABLE:
        raise HTTPException(status_code=500, detail="Excel support not available. Install openpyxl.")

    try:
        async with pool.acquire() as conn:
            # Get all outlets
            outlets = await conn.fetch("""
                SELECT "AcLocationID" as id, "AcLocationDesc" as name
                FROM "AcLocation"
                WHERE "IsActive" = 'Y'
                ORDER BY "AcLocationID"
            """)

        wb = Workbook()
        ws = wb.active
        ws.title = "Outlet Targets"

        # Headers
        headers = [
            'outlet_id', 'outlet_name', 'year_month',
            'total_sales', 'gross_profit', 'house_brand', 'focused_1', 'focused_2', 'focused_3',
            'pwp', 'clearance', 'transactions'
        ]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        # Pre-fill with outlet data and current month
        current_month = int(datetime.now().strftime('%Y%m'))
        for row, outlet in enumerate(outlets, 2):
            ws.cell(row=row, column=1, value=outlet['id'])
            ws.cell(row=row, column=2, value=outlet['name'])
            ws.cell(row=row, column=3, value=current_month)
            # Leave target columns empty for user to fill

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=outlet_target_template.xlsx"}
        )

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error generating template: {str(e)}")


@app.post("/api/v1/outlet-targets/upload")
async def upload_outlet_targets(
    file: UploadFile = File(...),
    token: str = Query(..., description="Session token")
):
    """Upload outlet targets from Excel file (Admin/Operations Manager only)."""
    # Verify session and permissions
    if token not in sessions:
        raise HTTPException(status_code=401, detail="Invalid session")

    user = sessions[token]['user']
    if not user['permissions'].get('can_upload_targets'):
        raise HTTPException(status_code=403, detail="Permission denied. Only Admin/Operations Manager can upload targets.")

    if not EXCEL_AVAILABLE:
        raise HTTPException(status_code=500, detail="Excel support not available. Install openpyxl.")

    try:
        # Read Excel file
        content = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb.active

        rows_processed = 0
        errors = []

        async with pool.acquire() as conn:
            # Ensure OutletTargets table exists with gross_profit_target
            await conn.execute("""
                CREATE TABLE IF NOT EXISTS "OutletTargets" (
                    outlet_id VARCHAR(20) NOT NULL,
                    year_month INT NOT NULL,
                    total_sales_target DECIMAL(15,2) DEFAULT 0,
                    gross_profit_target DECIMAL(15,2) DEFAULT 0,
                    house_brand_target DECIMAL(15,2) DEFAULT 0,
                    focused_item_1_target DECIMAL(15,2) DEFAULT 0,
                    focused_item_2_target DECIMAL(15,2) DEFAULT 0,
                    focused_item_3_target DECIMAL(15,2) DEFAULT 0,
                    pwp_target DECIMAL(15,2) DEFAULT 0,
                    stock_clearance_target DECIMAL(15,2) DEFAULT 0,
                    transaction_count_target INT DEFAULT 0,
                    updated_at TIMESTAMP DEFAULT NOW(),
                    updated_by VARCHAR(20),
                    PRIMARY KEY (outlet_id, year_month)
                )
            """)

            # Add gross_profit_target column if it doesn't exist (for existing tables)
            await conn.execute("""
                DO $$
                BEGIN
                    IF NOT EXISTS (
                        SELECT 1 FROM information_schema.columns
                        WHERE table_name = 'OutletTargets' AND column_name = 'gross_profit_target'
                    ) THEN
                        ALTER TABLE "OutletTargets" ADD COLUMN gross_profit_target DECIMAL(15,2) DEFAULT 0;
                    END IF;
                END $$;
            """)

            # Process rows (skip header)
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row[0]:  # Skip empty rows
                    continue

                try:
                    outlet_id = str(row[0]).strip()
                    # Skip outlet_name column (index 1)
                    year_month = int(row[2] or 0)
                    total_sales = float(row[3] or 0)
                    gross_profit = float(row[4] or 0)
                    house_brand = float(row[5] or 0)
                    focused_1 = float(row[6] or 0)
                    focused_2 = float(row[7] or 0)
                    focused_3 = float(row[8] or 0)
                    pwp = float(row[9] or 0)
                    clearance = float(row[10] or 0)
                    transactions = int(row[11] or 0)

                    if year_month < 202401 or year_month > 203012:
                        errors.append(f"Row {row_num}: Invalid year_month {year_month}")
                        continue

                    # Upsert outlet target
                    await conn.execute("""
                        INSERT INTO "OutletTargets" (
                            outlet_id, year_month, total_sales_target, gross_profit_target,
                            house_brand_target, focused_item_1_target, focused_item_2_target,
                            focused_item_3_target, pwp_target, stock_clearance_target,
                            transaction_count_target, updated_at, updated_by
                        ) VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9, $10, $11, NOW(), $12)
                        ON CONFLICT (outlet_id, year_month)
                        DO UPDATE SET
                            total_sales_target = EXCLUDED.total_sales_target,
                            gross_profit_target = EXCLUDED.gross_profit_target,
                            house_brand_target = EXCLUDED.house_brand_target,
                            focused_item_1_target = EXCLUDED.focused_item_1_target,
                            focused_item_2_target = EXCLUDED.focused_item_2_target,
                            focused_item_3_target = EXCLUDED.focused_item_3_target,
                            pwp_target = EXCLUDED.pwp_target,
                            stock_clearance_target = EXCLUDED.stock_clearance_target,
                            transaction_count_target = EXCLUDED.transaction_count_target,
                            updated_at = NOW(),
                            updated_by = EXCLUDED.updated_by
                    """, outlet_id, year_month, total_sales, gross_profit, house_brand,
                        focused_1, focused_2, focused_3, pwp, clearance, transactions, user['code'])

                    rows_processed += 1

                except Exception as row_error:
                    errors.append(f"Row {row_num}: {str(row_error)}")

        return {
            "success": True,
            "rows_processed": rows_processed,
            "errors": errors if errors else None
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Upload failed: {str(e)}")


@app.get("/api/v1/outlet-targets")
async def get_outlet_targets(
    outlet_id: Optional[str] = Query(None, description="Outlet ID (empty for all)"),
    outlet_ids: Optional[str] = Query(None, description="Comma-separated outlet IDs"),
    month: Optional[str] = Query(None, description="Month in YYYY-MM format")
):
    """Get outlet targets with current progress."""
    if month:
        try:
            period = datetime.strptime(month, "%Y-%m")
        except:
            raise HTTPException(status_code=400, detail="Invalid month format. Use YYYY-MM")
    else:
        period = datetime.now()

    year_month = int(period.strftime("%Y%m"))
    start_date = period.replace(day=1).date()
    if period.month == 12:
        end_date = period.replace(year=period.year+1, month=1, day=1).date() - timedelta(days=1)
    else:
        end_date = period.replace(month=period.month+1, day=1).date() - timedelta(days=1)

    try:
        async with pool.acquire() as conn:
            # Build outlet filter
            outlet_list = None
            if outlet_ids:
                outlet_list = [o.strip() for o in outlet_ids.split(',') if o.strip()]
            elif outlet_id:
                outlet_list = [outlet_id]

            # Get targets
            if outlet_list:
                targets = await conn.fetch("""
                    SELECT outlet_id, total_sales_target, gross_profit_target, house_brand_target,
                           focused_item_1_target, focused_item_2_target, focused_item_3_target,
                           pwp_target, stock_clearance_target, transaction_count_target
                    FROM "OutletTargets"
                    WHERE outlet_id = ANY($1) AND year_month = $2
                """, outlet_list, year_month)

                # Get current performance
                current = await conn.fetchrow("""
                    SELECT
                        COALESCE(SUM(total_sales), 0) as total_sales,
                        COALESCE(SUM(gross_profit), 0) as gross_profit,
                        COALESCE(SUM(house_brand_sales), 0) as house_brand,
                        COALESCE(SUM(focused_1_sales), 0) as focused_1,
                        COALESCE(SUM(focused_2_sales), 0) as focused_2,
                        COALESCE(SUM(focused_3_sales), 0) as focused_3,
                        COALESCE(SUM(pwp_sales), 0) as pwp,
                        COALESCE(SUM(clearance_sales), 0) as clearance,
                        COALESCE(SUM(transactions), 0) as transactions
                    FROM analytics.mv_outlet_daily_kpi
                    WHERE outlet_id = ANY($1)
                      AND sale_date BETWEEN $2 AND $3
                """, outlet_list, start_date, end_date)
            else:
                # All outlets
                targets = await conn.fetch("""
                    SELECT outlet_id, total_sales_target, gross_profit_target, house_brand_target,
                           focused_item_1_target, focused_item_2_target, focused_item_3_target,
                           pwp_target, stock_clearance_target, transaction_count_target
                    FROM "OutletTargets"
                    WHERE year_month = $1
                """, year_month)

                current = await conn.fetchrow("""
                    SELECT
                        COALESCE(SUM(total_sales), 0) as total_sales,
                        COALESCE(SUM(gross_profit), 0) as gross_profit,
                        COALESCE(SUM(house_brand_sales), 0) as house_brand,
                        COALESCE(SUM(focused_1_sales), 0) as focused_1,
                        COALESCE(SUM(focused_2_sales), 0) as focused_2,
                        COALESCE(SUM(focused_3_sales), 0) as focused_3,
                        COALESCE(SUM(pwp_sales), 0) as pwp,
                        COALESCE(SUM(clearance_sales), 0) as clearance,
                        COALESCE(SUM(transactions), 0) as transactions
                    FROM analytics.mv_outlet_daily_kpi
                    WHERE sale_date BETWEEN $1 AND $2
                """, start_date, end_date)

            # Sum up all targets
            total_target = {
                'total_sales': sum(float(t['total_sales_target'] or 0) for t in targets),
                'gross_profit': sum(float(t['gross_profit_target'] or 0) for t in targets),
                'house_brand': sum(float(t['house_brand_target'] or 0) for t in targets),
                'focused_1': sum(float(t['focused_item_1_target'] or 0) for t in targets),
                'focused_2': sum(float(t['focused_item_2_target'] or 0) for t in targets),
                'focused_3': sum(float(t['focused_item_3_target'] or 0) for t in targets),
                'pwp': sum(float(t['pwp_target'] or 0) for t in targets),
                'clearance': sum(float(t['stock_clearance_target'] or 0) for t in targets),
                'transactions': sum(int(t['transaction_count_target'] or 0) for t in targets)
            }

            def calc_progress(current_val, target_val):
                if not target_val or target_val == 0:
                    return None
                return round((float(current_val or 0) / float(target_val)) * 100, 1)

            result = {
                "total_sales": {
                    "target": total_target['total_sales'],
                    "current": float(current['total_sales'] or 0) if current else 0,
                    "progress": calc_progress(current['total_sales'] if current else 0, total_target['total_sales'])
                },
                "gross_profit": {
                    "target": total_target['gross_profit'],
                    "current": float(current['gross_profit'] or 0) if current else 0,
                    "progress": calc_progress(current['gross_profit'] if current else 0, total_target['gross_profit'])
                },
                "house_brand": {
                    "target": total_target['house_brand'],
                    "current": float(current['house_brand'] or 0) if current else 0,
                    "progress": calc_progress(current['house_brand'] if current else 0, total_target['house_brand'])
                },
                "focused_1": {
                    "target": total_target['focused_1'],
                    "current": float(current['focused_1'] or 0) if current else 0,
                    "progress": calc_progress(current['focused_1'] if current else 0, total_target['focused_1'])
                },
                "focused_2": {
                    "target": total_target['focused_2'],
                    "current": float(current['focused_2'] or 0) if current else 0,
                    "progress": calc_progress(current['focused_2'] if current else 0, total_target['focused_2'])
                },
                "focused_3": {
                    "target": total_target['focused_3'],
                    "current": float(current['focused_3'] or 0) if current else 0,
                    "progress": calc_progress(current['focused_3'] if current else 0, total_target['focused_3'])
                },
                "pwp": {
                    "target": total_target['pwp'],
                    "current": float(current['pwp'] or 0) if current else 0,
                    "progress": calc_progress(current['pwp'] if current else 0, total_target['pwp'])
                },
                "clearance": {
                    "target": total_target['clearance'],
                    "current": float(current['clearance'] or 0) if current else 0,
                    "progress": calc_progress(current['clearance'] if current else 0, total_target['clearance'])
                },
                "transactions": {
                    "target": total_target['transactions'],
                    "current": int(current['transactions'] or 0) if current else 0,
                    "progress": calc_progress(current['transactions'] if current else 0, total_target['transactions'])
                }
            }

            return {
                "success": True,
                "data": result,
                "period": month or period.strftime("%Y-%m"),
                "outlets_with_targets": len(targets)
            }

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error fetching outlet targets: {str(e)}")


# ============================================================================
# GEN Z MOTIVATION NOTIFICATION SYSTEM
# ============================================================================
# Designed to engage Gen Z staff with dopamine-triggering notifications:
# - Instant commission earned alerts
# - Ranking change celebrations
# - Target milestone achievements
# - Streak and badge unlocks

# Gen Z Achievement Messages - Short, punchy, emoji-rich
COMMISSION_MESSAGES = [
    "Ka-ching! 💸 You just earned RM{amount:.2f}!",
    "Money moves! 💰 RM{amount:.2f} commission unlocked!",
    "Bag secured! 🎒 +RM{amount:.2f} to your pocket!",
    "Cha-ching! 🤑 RM{amount:.2f} earned!",
    "Get that bread! 🍞 +RM{amount:.2f}!",
]

RANK_UP_MESSAGES = [
    "Level up! 🚀 You climbed to #{rank}!",
    "Main character energy! 🌟 Now ranked #{rank}!",
    "Plot twist! 📈 You're #{rank} now!",
    "Glow up alert! ✨ Jumped to #{rank}!",
    "Era upgrade! 🔥 #{rank} in the company!",
]

TARGET_MILESTONE_MESSAGES = {
    25: ["Quarter way there! ⚡ 25% of target done!", "Loading progress... 25% 📊"],
    50: ["Halfway slay! 🔥 50% target achieved!", "50% vibes only! ✌️"],
    75: ["Almost there bestie! 💪 75% complete!", "Three quarters down! 🎯"],
    100: ["TARGET CLEARED! 🏆 You did THAT!", "100% slayed! Legend status! 👑"],
    125: ["OVERACHIEVER! 🚀 125% - You're built different!", "Exceeded expectations! Extra credit! 🌟"],
    150: ["UNSTOPPABLE! 💎 150% - absolute legend!", "You're HIM/HER! 150%! 👑🔥"],
}

STREAK_MESSAGES = {
    3: "3-day streak! 🔥 Hat trick energy!",
    5: "5-day streak! ⭐ Consistency is key!",
    7: "WEEKLY WARRIOR! 🗓️ 7 days strong!",
    14: "TWO WEEK TITAN! 💪 Unstoppable!",
    30: "MONTHLY MASTER! 🏆 30-day legend!",
}

FIRST_TIME_ACHIEVEMENTS = {
    "first_sale": "First sale secured! 🎉 Welcome to the game!",
    "first_hb": "First House Brand sold! 🏠 Brand ambassador unlocked!",
    "first_100": "First RM100+ sale! 💯 Big fish energy!",
    "first_rank_top10": "TOP 10 debut! 🔟 You're in the elite now!",
    "first_target": "First target achieved! 🎯 Champion mode activated!",
}


class CommissionSaleEvent(BaseModel):
    """Event for commission-eligible sale."""
    staff_id: str
    staff_name: str
    product_name: str
    quantity: float
    sale_amount: float
    commission_earned: float
    sale_category: str  # house_brand, focused_1, etc.


class RankCheckResult(BaseModel):
    """Result of rank change detection."""
    staff_id: str
    staff_name: str
    old_rank: int
    new_rank: int
    rank_change: int  # positive = improved


@app.post("/api/v1/push/commission-earned")
async def notify_commission_earned(
    event: CommissionSaleEvent,
    api_key: str = Query(..., description="API key for authentication")
):
    """Instantly notify staff when they earn commission from a sale.

    Call this from sync service when detecting KPI-category sales.
    Gen Z staff love instant gratification - this provides dopamine hit!
    """
    expected_key = os.getenv('PUSH_API_KEY', 'flt-push-2024')
    if api_key != expected_key:
        raise HTTPException(status_code=401, detail="Invalid API key")

    if event.commission_earned <= 0:
        return {"success": False, "reason": "No commission to report"}

    # Generate engaging message
    message = random.choice(COMMISSION_MESSAGES).format(amount=event.commission_earned)

    # Add product context
    product_short = (event.product_name or "item")[:25]
    detail = f"Sold {event.quantity:.0f}x {product_short} (RM{event.sale_amount:.2f})"

    title_map = {
        "house_brand": "🏠 House Brand Commission!",
        "focused_1": "🎯 Focused 1 Commission!",
        "focused_2": "⭐ Focused 2 Commission!",
        "focused_3": "🌟 Focused 3 Commission!",
        "pwp": "🛒 PWP Commission!",
        "clearance": "📦 Clearance Commission!",
    }
    title = title_map.get(event.sale_category, "💰 Commission Earned!")

    # Save to notifications table
    try:
        async with pool.acquire() as conn:
            await conn.execute("""
                INSERT INTO kpi.notifications (staff_id, title, message, type, data)
                VALUES ($1, $2, $3, 'commission', $4)
            """, event.staff_id, title, f"{message}\n{detail}",
            json.dumps({
                "commission": event.commission_earned,
                "sale_amount": event.sale_amount,
                "product": event.product_name,
                "category": event.sale_category
            }))
    except Exception as e:
        print(f"Failed to save commission notification: {e}")

    # Send push
    result = await send_push_to_staff(
        staff_id=event.staff_id,
        title=title,
        message=f"{message}\n{detail}",
        data={
            "type": "commission",
            "commission": event.commission_earned,
            "category": event.sale_category
        }
    )

    return result


@app.post("/api/v1/push/commission-check")
async def check_and_notify_commission_changes(
    api_key: str = Query(..., description="API key for authentication")
):
    """Check for new commission earnings and notify staff.

    Schedule every 10 minutes: */10 * * * * (during business hours 9am-10pm)

    Compares current commission from mv_staff_daily_commission with cached values.
    Sends notification when staff earns new commission (delta > 0).

    This is the REAL-TIME commission notification system that Gen Z staff want!
    """
    expected_key = os.getenv('PUSH_API_KEY', 'flt-push-2024')
    if api_key != expected_key:
        raise HTTPException(status_code=401, detail="Invalid API key")

    if not PUSH_AVAILABLE:
        return {"success": False, "error": "Push notifications not available"}

    notified = []
    today = date.today()

    try:
        async with pool.acquire() as conn:
            # Ensure cache table exists
            await conn.execute("""
                CREATE TABLE IF NOT EXISTS kpi.staff_commission_cache (
                    staff_id TEXT PRIMARY KEY,
                    commission NUMERIC DEFAULT 0,
                    checked_at TIMESTAMP DEFAULT NOW()
                )
            """)

            # Get current commission from MV (today only)
            current_commissions = await conn.fetch("""
                SELECT staff_id, SUM(commission) as total_commission
                FROM analytics.mv_staff_daily_commission
                WHERE sale_date = $1
                GROUP BY staff_id
                HAVING SUM(commission) > 0
            """, today)

            # Get previous cached values
            prev_cache = await conn.fetch("""
                SELECT staff_id, commission FROM kpi.staff_commission_cache
            """)
            prev_map = {r['staff_id']: float(r['commission']) for r in prev_cache}

            # Get subscribed staff
            subscribed = await conn.fetch("""
                SELECT DISTINCT staff_id FROM kpi.push_subscriptions
            """)
            subscribed_ids = {r['staff_id'] for r in subscribed}

            # Check each staff for commission increases
            for row in current_commissions:
                staff_id = row['staff_id']
                current = float(row['total_commission'])
                previous = prev_map.get(staff_id, 0.0)
                delta = current - previous

                # Only notify if commission increased and staff is subscribed
                if delta > 0 and staff_id in subscribed_ids:
                    # Get staff name
                    staff_row = await conn.fetchrow("""
                        SELECT staff_name FROM kpi.staff_list_master WHERE staff_id = $1
                    """, staff_id)
                    staff_name = staff_row['staff_name'] if staff_row else staff_id

                    # Generate notification
                    message = random.choice(COMMISSION_MESSAGES).format(amount=delta)
                    title = "💰 Commission Earned!"

                    # Add context
                    if delta >= 10:
                        title = "🔥 Big Commission Alert!"
                    elif delta >= 5:
                        title = "⭐ Nice Commission!"

                    detail = f"Today's total: RM{current:.2f} (+RM{delta:.2f})"

                    # Save to notifications table
                    try:
                        await conn.execute("""
                            INSERT INTO kpi.notifications (staff_id, title, message, type, data)
                            VALUES ($1, $2, $3, 'commission', $4)
                        """, staff_id, title, f"{message}\n{detail}",
                        json.dumps({
                            "delta": delta,
                            "total": current,
                            "date": str(today)
                        }))
                    except Exception as e:
                        print(f"Failed to save commission notification: {e}")

                    # Send push notification
                    result = await send_push_to_staff(
                        staff_id=staff_id,
                        title=title,
                        message=f"{message}\n{detail}",
                        data={
                            "type": "commission",
                            "delta": delta,
                            "total": current
                        }
                    )

                    notified.append({
                        "staff_id": staff_id,
                        "staff_name": staff_name,
                        "delta": delta,
                        "total": current,
                        "push_sent": result.get("success", False)
                    })

            # Update cache with current values
            for row in current_commissions:
                await conn.execute("""
                    INSERT INTO kpi.staff_commission_cache (staff_id, commission, checked_at)
                    VALUES ($1, $2, NOW())
                    ON CONFLICT (staff_id) DO UPDATE SET
                        commission = EXCLUDED.commission,
                        checked_at = NOW()
                """, row['staff_id'], row['total_commission'])

            # Clean up cache at midnight (reset for new day)
            # This is handled by checking if cache has stale date
            await conn.execute("""
                DELETE FROM kpi.staff_commission_cache
                WHERE checked_at::date < $1
            """, today)

    except Exception as e:
        print(f"Commission check error: {e}")
        import traceback
        traceback.print_exc()
        return {"success": False, "error": str(e)}

    return {
        "success": True,
        "checked": len(current_commissions) if 'current_commissions' in dir() else 0,
        "notified": len(notified),
        "details": notified
    }


@app.post("/api/v1/push/rank-change-check")
async def check_and_notify_rank_changes(
    api_key: str = Query(..., description="API key for authentication")
):
    """Check for significant rank changes and notify staff.

    Schedule hourly: 0 * * * * (every hour during business hours)

    Triggers notification when:
    - Staff moves up 3+ positions
    - Staff enters top 10
    - Staff becomes #1
    """
    expected_key = os.getenv('PUSH_API_KEY', 'flt-push-2024')
    if api_key != expected_key:
        raise HTTPException(status_code=401, detail="Invalid API key")

    if not PUSH_AVAILABLE:
        return {"success": False, "error": "Push notifications not available"}

    notified = []

    try:
        async with pool.acquire() as conn:
            # Get current rankings vs cached previous rankings
            # We store previous rankings in kpi.staff_rank_cache

            # Ensure cache table exists
            await conn.execute("""
                CREATE TABLE IF NOT EXISTS kpi.staff_rank_cache (
                    staff_id TEXT PRIMARY KEY,
                    company_rank INTEGER,
                    checked_at TIMESTAMP DEFAULT NOW()
                )
            """)

            # Get current rankings from MV
            current_ranks = await conn.fetch("""
                SELECT staff_id, company_rank_sales as company_rank
                FROM analytics.mv_staff_rankings
                WHERE company_rank_sales IS NOT NULL
            """)

            # Get previous rankings
            prev_ranks = await conn.fetch("""
                SELECT staff_id, company_rank FROM kpi.staff_rank_cache
            """)
            prev_map = {r['staff_id']: r['company_rank'] for r in prev_ranks}

            # Check subscribed staff for significant changes
            subscribed = await conn.fetch("""
                SELECT DISTINCT staff_id FROM kpi.push_subscriptions
            """)
            subscribed_ids = {r['staff_id'] for r in subscribed}

            for row in current_ranks:
                staff_id = row['staff_id']
                new_rank = row['company_rank']
                old_rank = prev_map.get(staff_id)

                if staff_id not in subscribed_ids:
                    continue

                if old_rank is None:
                    # First time seeing this staff, just cache
                    continue

                rank_change = old_rank - new_rank  # positive = improved

                # Notify if significant improvement
                should_notify = False
                special_message = None

                if new_rank == 1 and old_rank != 1:
                    should_notify = True
                    special_message = "👑 CROWNED #1! You're the TOP performer! 👑"
                elif new_rank <= 10 and old_rank > 10:
                    should_notify = True
                    special_message = f"🏅 TOP 10 ENTRY! You broke into #{new_rank}!"
                elif rank_change >= 5:
                    should_notify = True
                    special_message = f"🚀 MASSIVE CLIMB! Jumped {rank_change} spots to #{new_rank}!"
                elif rank_change >= 3:
                    should_notify = True

                if should_notify:
                    message = special_message or random.choice(RANK_UP_MESSAGES).format(rank=new_rank)

                    try:
                        await send_push_to_staff(
                            staff_id=staff_id,
                            title="📈 Rank Up!",
                            message=message,
                            data={
                                "type": "rank_change",
                                "old_rank": old_rank,
                                "new_rank": new_rank,
                                "change": rank_change
                            }
                        )
                        notified.append({
                            "staff_id": staff_id,
                            "old_rank": old_rank,
                            "new_rank": new_rank
                        })
                    except Exception as e:
                        print(f"Failed to notify {staff_id}: {e}")

            # Update cache with current rankings
            for row in current_ranks:
                await conn.execute("""
                    INSERT INTO kpi.staff_rank_cache (staff_id, company_rank, checked_at)
                    VALUES ($1, $2, NOW())
                    ON CONFLICT (staff_id) DO UPDATE SET
                        company_rank = EXCLUDED.company_rank,
                        checked_at = NOW()
                """, row['staff_id'], row['company_rank'])

        return {
            "success": True,
            "notified_count": len(notified),
            "notified": notified,
            "timestamp": datetime.now().isoformat()
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Rank check failed: {str(e)}")


@app.post("/api/v1/push/milestone-check")
async def check_and_notify_milestones(
    api_key: str = Query(..., description="API key for authentication")
):
    """Check milestone progress and notify - both achievements AND falling behind.

    Logic:
    - Staff achieved milestone EARLY → Celebrate! "X days ahead of schedule!"
    - Staff BEHIND expected milestone → Motivate! "Time to catch up!"

    E.g., Day 15 of 30 = expected 50%.
    - If staff at 60% → "Ahead! Keep it up!"
    - If staff at 35% → "15% behind pace, you can catch up!"

    Milestones: 25%, 50%, 75%, 100%
    Only notifies once per milestone per month (for achievements).
    Behind-pace alerts sent once per week max.
    """
    expected_key = os.getenv('PUSH_API_KEY', 'flt-push-2024')
    if api_key != expected_key:
        raise HTTPException(status_code=401, detail="Invalid API key")

    if not PUSH_AVAILABLE:
        return {"success": False, "error": "Push notifications not available"}

    today = date.today()
    year_month = int(today.strftime('%Y%m'))

    # Calculate expected progress based on day of month
    days_in_month = (date(today.year, today.month + 1 if today.month < 12 else 1, 1) - timedelta(days=1)).day if today.month < 12 else 31
    day_of_month = today.day
    expected_progress_pct = (day_of_month / days_in_month) * 100

    notified_ahead = []
    notified_behind = []

    # Gen Z motivational messages for falling behind
    BEHIND_MESSAGES = [
        "Time to level up! You've got this! 💪",
        "Comeback arc starts NOW! 🔥",
        "Still time to turn it around! Let's go! ⚡",
        "Challenge accepted? Let's close that gap! 🎯",
        "Plot twist incoming - make it happen! 📈",
    ]

    try:
        async with pool.acquire() as conn:
            # Ensure milestone tracking table exists
            await conn.execute("""
                CREATE TABLE IF NOT EXISTS kpi.milestone_notifications (
                    staff_id TEXT,
                    year_month INTEGER,
                    milestone INTEGER,
                    notified_at TIMESTAMP DEFAULT NOW(),
                    PRIMARY KEY (staff_id, year_month, milestone)
                )
            """)

            # Ensure behind-pace tracking table exists (to avoid spamming)
            await conn.execute("""
                CREATE TABLE IF NOT EXISTS kpi.behind_pace_notifications (
                    staff_id TEXT,
                    year_month INTEGER,
                    last_notified DATE,
                    PRIMARY KEY (staff_id, year_month)
                )
            """)

            # Get subscribed staff with targets and MTD
            staff_progress = await conn.fetch("""
                WITH subscribed AS (
                    SELECT DISTINCT staff_id FROM kpi.push_subscriptions
                ),
                mtd AS (
                    SELECT staff_id, SUM(total_sales) as mtd_sales
                    FROM analytics.mv_staff_daily_kpi
                    WHERE EXTRACT(YEAR FROM sale_date) = $1
                      AND EXTRACT(MONTH FROM sale_date) = $2
                    GROUP BY staff_id
                )
                SELECT
                    s.staff_id,
                    COALESCE(m.mtd_sales, 0) as mtd_sales,
                    COALESCE(t.total_sales_target, 0) as target
                FROM subscribed s
                LEFT JOIN mtd m ON s.staff_id = m.staff_id
                LEFT JOIN "KPITargets" t ON s.staff_id = t.salesman_id AND t.year_month = $3
                WHERE COALESCE(t.total_sales_target, 0) > 0
            """, today.year, today.month, year_month)

            # Get already-notified milestones (for achievements)
            notified_milestones = await conn.fetch("""
                SELECT staff_id, milestone FROM kpi.milestone_notifications
                WHERE year_month = $1
            """, year_month)
            notified_set = {(r['staff_id'], r['milestone']) for r in notified_milestones}

            # Get recent behind-pace notifications (max once per week)
            recent_behind = await conn.fetch("""
                SELECT staff_id, last_notified FROM kpi.behind_pace_notifications
                WHERE year_month = $1 AND last_notified > $2
            """, year_month, today - timedelta(days=7))
            recent_behind_set = {r['staff_id'] for r in recent_behind}

            # Milestones to check
            milestones = [25, 50, 75, 100]

            for staff in staff_progress:
                staff_id = staff['staff_id']
                mtd = float(staff['mtd_sales'] or 0)
                target = float(staff['target'] or 0)

                if target <= 0:
                    continue

                progress_pct = (mtd / target) * 100
                gap_pct = expected_progress_pct - progress_pct

                # === CHECK FOR ACHIEVED MILESTONES (EARLY) ===
                for milestone in milestones:
                    if progress_pct >= milestone and (staff_id, milestone) not in notified_set:
                        # Calculate if achieved early
                        expected_day_for_milestone = (milestone / 100) * days_in_month
                        days_ahead = expected_day_for_milestone - day_of_month

                        if days_ahead >= 1:
                            # Achieved EARLY - celebrate!
                            messages = TARGET_MILESTONE_MESSAGES.get(milestone, [f"{milestone}% achieved! 🎯"])
                            message = random.choice(messages) if isinstance(messages, list) else messages
                            ahead_msg = f"{int(days_ahead)} days ahead of schedule!"

                            if milestone >= 100:
                                title = "🏆 TARGET SMASHED EARLY!"
                            else:
                                title = "⚡ Ahead of Target!"
                            full_message = f"{message}\n{ahead_msg}"
                        else:
                            # Achieved on time or slightly late - still celebrate
                            messages = TARGET_MILESTONE_MESSAGES.get(milestone, [f"{milestone}% achieved! 🎯"])
                            message = random.choice(messages) if isinstance(messages, list) else messages

                            if milestone >= 100:
                                title = "🏆 TARGET ACHIEVED!"
                            else:
                                title = "🎯 Milestone Reached!"
                            full_message = message

                        try:
                            await send_push_to_staff(
                                staff_id=staff_id,
                                title=title,
                                message=full_message,
                                data={
                                    "type": "milestone_achieved",
                                    "milestone": milestone,
                                    "days_ahead": max(0, days_ahead),
                                    "mtd": mtd,
                                    "target": target
                                }
                            )

                            # Record notification
                            await conn.execute("""
                                INSERT INTO kpi.milestone_notifications (staff_id, year_month, milestone)
                                VALUES ($1, $2, $3)
                                ON CONFLICT DO NOTHING
                            """, staff_id, year_month, milestone)

                            notified_ahead.append({
                                "staff_id": staff_id,
                                "milestone": milestone,
                                "status": "achieved",
                                "days_ahead": round(max(0, days_ahead), 1)
                            })
                        except Exception as e:
                            print(f"Failed to notify {staff_id} for {milestone}%: {e}")

                # === CHECK FOR FALLING BEHIND ===
                # Only alert if significantly behind (>10% gap) and not recently notified
                if gap_pct >= 10 and staff_id not in recent_behind_set:
                    # Find which milestone they should have hit by now
                    expected_milestone = None
                    for m in milestones:
                        if expected_progress_pct >= m and progress_pct < m:
                            expected_milestone = m
                            break

                    if expected_milestone or gap_pct >= 15:
                        motivation = random.choice(BEHIND_MESSAGES)
                        gap_amount = (gap_pct / 100) * target

                        if expected_milestone:
                            title = f"📊 {expected_milestone}% Check-In"
                            msg = f"You're at {progress_pct:.0f}% (expected {expected_progress_pct:.0f}%)\nGap: RM{gap_amount:,.0f} | {motivation}"
                        else:
                            title = "📊 Progress Check"
                            msg = f"Currently {progress_pct:.0f}% vs expected {expected_progress_pct:.0f}%\n{motivation}"

                        try:
                            await send_push_to_staff(
                                staff_id=staff_id,
                                title=title,
                                message=msg,
                                data={
                                    "type": "behind_pace",
                                    "current_pct": progress_pct,
                                    "expected_pct": expected_progress_pct,
                                    "gap_pct": gap_pct,
                                    "mtd": mtd,
                                    "target": target
                                }
                            )

                            # Record to avoid spamming
                            await conn.execute("""
                                INSERT INTO kpi.behind_pace_notifications (staff_id, year_month, last_notified)
                                VALUES ($1, $2, $3)
                                ON CONFLICT (staff_id, year_month) DO UPDATE SET last_notified = $3
                            """, staff_id, year_month, today)

                            notified_behind.append({
                                "staff_id": staff_id,
                                "status": "behind",
                                "current_pct": round(progress_pct, 1),
                                "expected_pct": round(expected_progress_pct, 1),
                                "gap_pct": round(gap_pct, 1)
                            })
                        except Exception as e:
                            print(f"Failed to notify {staff_id} for behind pace: {e}")

        return {
            "success": True,
            "notified_ahead": len(notified_ahead),
            "notified_behind": len(notified_behind),
            "details": {
                "ahead": notified_ahead,
                "behind": notified_behind
            },
            "expected_progress": round(expected_progress_pct, 1),
            "day_of_month": day_of_month,
            "days_in_month": days_in_month,
            "timestamp": datetime.now().isoformat()
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Milestone check failed: {str(e)}")


@app.post("/api/v1/push/streak-check")
async def check_and_notify_streaks(
    api_key: str = Query(..., description="API key for authentication")
):
    """Check for weekly consistency and consecutive week streaks.

    Schedule: Daily at 9pm MYT (part of daily milestones)

    Tracks:
    1. Weekly consistency: Days on target this week (accounts for 1-2 days off)
    2. Weekly streak: Consecutive weeks meeting weekly target (MTD pace)

    Staff typically work 5-6 days/week, so we celebrate:
    - 4/5 days on target: "Great week! 4 days on target!"
    - 5/5 days on target: "Perfect week! All working days crushed!"
    - 2-week streak: Consecutive weeks on pace
    - 4-week streak: Monthly champion
    """
    expected_key = os.getenv('PUSH_API_KEY', 'flt-push-2024')
    if api_key != expected_key:
        raise HTTPException(status_code=401, detail="Invalid API key")

    if not PUSH_AVAILABLE:
        return {"success": False, "error": "Push notifications not available"}

    today = date.today()
    year_month = int(today.strftime('%Y%m'))
    notified = []

    # Only run on Saturdays (end of work week) or Sundays
    # This gives weekly summary rather than daily spam
    if today.weekday() not in [5, 6]:  # Saturday = 5, Sunday = 6
        return {
            "success": True,
            "message": "Streak check only runs on weekends (end of work week)",
            "notified_count": 0
        }

    try:
        async with pool.acquire() as conn:
            # Create streak tracking table
            await conn.execute("""
                CREATE TABLE IF NOT EXISTS kpi.weekly_streak_notifications (
                    staff_id TEXT,
                    week_start DATE,
                    notified_at TIMESTAMP DEFAULT NOW(),
                    PRIMARY KEY (staff_id, week_start)
                )
            """)

            # Get week boundaries (Mon-Sat)
            week_start = today - timedelta(days=today.weekday())  # Monday
            week_end = today

            # Get subscribed staff with targets
            subscribed = await conn.fetch("""
                SELECT DISTINCT ps.staff_id, m.staff_name,
                       COALESCE(t.total_sales_target, 0) as monthly_target
                FROM kpi.push_subscriptions ps
                JOIN kpi.staff_list_master m ON ps.staff_id = m.staff_id
                LEFT JOIN "KPITargets" t ON ps.staff_id = t.salesman_id AND t.year_month = $1
            """, year_month)

            # Get already notified this week
            notified_this_week = await conn.fetch("""
                SELECT staff_id FROM kpi.weekly_streak_notifications
                WHERE week_start = $1
            """, week_start)
            already_notified = {r['staff_id'] for r in notified_this_week}

            for staff in subscribed:
                staff_id = staff['staff_id']
                staff_name = staff['staff_name']
                monthly_target = float(staff['monthly_target'] or 0)

                if staff_id in already_notified:
                    continue

                if monthly_target <= 0:
                    continue

                # Calculate daily target (assuming ~26 working days)
                daily_target = monthly_target / 26

                # Get this week's daily performance
                week_data = await conn.fetch("""
                    SELECT sale_date, SUM(total_sales) as daily_total
                    FROM analytics.mv_staff_daily_kpi
                    WHERE staff_id = $1
                      AND sale_date >= $2
                      AND sale_date <= $3
                    GROUP BY sale_date
                """, staff_id, week_start, week_end)

                working_days = len(week_data)
                days_on_target = sum(1 for d in week_data if float(d['daily_total'] or 0) >= daily_target)

                # Only celebrate if they worked at least 3 days this week
                if working_days >= 3:
                    if working_days == days_on_target and days_on_target >= 4:
                        # Perfect week!
                        title = "🌟 Perfect Week!"
                        message = f"All {working_days} working days on target! Absolute beast mode!"
                        notif_type = "perfect_week"
                    elif days_on_target >= 4:
                        # Great week
                        title = "✨ Great Week!"
                        message = f"{days_on_target}/{working_days} days on target! Solid consistency!"
                        notif_type = "great_week"
                    else:
                        # Not enough days on target - skip notification
                        continue

                    try:
                        await send_push_to_staff(
                            staff_id=staff_id,
                            title=title,
                            message=message,
                            data={"type": notif_type, "days_worked": working_days, "days_on_target": days_on_target}
                        )

                        # Save to notifications
                        await conn.execute("""
                            INSERT INTO kpi.notifications (staff_id, title, message, type, data)
                            VALUES ($1, $2, $3, 'weekly_consistency', $4)
                        """, staff_id, title, message, json.dumps({
                            "days_worked": working_days,
                            "days_on_target": days_on_target,
                            "week_start": str(week_start)
                        }))

                        # Mark as notified
                        await conn.execute("""
                            INSERT INTO kpi.weekly_streak_notifications (staff_id, week_start)
                            VALUES ($1, $2) ON CONFLICT DO NOTHING
                        """, staff_id, week_start)

                        notified.append({
                            "staff_id": staff_id,
                            "staff_name": staff_name,
                            "days_worked": working_days,
                            "days_on_target": days_on_target,
                            "type": notif_type
                        })
                    except Exception as e:
                        print(f"Failed to notify weekly consistency for {staff_id}: {e}")

        return {
            "success": True,
            "notified_count": len(notified),
            "notified": notified,
            "timestamp": datetime.now().isoformat()
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Streak check failed: {str(e)}")


@app.get("/api/v1/achievements/{staff_id}")
async def get_staff_achievements(staff_id: str):
    """Get all achievements/badges for a staff member.

    Returns earned badges and progress toward next ones.
    Great for gamification display in the app.
    """
    today = date.today()
    year_month = int(today.strftime('%Y%m'))

    try:
        async with pool.acquire() as conn:
            # Get various achievement data
            stats = await conn.fetchrow("""
                WITH mtd AS (
                    SELECT
                        SUM(total_sales) as mtd_sales,
                        SUM(house_brand_sales) as mtd_hb,
                        SUM(transactions) as mtd_trans,
                        COUNT(DISTINCT sale_date) as active_days
                    FROM analytics.mv_staff_daily_kpi
                    WHERE staff_id = $1
                      AND EXTRACT(YEAR FROM sale_date) = $2
                      AND EXTRACT(MONTH FROM sale_date) = $3
                ),
                all_time AS (
                    SELECT
                        SUM(total_sales) as total_sales,
                        COUNT(DISTINCT sale_date) as total_days
                    FROM analytics.mv_staff_daily_kpi
                    WHERE staff_id = $1
                ),
                ranking AS (
                    SELECT company_rank_sales FROM analytics.mv_staff_rankings
                    WHERE staff_id = $1
                )
                SELECT
                    COALESCE(m.mtd_sales, 0) as mtd_sales,
                    COALESCE(m.mtd_hb, 0) as mtd_hb,
                    COALESCE(m.mtd_trans, 0) as mtd_trans,
                    COALESCE(m.active_days, 0) as active_days,
                    COALESCE(a.total_sales, 0) as total_sales,
                    COALESCE(a.total_days, 0) as total_days,
                    r.company_rank_sales as current_rank
                FROM mtd m, all_time a
                LEFT JOIN ranking r ON true
            """, staff_id, today.year, today.month)

            # Get target
            target = await conn.fetchrow("""
                SELECT total_sales_target FROM "KPITargets"
                WHERE salesman_id = $1 AND year_month = $2
            """, staff_id, year_month)

            target_sales = float(target['total_sales_target'] or 0) if target else 0

            achievements = []

            # Sales Milestones
            total_sales = float(stats['total_sales'] or 0)
            if total_sales >= 1000000:
                achievements.append({"id": "million_club", "name": "Million Club", "icon": "💎", "desc": "RM1M+ lifetime sales"})
            elif total_sales >= 500000:
                achievements.append({"id": "half_million", "name": "Half Million Hero", "icon": "🏆", "desc": "RM500K+ lifetime sales"})
            elif total_sales >= 100000:
                achievements.append({"id": "100k_club", "name": "100K Club", "icon": "🥇", "desc": "RM100K+ lifetime sales"})

            # Ranking Achievements
            current_rank = stats['current_rank']
            if current_rank == 1:
                achievements.append({"id": "top1", "name": "The Champion", "icon": "👑", "desc": "Currently #1 in company"})
            elif current_rank and current_rank <= 3:
                achievements.append({"id": "top3", "name": "Podium Finisher", "icon": "🥈", "desc": "Top 3 in company"})
            elif current_rank and current_rank <= 10:
                achievements.append({"id": "top10", "name": "Elite 10", "icon": "🔟", "desc": "Top 10 in company"})

            # Consistency Achievements
            active_days = stats['active_days']
            if active_days >= 25:
                achievements.append({"id": "consistent", "name": "Iron Will", "icon": "💪", "desc": "25+ active days this month"})
            elif active_days >= 20:
                achievements.append({"id": "regular", "name": "Regular", "icon": "📅", "desc": "20+ active days this month"})

            # Target Achievements
            mtd_sales = float(stats['mtd_sales'] or 0)
            if target_sales > 0:
                progress = (mtd_sales / target_sales) * 100
                if progress >= 150:
                    achievements.append({"id": "overachiever", "name": "Overachiever", "icon": "🚀", "desc": "150%+ of target"})
                elif progress >= 100:
                    achievements.append({"id": "target_met", "name": "Target Met", "icon": "🎯", "desc": "Monthly target achieved"})

            return {
                "success": True,
                "staff_id": staff_id,
                "achievements": achievements,
                "stats": {
                    "mtd_sales": round(mtd_sales, 2),
                    "total_sales": round(total_sales, 2),
                    "current_rank": current_rank,
                    "active_days": active_days,
                    "target_progress": round((mtd_sales / target_sales * 100), 1) if target_sales > 0 else None
                }
            }

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error fetching achievements: {str(e)}")


# ============================================================================
# Twice-Daily Progress Check (12pm and 6pm)
# ============================================================================

# Midday check-in messages (12pm) - encouraging, motivating to push harder
MIDDAY_MESSAGES_AHEAD = [
    "Killing it! 🔥 Keep this energy going!",
    "You're on fire today! 💪 Maintain the momentum!",
    "Slaying the daily target! ✨ Don't stop now!",
    "Main character energy! 🌟 Afternoon = more wins!",
]

MIDDAY_MESSAGES_BEHIND = [
    "Afternoon comeback loading... 💪",
    "Second half is YOUR half! 🏃",
    "Plot twist incoming! 📈",
    "Time to turn it around! ⚡",
]

MIDDAY_MESSAGES_ON_TRACK = [
    "On track! Solid morning! 👍",
    "Looking good! Keep the pace! 🎯",
    "Nice steady progress! 📊",
]

# End of day messages (6pm) - celebrating or encouraging for tomorrow
EOD_MESSAGES_CRUSHED = [
    "DAILY TARGET CRUSHED! 🏆 Legend status!",
    "You did THAT today! 💅 Rest well, champion!",
    "Target? Destroyed! 💥 Tomorrow we go again!",
]

EOD_MESSAGES_CLOSE = [
    "So close! 🎯 A few more sales and you're there!",
    "Almost at target! Final push! 💪",
    "You've got this! Just a bit more! ⚡",
]

EOD_MESSAGES_BEHIND = [
    "Tomorrow is a new day! 🌅 Fresh start incoming!",
    "Every day is a chance to level up! 📈",
    "Rest up, comeback arc starts tomorrow! 💪",
]


@app.post("/api/v1/push/daily-progress-check")
async def send_daily_progress_check(
    time_of_day: str = Query(..., description="midday or evening"),
    api_key: str = Query(..., description="API key for authentication")
):
    """Send twice-daily progress check comparing today's sales vs daily target.

    Schedule:
    - Midday (12pm): 0 12 * * 1-6 (Mon-Sat at noon)
    - Evening (6pm): 0 18 * * 1-6 (Mon-Sat at 6pm)

    Compares today's sales against daily target (monthly target / days in month).
    """
    expected_key = os.getenv('PUSH_API_KEY', 'flt-push-2024')
    if api_key != expected_key:
        raise HTTPException(status_code=401, detail="Invalid API key")

    if not PUSH_AVAILABLE:
        return {"success": False, "error": "Push notifications not available"}

    if time_of_day not in ["midday", "evening"]:
        raise HTTPException(status_code=400, detail="time_of_day must be 'midday' or 'evening'")

    today = date.today()
    year_month = int(today.strftime('%Y%m'))

    # Calculate daily target from monthly target
    days_in_month = (date(today.year, today.month + 1 if today.month < 12 else 1, 1) - timedelta(days=1)).day if today.month < 12 else 31
    # Assume ~26 working days per month for realistic daily target
    working_days = 26

    sent_count = 0
    errors = []

    try:
        async with pool.acquire() as conn:
            # Get subscribed staff with today's sales and monthly target
            staff_data = await conn.fetch("""
                WITH subscribed AS (
                    SELECT DISTINCT staff_id FROM kpi.push_subscriptions
                ),
                today_sales AS (
                    SELECT staff_id, SUM(total_sales) as today_total
                    FROM analytics.mv_staff_daily_kpi
                    WHERE sale_date = $1
                    GROUP BY staff_id
                ),
                mtd_sales AS (
                    SELECT staff_id, SUM(total_sales) as mtd_total
                    FROM analytics.mv_staff_daily_kpi
                    WHERE EXTRACT(YEAR FROM sale_date) = $2
                      AND EXTRACT(MONTH FROM sale_date) = $3
                    GROUP BY staff_id
                )
                SELECT
                    s.staff_id,
                    COALESCE(d.today_total, 0) as today_sales,
                    COALESCE(m.mtd_total, 0) as mtd_sales,
                    COALESCE(t.total_sales_target, 0) as monthly_target
                FROM subscribed s
                LEFT JOIN today_sales d ON s.staff_id = d.staff_id
                LEFT JOIN mtd_sales m ON s.staff_id = m.staff_id
                LEFT JOIN "KPITargets" t ON s.staff_id = t.salesman_id AND t.year_month = $4
            """, today, today.year, today.month, year_month)

            for staff in staff_data:
                staff_id = staff['staff_id']
                today_sales = float(staff['today_sales'] or 0)
                mtd_sales = float(staff['mtd_sales'] or 0)
                monthly_target = float(staff['monthly_target'] or 0)

                # Calculate daily target
                daily_target = monthly_target / working_days if monthly_target > 0 else 0

                if daily_target <= 0:
                    # No target set - still send a simple update
                    if today_sales > 0:
                        title = "📊 Today's Progress"
                        msg = f"Today: RM{today_sales:,.0f} | MTD: RM{mtd_sales:,.0f}"
                    else:
                        continue  # Skip if no target and no sales
                else:
                    # Calculate progress against daily target
                    daily_progress_pct = (today_sales / daily_target) * 100

                    if time_of_day == "midday":
                        # Midday: expect ~50% of daily target by noon
                        if daily_progress_pct >= 60:
                            title = "🔥 Midday Check-In"
                            motivation = random.choice(MIDDAY_MESSAGES_AHEAD)
                        elif daily_progress_pct >= 40:
                            title = "📊 Midday Check-In"
                            motivation = random.choice(MIDDAY_MESSAGES_ON_TRACK)
                        else:
                            title = "💪 Midday Check-In"
                            motivation = random.choice(MIDDAY_MESSAGES_BEHIND)

                        gap = daily_target - today_sales
                        if gap > 0:
                            msg = f"Today: RM{today_sales:,.0f} / RM{daily_target:,.0f} ({daily_progress_pct:.0f}%)\nGap: RM{gap:,.0f} | {motivation}"
                        else:
                            msg = f"Today: RM{today_sales:,.0f} / RM{daily_target:,.0f} ({daily_progress_pct:.0f}%)\n{motivation}"

                    else:  # evening
                        if daily_progress_pct >= 100:
                            title = "🏆 Daily Target Achieved!"
                            motivation = random.choice(EOD_MESSAGES_CRUSHED)
                            msg = f"Today: RM{today_sales:,.0f} ({daily_progress_pct:.0f}% of daily target!)\n{motivation}"
                        elif daily_progress_pct >= 80:
                            title = "⚡ Almost There!"
                            motivation = random.choice(EOD_MESSAGES_CLOSE)
                            gap = daily_target - today_sales
                            msg = f"Today: RM{today_sales:,.0f} / RM{daily_target:,.0f} ({daily_progress_pct:.0f}%)\nJust RM{gap:,.0f} more! {motivation}"
                        else:
                            title = "🌙 Day's Wrap Up"
                            motivation = random.choice(EOD_MESSAGES_BEHIND)
                            msg = f"Today: RM{today_sales:,.0f} / RM{daily_target:,.0f} ({daily_progress_pct:.0f}%)\n{motivation}"

                try:
                    await send_push_to_staff(
                        staff_id=staff_id,
                        title=title,
                        message=msg,
                        data={
                            "type": f"daily_progress_{time_of_day}",
                            "today_sales": today_sales,
                            "daily_target": daily_target,
                            "mtd_sales": mtd_sales,
                            "monthly_target": monthly_target
                        }
                    )
                    sent_count += 1
                except Exception as e:
                    errors.append(f"{staff_id}: {str(e)}")

        return {
            "success": True,
            "time_of_day": time_of_day,
            "sent_count": sent_count,
            "errors": errors[:10] if errors else [],
            "timestamp": datetime.now().isoformat()
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Daily progress check failed: {str(e)}")


@app.post("/api/v1/push/daily-rank-check")
async def run_daily_rank_and_milestone_check(
    api_key: str = Query(..., description="API key for authentication")
):
    """Daily check for rank changes and ahead-of-target milestones.

    Schedule: 0 21 * * * (daily at 9pm, after business hours)

    Runs:
    - Rank change check (notify significant rank improvements)
    - Ahead-of-target milestone check (notify early achievers)
    """
    expected_key = os.getenv('PUSH_API_KEY', 'flt-push-2024')
    if api_key != expected_key:
        raise HTTPException(status_code=401, detail="Invalid API key")

    results = {}

    # Run rank check
    try:
        rank_result = await check_and_notify_rank_changes(api_key=api_key)
        results["rank_check"] = {"success": True, "notified": rank_result.get("notified_count", 0)}
    except Exception as e:
        results["rank_check"] = {"success": False, "error": str(e)}

    # Run milestone check (both ahead and behind)
    try:
        milestone_result = await check_and_notify_milestones(api_key=api_key)
        results["milestone_check"] = {
            "success": True,
            "notified_ahead": milestone_result.get("notified_ahead", 0),
            "notified_behind": milestone_result.get("notified_behind", 0)
        }
    except Exception as e:
        results["milestone_check"] = {"success": False, "error": str(e)}

    return {
        "success": True,
        "results": results,
        "timestamp": datetime.now().isoformat()
    }


# ============================================================================
# THRESHOLD-BASED COMMISSION ALERTS
# ============================================================================
# Instead of notifying every small commission increase, we notify at meaningful
# thresholds to avoid spam while still providing dopamine hits.

# Average commission ~RM400/month = ~RM15/day
# Thresholds set to celebrate meaningful achievements
COMMISSION_THRESHOLDS = [
    (10, "Nice Start!", "RM10+ commission today - keep it up!"),
    (20, "Double Digits!", "RM20+ commission - you're rolling!"),
    (30, "On a Roll!", "RM30+ commission - above average day!"),
    (50, "Half Century!", "RM50+ commission - crushing it!"),
    (100, "TRIPLE DIGITS!", "RM100+ commission - absolute legend!"),
]


@app.post("/api/v1/push/commission-threshold-check")
async def check_commission_thresholds(
    api_key: str = Query(..., description="API key for authentication")
):
    """Check cumulative daily commission against thresholds.

    Notifies when staff crosses RM50, RM100, RM200, RM500, RM1000 thresholds.
    Each threshold is notified ONCE per day to avoid spam.

    Schedule: Every 30 min during business hours (9am-9pm MYT)
    """
    expected_key = os.getenv('PUSH_API_KEY', 'flt-push-2024')
    if api_key != expected_key:
        raise HTTPException(status_code=401, detail="Invalid API key")

    if not PUSH_AVAILABLE:
        return {"success": False, "error": "Push notifications not available"}

    today = date.today()
    notified = []

    try:
        async with pool.acquire() as conn:
            # Create threshold tracking table if not exists
            await conn.execute("""
                CREATE TABLE IF NOT EXISTS kpi.commission_threshold_notifications (
                    staff_id TEXT,
                    threshold INTEGER,
                    sale_date DATE,
                    notified_at TIMESTAMP DEFAULT NOW(),
                    PRIMARY KEY (staff_id, threshold, sale_date)
                )
            """)

            # Clean up old records (keep only last 7 days)
            await conn.execute("""
                DELETE FROM kpi.commission_threshold_notifications
                WHERE sale_date < CURRENT_DATE - INTERVAL '7 days'
            """)

            # Get today's commission for all staff from cache (faster than MV)
            staff_commissions = await conn.fetch("""
                SELECT
                    c.staff_id,
                    m.staff_name,
                    COALESCE(
                        (SELECT SUM(commission) FROM analytics.mv_staff_daily_commission
                         WHERE staff_id = c.staff_id AND sale_date = CURRENT_DATE), 0
                    ) as today_commission
                FROM kpi.push_subscriptions c
                JOIN kpi.staff_list_master m ON c.staff_id = m.staff_id
                GROUP BY c.staff_id, m.staff_name
            """)

            # Get already-notified thresholds for today
            notified_thresholds = await conn.fetch("""
                SELECT staff_id, threshold
                FROM kpi.commission_threshold_notifications
                WHERE sale_date = $1
            """, today)
            notified_set = {(r['staff_id'], r['threshold']) for r in notified_thresholds}

            # Check each staff against thresholds
            for staff in staff_commissions:
                staff_id = staff['staff_id']
                staff_name = staff['staff_name']
                commission = float(staff['today_commission'] or 0)

                # Check each threshold
                for threshold, title_text, message_text in COMMISSION_THRESHOLDS:
                    # Skip if already notified this threshold today
                    if (staff_id, threshold) in notified_set:
                        continue

                    # Check if commission crossed this threshold
                    if commission >= threshold:
                        # Send notification
                        full_title = f"💰 {title_text}"
                        full_message = f"{message_text}\nTotal today: RM{commission:.2f}"

                        # Save to notifications table
                        try:
                            await conn.execute("""
                                INSERT INTO kpi.notifications (staff_id, title, message, type, data)
                                VALUES ($1, $2, $3, 'commission_threshold', $4)
                            """, staff_id, full_title, full_message, json.dumps({
                                "threshold": threshold,
                                "commission": commission,
                                "date": str(today)
                            }))
                        except Exception as e:
                            print(f"Failed to save notification: {e}")

                        # Send push
                        result = await send_push_to_staff(
                            staff_id=staff_id,
                            title=full_title,
                            message=full_message,
                            data={"type": "commission_threshold", "threshold": threshold}
                        )

                        # Mark as notified
                        await conn.execute("""
                            INSERT INTO kpi.commission_threshold_notifications
                            (staff_id, threshold, sale_date)
                            VALUES ($1, $2, $3)
                            ON CONFLICT DO NOTHING
                        """, staff_id, threshold, today)

                        notified.append({
                            "staff_id": staff_id,
                            "staff_name": staff_name,
                            "threshold": threshold,
                            "commission": commission,
                            "push_sent": result.get("success", False)
                        })

            return {
                "success": True,
                "checked": len(staff_commissions),
                "notified": len(notified),
                "details": notified
            }

    except Exception as e:
        print(f"Commission threshold check error: {e}")
        import traceback
        traceback.print_exc()
        return {"success": False, "error": str(e)}


# ============================================================================
# FIRST-TIME ACHIEVEMENTS
# ============================================================================
# Celebrate "firsts" - these only happen once per staff member's career.

@app.post("/api/v1/push/first-time-achievements")
async def check_first_time_achievements(
    api_key: str = Query(..., description="API key for authentication")
):
    """Check and notify first-time achievements.

    Achievements tracked:
    - first_sale: First ever transaction
    - first_100: First RM100+ single transaction
    - first_hb: First House Brand sale
    - first_top10: First time entering company Top 10
    - first_target: First time hitting monthly target

    Each achievement is notified ONCE per staff member (lifetime).

    Schedule: Daily at 9pm MYT (as part of cron_milestones.py)
    """
    expected_key = os.getenv('PUSH_API_KEY', 'flt-push-2024')
    if api_key != expected_key:
        raise HTTPException(status_code=401, detail="Invalid API key")

    if not PUSH_AVAILABLE:
        return {"success": False, "error": "Push notifications not available"}

    notified = []

    try:
        async with pool.acquire() as conn:
            # Create first-time achievements tracking table
            await conn.execute("""
                CREATE TABLE IF NOT EXISTS kpi.first_time_achievements (
                    staff_id TEXT,
                    achievement_id TEXT,
                    achieved_at TIMESTAMP DEFAULT NOW(),
                    PRIMARY KEY (staff_id, achievement_id)
                )
            """)

            # Get subscribed staff
            subscribed = await conn.fetch("""
                SELECT DISTINCT ps.staff_id, m.staff_name
                FROM kpi.push_subscriptions ps
                JOIN kpi.staff_list_master m ON ps.staff_id = m.staff_id
            """)

            # Get already-achieved for all staff
            achieved = await conn.fetch("""
                SELECT staff_id, achievement_id FROM kpi.first_time_achievements
            """)
            achieved_set = {(r['staff_id'], r['achievement_id']) for r in achieved}

            for staff in subscribed:
                staff_id = staff['staff_id']
                staff_name = staff['staff_name']

                # Check First RM100+ Commission Sale (single transaction with commission >= 100)
                if (staff_id, 'first_100') not in achieved_set:
                    has_big_sale = await conn.fetchval("""
                        SELECT 1 FROM "AcCSD" d
                        JOIN "AcCSM" m ON d."DocumentNo" = m."DocumentNo"
                        JOIN "AcStockCompany" sc ON d."AcStockID" = sc."AcStockID"
                            AND d."AcStockUOMID" = sc."AcStockUOMID"
                        WHERE d."AcSalesmanID" = $1
                          AND d."ItemTotal" >= 100
                          AND sc."AcStockUDGroup1ID" IN ('FLTHB', 'FLTF1', 'FLTF2', 'FLTF3', 'FLTSC')
                        LIMIT 1
                    """, staff_id)

                    if has_big_sale:
                        await notify_first_time(conn, staff_id, staff_name, 'first_100',
                            "Big Fish!", "First RM100+ commission sale! Big money energy!",
                            notified)

                # Check First Top 10
                if (staff_id, 'first_top10') not in achieved_set:
                    in_top10 = await conn.fetchval("""
                        SELECT 1 FROM analytics.mv_staff_rankings
                        WHERE staff_id = $1 AND company_rank_sales <= 10
                    """, staff_id)

                    if in_top10:
                        await notify_first_time(conn, staff_id, staff_name, 'first_top10',
                            "Elite Status!", FIRST_TIME_ACHIEVEMENTS.get("first_rank_top10", "TOP 10 debut!"),
                            notified)

                # Check First Target (any month where they hit 100%)
                if (staff_id, 'first_target') not in achieved_set:
                    hit_target = await conn.fetchval("""
                        WITH monthly AS (
                            SELECT
                                EXTRACT(YEAR FROM sale_date) as yr,
                                EXTRACT(MONTH FROM sale_date) as mo,
                                SUM(total_sales) as total
                            FROM analytics.mv_staff_daily_kpi
                            WHERE staff_id = $1
                            GROUP BY 1, 2
                        )
                        SELECT 1 FROM monthly m
                        JOIN "KPITargets" t ON
                            t.salesman_id = $1 AND
                            t.year_month = (m.yr * 100 + m.mo)::integer
                        WHERE m.total >= t.total_sales_target AND t.total_sales_target > 0
                        LIMIT 1
                    """, staff_id)

                    if hit_target:
                        await notify_first_time(conn, staff_id, staff_name, 'first_target',
                            "Target Achieved!", FIRST_TIME_ACHIEVEMENTS.get("first_target", "First target achieved!"),
                            notified)

            return {
                "success": True,
                "checked": len(subscribed),
                "notified": len(notified),
                "details": notified
            }

    except Exception as e:
        print(f"First-time achievements error: {e}")
        import traceback
        traceback.print_exc()
        return {"success": False, "error": str(e)}


async def notify_first_time(conn, staff_id: str, staff_name: str, achievement_id: str,
                           title: str, message: str, notified_list: list):
    """Helper to send first-time achievement notification and track it."""
    full_title = f"🏆 {title}"

    # Save to notifications table
    try:
        await conn.execute("""
            INSERT INTO kpi.notifications (staff_id, title, message, type, data)
            VALUES ($1, $2, $3, 'first_time', $4)
        """, staff_id, full_title, message, json.dumps({
            "achievement_id": achievement_id
        }))
    except Exception as e:
        print(f"Failed to save first-time notification: {e}")

    # Send push
    result = await send_push_to_staff(
        staff_id=staff_id,
        title=full_title,
        message=message,
        data={"type": "first_time", "achievement_id": achievement_id}
    )

    # Mark as achieved
    await conn.execute("""
        INSERT INTO kpi.first_time_achievements (staff_id, achievement_id)
        VALUES ($1, $2) ON CONFLICT DO NOTHING
    """, staff_id, achievement_id)

    notified_list.append({
        "staff_id": staff_id,
        "staff_name": staff_name,
        "achievement_id": achievement_id,
        "title": title,
        "push_sent": result.get("success", False)
    })


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
